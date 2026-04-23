import sys
import os
# Qt WebEngine üzerinde Windows GPU/VSync çakışmaları için disable GPU
os.environ["QTWEBENGINE_DISABLE_GPU"] = "1"
os.environ["QT_OPENGL"] = "software"
os.environ["QT_DEBUG_PLUGINS"] = "0"
os.environ["QTWEBENGINE_CHROMIUM_FLAGS"] = "--disable-gpu --disable-software-rasterizer --log-level=3 --ignore-gpu-blocklist"
sys.argv.extend(["--disable-gpu", "--disable-software-rasterizer", "--disable-gpu-compositing"])
import shutil
import json
import random
import unicodedata
import urllib.request
import urllib.error
import subprocess
import tempfile
import webbrowser
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
import pandas as pd
import uuid

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QFormLayout, QLineEdit, QComboBox, QDateEdit, QDoubleSpinBox,
    QPushButton, QLabel, QToolButton, QSizePolicy, QTableWidget, QTableWidgetItem, QTabWidget,
    QMessageBox, QHeaderView, QGroupBox, QFileDialog, QGridLayout, QMenu,
    QStackedWidget, QFrame, QDialog, QDialogButtonBox, QListWidget,
    QListWidgetItem, QAbstractItemView, QScrollArea, QCheckBox
)
from PyQt6.QtCore import Qt, QDate, QTimer, QPropertyAnimation, QVariantAnimation, QEasingCurve, pyqtProperty, QUrl, QSize
from PyQt6.QtGui import QFont, QColor, QAction, QPainter, QLinearGradient, QFontDatabase, QMovie, QIcon
from PyQt6.QtWidgets import QProgressBar
from PyQt6.QtMultimedia import QMediaPlayer, QAudioOutput
from PyQt6.QtMultimediaWidgets import QVideoWidget
try:
    from PyQt6.QtWebEngineWidgets import QWebEngineView
    from PyQt6.QtWebEngineCore import QWebEngineSettings
    import_webengine_success = True
except ImportError:
    QWebEngineView = None
    import_webengine_success = False

try:
    from PyQt6.QtLottie import QLottieWidget  # type: ignore
except ImportError:
    QLottieWidget = None

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

try:
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas  # type: ignore[import-not-found]
    from matplotlib.figure import Figure  # type: ignore[import-not-found]
    MATPLOTLIB_AVAILABLE = True
except Exception:
    FigureCanvas = None
    Figure = None
    MATPLOTLIB_AVAILABLE = False

if MATPLOTLIB_AVAILABLE:
    class NonBlockingFigureCanvas(FigureCanvas):
        def wheelEvent(self, event):
            # Do not consume wheel event so parent scroll can continue.
            event.ignore()

APP_VERSION = "1.0.0"
APP_NAME = "GelirGiderApp"
APP_EXE_NAME = "fatura_masaustu.exe"


def get_source_root() -> Path:
    return Path(os.path.dirname(os.path.abspath(__file__)))


def get_bundle_root() -> Path:
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS)
    return get_source_root()


def get_install_root() -> Path:
    return Path(os.environ.get("LOCALAPPDATA", str(Path.home()))) / APP_NAME


def sync_runtime_data(bundle_data_dir: Path, app_data_dir: Path) -> None:
    if not bundle_data_dir.exists():
        return

    app_data_dir.mkdir(parents=True, exist_ok=True)
    preserve_dirs = {"evraklar"}
    replace_dirs = {"anim", "dil", "font"}
    preserve_files = {"app_ayarlar.json", "fatura_kayitlari.json"}

    for item in bundle_data_dir.iterdir():
        target = app_data_dir / item.name
        if item.is_dir():
            if item.name in replace_dirs:
                if target.exists():
                    shutil.rmtree(target, ignore_errors=True)
                shutil.copytree(item, target)
            elif item.name in preserve_dirs:
                if not target.exists():
                    shutil.copytree(item, target)
            elif not target.exists():
                shutil.copytree(item, target)
        else:
            if item.name in preserve_files:
                if not target.exists():
                    shutil.copy2(item, target)
            else:
                shutil.copy2(item, target)


def bootstrap_runtime_once() -> None:
    if not getattr(sys, "frozen", False):
        return

    current_exe = Path(sys.executable)
    install_root = get_install_root()
    install_root.mkdir(parents=True, exist_ok=True)
    target_exe = install_root / APP_EXE_NAME
    bundle_data_dir = get_bundle_root() / "data"
    app_data_dir = install_root / "data"

    if current_exe.parent != install_root:
        shutil.copy2(current_exe, target_exe)
        sync_runtime_data(bundle_data_dir, app_data_dir)
        subprocess.Popen([str(target_exe)], cwd=str(install_root))
        sys.exit(0)

    sync_runtime_data(bundle_data_dir, app_data_dir)


bootstrap_runtime_once()

APP_ROOT_DIR = get_install_root() if getattr(sys, "frozen", False) else get_source_root()
APP_DATA_DIR = APP_ROOT_DIR / "data"
APP_ANIM_DIR = APP_DATA_DIR / "anim"
APP_LANG_DIR = APP_DATA_DIR / "dil"
APP_FONT_DIR = APP_DATA_DIR / "font"
APP_EVRAK_DIR = APP_DATA_DIR / "evraklar"


def app_data_path(*parts: str) -> str:
    return str(APP_DATA_DIR.joinpath(*parts))

KDV_OPTIONS = [0, 1, 8, 10, 18, 20]
TEVKIFAT_OPTIONS = [
    ("yok_tevkifat_key", 0),  # Will translate on the fly
    ("2/10", 2),
    ("3/10", 3),
    ("4/10", 4),
    ("5/10", 5),
    ("7/10", 7),
    ("9/10", 9),
    ("Tam (10/10)", 10)
]

# --- APP SETTINGS & LOCALIZATION ---
SETTINGS_FILE = app_data_path("app_ayarlar.json")
app_config = {
    "lang": "tr",
    "theme": "dark",
    "update_enabled": True,
    "update_repo": "umtark/Yonetim_Paneli_Genel",
    "update_asset_keyword": "fatura",
    "update_interval_hours": 12,
    "update_last_check": "",
}

if os.path.exists(SETTINGS_FILE):
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            app_config.update(json.load(f))
    except Exception:
        pass

def save_app_settings():
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(app_config, f, indent=4)

locale_data = {}
try:
    with open(str(APP_LANG_DIR / f"{app_config['lang']}.json"), "r", encoding="utf-8") as f:
        locale_data = json.load(f)
except Exception:
    pass

def _t(key, default=None):
    return locale_data.get(key, default if default is not None else key)


def get_custom_font(family, size, weight=QFont.Weight.Normal):
    f = QFont(family, size, weight)
    f.setStyleStrategy(QFont.StyleStrategy.PreferAntialias)
    f.setHintingPreference(QFont.HintingPreference.PreferFullHinting)
    return f

def register_application_fonts():
    for font_filename in [
        "Quicksand-Light.ttf",
        "Quicksand-Regular.ttf",
        "Quicksand-Medium.ttf",
        "Quicksand-SemiBold.ttf",
        "Quicksand-Bold.ttf",
    ]:
        font_path = APP_FONT_DIR / font_filename
        if font_path.exists():
            QFontDatabase.addApplicationFont(str(font_path))


THEME_COLOR_PALETTES = {
    "shared": {
        "white": "#ffffff",
        "primary_accent": "#7c3aed",
        "primary_hover": "#8b5cf6",
        "primary_active": "#6d28d9",
        "success_bg": "#059669",
        "success_hover": "#10b981",
        "info_bg": "#2563eb",
        "info_hover": "#1d4ed8",
        "info_active": "#1e40af",
        "danger_bg": "#dc2626",
        "danger_hover": "#ef4444",
        "warning_bg": "#f59e0b",
        "warning_hover": "#d97706",
        "indigo_bg": "#6366f1",
        "indigo_hover": "#4f46e5",
        "rose_bg": "#e11d48",
        "emerald_bg": "#16a34a",
        "muted_hint": "#64748b",
        "muted_border": "#94a3b8",
        "muted_panel": "#544545",
        "ticker_dark_bg": "rgba(40, 40, 40, 0.8)",
        "ticker_dark_border": "#4b5563",
        "ticker_dark_text": "#fbbf24",
        "ticker_light_border": "#fcd34d",
        "ticker_light_text": "#b45309",
        "positive_text": "#10b981",
        "positive_soft_text": "#4ade80",
        "negative_text": "#ef4444",
        "negative_soft_text": "#f87171",
        "warning_text_light": "#92400e",
        "critical_text_light": "#991b1b",
        "warning_text_dark": "#f59e0b",
        "critical_text_dark": "#ef4444",
        "overlay_soft": "rgba(255,255,255,0.1)",
        "pulse_base": "#94a3b8",
        "selection_hint": "rgba(240, 240, 240, 50)",
        "report_muted": "#64748b",
        "report_header_teal": "#0f766e",
        "report_grid": "#94a3b8",
        "report_row_bg": "#f8fafc",
    },
    "light": {
        "main_bg": "#f8fafc",
        "surface_bg": "#ffffff",
        "surface_alt": "#f1f5f9",
        "surface_soft": "#eef2ff",
        "text": "#1e293b",
        "title_text": "#0f172a",
        "subtitle_text": "#475569",
        "group_border": "#cbd5e1",
        "group_title": "#4338ca",
        "input_bg": "#ffffff",
        "input_border": "#cbd5e1",
        "focus_border": "#8b5cf6",
        "table_bg": "#ffffff",
        "gridline": "#e2e8f0",
        "button_bg": "#2563eb",
        "button_hover": "#1d4ed8",
        "button_pressed": "#1e40af",
        "button_border": "#1d4ed8",
        "button_text": "#ffffff",
        "tab_bg": "#f1f5f9",
        "tab_text": "#475569",
        "tab_selected_bg": "#8b5cf6",
        "tab_selected_text": "#ffffff",
        "message_bg": "#ffffff",
        "message_border": "#cbd5e1",
        "menu_bg": "#ffffff",
        "menu_text": "#1e293b",
        "menu_selected_bg": "#8b5cf6",
        "menu_selected_text": "#ffffff",
        "sidebar_bg": "#ffffff",
        "sidebar_border": "#cbd5e1",
        "sidebar_logo": "#4338ca",
        "sidebar_text": "#1e293b",
        "sidebar_hover_bg": "#e2e8f0",
        "sidebar_hover_text": "#0f172a",
        "sidebar_checked_bg": "#1d4ed8",
        "sidebar_checked_text": "#0f172a",
        "sidebar_checked_border": "#93c5fd",
        "sidebar_checked_hover_border": "#bfdbfe",
        "hamburger_color": "#7c3aed",
        "hamburger_hover_text": "#4338ca",
        "hamburger_hover_bg": "#D000FF2A",
        "license_dialog_bg": "#ffffff",
        "license_info_bg": "#f8fafc",
        "license_info_text": "#0f172a",
        "license_error": "#b91c1c",
        "invoice_group_bg": "#ffffff",
        "invoice_title": "#1d4ed8",
        "invoice_label": "#334155",
        "vehicle_group_bg": "#ffffff",
        "vehicle_group_border": "#94a3b8",
        "vehicle_group_title": "#1e40af",
        "vehicle_label": "#334155",
        "settings_title": "#8b5cf6",
        "subtle_text": "#4b5563",
        "dashboard_page_bg": "qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #f8fafc, stop:1 #eef2ff)",
        "dashboard_info_bg": "#e0f2fe",
        "dashboard_info_border": "#7dd3fc",
        "dashboard_info_text": "#0c4a6e",
        "dashboard_kpi_title": "#1e3a8a",
        "dashboard_kpi_value": "#0f172a",
        "dashboard_badge_bg": "#ffffff",
        "dashboard_badge_border": "#cbd5e1",
        "dashboard_badge_text": "#A0177E",
        "dashboard_alerts_border": "#2563eb",
        "dashboard_alerts_title": "#1e40af",
        "dashboard_alerts_header_bg": "#dbeafe",
        "dashboard_alerts_header_text": "#1e3a8a",
        "dashboard_alerts_panel_bg": "#eff6ff",
        "dashboard_events_border": "#7dbad0",
        "dashboard_events_title": "#1e40af",
        "dashboard_events_header_bg": "#bfdbfe",
        "dashboard_events_header_text": "#1e3a8a",
        "dashboard_selected_bg": "#dbeafe",
        "dashboard_selected_text": "#8e136b",
        "dashboard_card_blue_bg": "#e0ecff",
        "dashboard_card_red_bg": "#fee2e2",
        "dashboard_card_green_bg": "#dcfce7",
        "dashboard_card_amber_bg": "#fef3c7",
        "dashboard_card_gradient_end": "#544545",
    },
    "dark": {
        "main_bg": "#1a1625",
        "surface_bg": "#231f30",
        "surface_alt": "#2d273f",
        "surface_soft": "#201934",
        "text": "#e2e8f0",
        "title_text": "#f5f3ff",
        "subtitle_text": "#c4b5fd",
        "group_border": "#4a3b69",
        "group_title": "#c4b5fd",
        "input_bg": "#231f30",
        "input_border": "#4a3b69",
        "focus_border": "#8b5cf6",
        "table_bg": "#231f30",
        "gridline": "#3b3253",
        "button_bg": "#4c3a73",
        "button_hover": "#5b478a",
        "button_pressed": "#3a2b5c",
        "button_border": "#5b478a",
        "button_text": "#ffffff",
        "tab_bg": "#231f30",
        "tab_text": "#e2e8f0",
        "tab_selected_bg": "#4c3a73",
        "tab_selected_text": "#ffffff",
        "message_bg": "#1a1625",
        "message_border": "#4a3b69",
        "menu_bg": "#231f30",
        "menu_text": "#e2e8f0",
        "menu_selected_bg": "#4c3a73",
        "menu_selected_text": "#ffffff",
        "sidebar_bg": "#231f30",
        "sidebar_border": "#4a3b69",
        "sidebar_logo": "#c4b5fd",
        "sidebar_text": "#e2e8f0",
        "sidebar_hover_bg": "#4a3b69",
        "sidebar_hover_text": "#ffffff",
        "sidebar_checked_bg": "#7c3aed",
        "sidebar_checked_text": "#ffffff",
        "sidebar_checked_border": "#c4b5fd",
        "sidebar_checked_hover_border": "#ddd6fe",
        "hamburger_color": "#a78bfa",
        "hamburger_hover_text": "#ffffff",
        "hamburger_hover_bg": "rgba(255,255,255,0.1)",
        "license_dialog_bg": "#1a1625",
        "license_info_bg": "#231f30",
        "license_info_text": "#e2e8f0",
        "license_error": "#fca5a5",
        "invoice_group_bg": "transparent",
        "invoice_title": "#c4b5fd",
        "invoice_label": "#e2e8f0",
        "vehicle_group_bg": "rgba(35,31,48,0.88)",
        "vehicle_group_border": "#4a3b69",
        "vehicle_group_title": "#c4b5fd",
        "vehicle_label": "#cbd5e1",
        "settings_title": "#8b5cf6",
        "subtle_text": "#cbd5e1",
        "dashboard_page_bg": "qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #171327, stop:1 #201934)",
        "dashboard_info_bg": "rgba(124, 58, 237, 0.20)",
        "dashboard_info_border": "rgba(167, 139, 250, 0.45)",
        "dashboard_info_text": "#ddd6fe",
        "dashboard_kpi_title": "#c4b5fd",
        "dashboard_kpi_value": "#ffffff",
        "dashboard_badge_bg": "rgba(15, 23, 42, 0.50)",
        "dashboard_badge_border": "rgba(148, 163, 184, 0.50)",
        "dashboard_badge_text": "#e2e8f0",
        "dashboard_alerts_border": "#dc2626",
        "dashboard_alerts_title": "#ef4444",
        "dashboard_alerts_header_bg": "#7f1d1d",
        "dashboard_alerts_header_text": "#ffffff",
        "dashboard_alerts_panel_bg": "transparent",
        "dashboard_events_border": "#2563eb",
        "dashboard_events_title": "#3b82f6",
        "dashboard_events_header_bg": "#1e3a8a",
        "dashboard_events_header_text": "#ffffff",
        "dashboard_selected_bg": "#1e3a8a",
        "dashboard_selected_text": "#ffffff",
        "dashboard_card_blue_bg": "rgba(30,58,138,0.35)",
        "dashboard_card_red_bg": "rgba(127,29,29,0.35)",
        "dashboard_card_green_bg": "rgba(6,78,59,0.35)",
        "dashboard_card_amber_bg": "rgba(120,53,15,0.35)",
        "dashboard_card_gradient_end": "rgba(35,31,48,0.88)",
    },
}


def get_theme_palette(theme="dark"):
    palette = dict(THEME_COLOR_PALETTES["shared"])
    palette.update(THEME_COLOR_PALETTES["light" if theme == "light" else "dark"])
    return palette


def build_button_style(background_color, text_color, hover_color=None, padding="8px 16px", border_radius=6, font_size=None, border_color=None, extra=""):
    hover_color = hover_color or background_color
    font_size_rule = f"font-size: {font_size}; " if font_size else ""
    border_rule = f"border: 1px solid {border_color}; " if border_color else "border: none; "
    extra_rule = f"{extra.strip()} " if extra else ""
    return (
        f"QPushButton {{ background-color: {background_color}; color: {text_color}; padding: {padding}; {font_size_rule}font-weight: bold; border-radius: {border_radius}px; {border_rule}{extra_rule}}} "
        f"QPushButton:hover {{ background-color: {hover_color}; }}"
    )


def get_stylesheet(theme="dark"):
    palette = get_theme_palette(theme)
    return """
        QMainWindow {{ background-color: {main_bg}; }}
        QWidget {{ background-color: {main_bg}; color: {text}; font-family: Quicksand; font-weight: 500; }}
        QGroupBox {{ border: 1px solid {group_border}; border-radius: 6px; margin-top: 10px; padding-top: 15px; font-weight: bold; background-color: transparent; }}
        QGroupBox::title {{ subcontrol-origin: margin; subcontrol-position: top center; padding: 0 5px; color: {group_title}; background-color: {main_bg}; }}
        QLineEdit, QComboBox, QDateEdit, QDoubleSpinBox {{ background-color: {input_bg}; border: 1px solid {input_border}; border-radius: 4px; padding: 5px; color: {text}; }}
        QLineEdit:focus, QComboBox:focus, QDateEdit:focus, QDoubleSpinBox:focus {{ border: 1px solid {focus_border}; }}
        QTableWidget {{ background-color: {table_bg}; gridline-color: {gridline}; color: {text}; border: 1px solid {input_border}; }}
        QHeaderView::section {{ background-color: {surface_alt}; color: {group_title}; padding: 4px; border: 1px solid {input_border}; font-weight: bold; }}
        QPushButton {{ background-color: {button_bg}; color: {button_text}; border: 1px solid {button_border}; border-radius: 6px; padding: 6px; font-weight: bold; }}
        QPushButton:hover {{ background-color: {button_hover}; }}
        QPushButton:pressed {{ background-color: {button_pressed}; }}
        QTabWidget::pane {{ border: 1px solid {group_border}; }}
        QTabBar::tab {{ background-color: {tab_bg}; color: {tab_text}; padding: 8px 16px; border: 1px solid {group_border}; border-bottom: none; border-top-left-radius: 4px; border-top-right-radius: 4px; }}
        QTabBar::tab:selected {{ background-color: {tab_selected_bg}; color: {tab_selected_text}; font-weight: bold; }}
        QMessageBox {{ background-color: {message_bg}; border: 1px solid {message_border}; color: {text}; }}
        QMessageBox QLabel {{ background-color: transparent; color: {text}; font-size: 14px; padding: 5px; }}
        QMessageBox QPushButton {{ background-color: {button_bg}; color: {button_text}; padding: 8px 16px; font-weight: bold; border-radius: 6px; min-width: 60px; }}
        QMessageBox QPushButton:hover {{ background-color: {button_hover}; }}
        QMenu {{ background-color: {menu_bg}; color: {menu_text}; border: 1px solid {group_border}; }}
        QMenu::item {{ padding: 6px 20px; }}
        QMenu::item:selected {{ background-color: {menu_selected_bg}; color: {menu_selected_text}; }}
        QLabel {{ background-color: transparent; }}

        QFrame#SidebarFrame {{ background-color: {sidebar_bg}; border-right: 1px solid {sidebar_border}; }}
        QLabel#LogoLabel {{ color: {sidebar_logo}; text-align: center; margin-bottom: 20px; font-size: 18px; font-weight: bold; }}
        QToolButton#SidebarButton {{ background-color: transparent; color: {sidebar_text}; border: 2px solid transparent; text-align: center; padding-top: 10px; padding-bottom: 10px; font-size: 16px; border-radius: 6px; }}
        QToolButton#SidebarButton:hover {{ background-color: {sidebar_hover_bg}; color: {sidebar_hover_text}; }}
        QToolButton#SidebarButton:checked {{ background-color: {sidebar_checked_bg}; color: {sidebar_checked_text}; font-weight: bold; border: 2px solid {sidebar_checked_border}; padding-left: 8px; padding-right: 8px; }}
        QToolButton#SidebarButton:checked:hover {{ background-color: {sidebar_checked_bg}; color: {sidebar_checked_text}; border: 2px solid {sidebar_checked_hover_border}; }}
    """.format(**palette)


def get_dashboard_theme(theme="dark"):
    palette = get_theme_palette(theme)
    return {
        "page_bg": palette["dashboard_page_bg"],
        "title": palette["title_text"],
        "subtitle": palette["subtitle_text"],
        "info_bg": palette["dashboard_info_bg"],
        "info_border": palette["dashboard_info_border"],
        "info_text": palette["dashboard_info_text"],
        "kpi_title": palette["dashboard_kpi_title"],
        "kpi_value": palette["dashboard_kpi_value"],
        "badge_bg": palette["dashboard_badge_bg"],
        "badge_border": palette["dashboard_badge_border"],
        "badge_text": palette["dashboard_badge_text"],
        "alerts_border": palette["dashboard_alerts_border"],
        "alerts_title": palette["dashboard_alerts_title"],
        "alerts_header_bg": palette["dashboard_alerts_header_bg"],
        "alerts_header_text": palette["dashboard_alerts_header_text"],
        "alerts_panel_bg": palette["dashboard_alerts_panel_bg"],
        "events_border": palette["dashboard_events_border"],
        "events_title": palette["dashboard_events_title"],
        "events_header_bg": palette["dashboard_events_header_bg"],
        "events_header_text": palette["dashboard_events_header_text"],
        "table_selected_bg": palette["dashboard_selected_bg"],
        "table_selected_text": palette["dashboard_selected_text"],
        "card_blue_bg": palette["dashboard_card_blue_bg"],
        "card_red_bg": palette["dashboard_card_red_bg"],
        "card_green_bg": palette["dashboard_card_green_bg"],
        "card_amber_bg": palette["dashboard_card_amber_bg"],
        "card_gradient_end": palette["dashboard_card_gradient_end"],
        "warning_text": palette["warning_text_light" if theme == "light" else "warning_text_dark"],
        "critical_text": palette["critical_text_light" if theme == "light" else "critical_text_dark"],
        "positive_text": palette["positive_text"],
    }


def get_license_dialog_theme(theme="dark"):
    palette = get_theme_palette(theme)
    return {
        "dialog": f"font-family: Quicksand; font-size: 14px; background-color: {palette['license_dialog_bg']}; color: {palette['text']};",
        "info_frame": f"background-color: {palette['license_info_bg']}; border: 1px solid {palette['input_border']}; border-radius: 10px; padding: 14px;",
        "info_text": f"font-size: 13px; color: {palette['license_info_text']};",
        "input": f"font-size: 15px; font-weight: bold; padding: 8px; border: 1px solid {palette['input_border']}; border-radius: 6px; background-color: {palette['input_bg']}; color: {palette['title_text'] if theme == 'light' else palette['white']};",
        "button": f"background-color: {palette['info_bg']}; color: {palette['white']}; padding: 10px; font-weight: bold; border-radius: 6px;",
        "error": f"color: {palette['license_error']}; font-size: 12px;",
    }


def get_invoice_form_styles(theme="dark"):
    palette = get_theme_palette(theme)
    return f"""
        QGroupBox {{ font-weight: bold; border: 1px solid {palette['group_border']}; border-radius: 8px; margin-top: 15px; background-color: {palette['invoice_group_bg']}; }}
        QGroupBox::title {{ subcontrol-origin: margin; subcontrol-position: top left; padding: 0 10px; color: {palette['invoice_title']}; background-color: {palette['main_bg']}; }}
        QLabel {{ font-weight: bold; background-color: transparent; color: {palette['invoice_label']}; }}
        QLineEdit, QComboBox, QDateEdit {{ padding: 6px; border: 1px solid {palette['input_border']}; border-radius: 4px; background-color: {palette['input_bg']}; color: {palette['text'] if theme == 'light' else palette['white']}; }}
    """


def get_vehicle_summary_styles(theme="dark"):
    palette = get_theme_palette(theme)
    return {
        "group": f"QGroupBox {{ border: 1px solid {palette['vehicle_group_border']}; border-radius: 8px; margin-top: 12px; background-color: {palette['vehicle_group_bg']}; }} QGroupBox::title {{ subcontrol-origin: margin; subcontrol-position: top center; padding: 0 10px; color: {palette['vehicle_group_title']}; }}",
        "label": f"QLabel {{ background-color: transparent; border: none; color: {palette['vehicle_label']}; }}",
    }
# -----------------------------------
# (QLineEdit remvoed to favor native PyQt translators)

def hesapla_tutarlar(matrah: float, kdv_orani: float, tevkifat_payi: int = 0) -> tuple[float, float, float, float]:
    kdv_tutari = matrah * (kdv_orani / 100)
    tevkifat_tutari = kdv_tutari * (tevkifat_payi / 10)
    odenecek_kdv = kdv_tutari - tevkifat_tutari
    toplam = matrah + odenecek_kdv
    return round(kdv_tutari, 2), round(tevkifat_tutari, 2), round(odenecek_kdv, 2), round(toplam, 2)

def records_to_df(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(
            columns=[_t("tarih_hdr", "Tarih"), _t("tip_hdr", "Tip"), _t("fatura_no_hdr", "Fatura No"), _t("firma_hdr", "Firma"), _t("aciklama_hdr", "Aciklama"), _t("matrah_hdr", "Matrah"), _t("kdv_yuzde_hdr", "KDV %"), _t("tevkifat_hdr", "Tevkifat"), _t("tev_tutari_hdr", "Tev.Tutarı"), _t("toplam_hdr", "Toplam")]
        )
    return pd.DataFrame(records)

def format_number(value: float) -> str:
    if pd.isnull(value): return "0,00"
    return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def format_currency(value: float) -> str:
    return format_number(value) + " TL"

def hesapla_ozet(df: pd.DataFrame) -> dict:
    if df.empty:
        return {
            "giden_toplam": 0.0, "gelen_toplam": 0.0,
            "giden_kdv": 0.0, "gelen_kdv": 0.0,
            "net_fark": 0.0, "kdv_fark": 0.0,
        }

    giden = df[df["Tip"] == "Giden"]
    gelen = df[df["Tip"] == "Gelen"]

    giden_toplam = float(giden["Toplam"].sum())
    gelen_toplam = float(gelen["Toplam"].sum())
    giden_kdv = float(giden["KDV Tutari"].sum()) if not df.empty and "KDV Tutari" in df else 0.0
    gelen_kdv = float(gelen["KDV Tutari"].sum()) if not df.empty and "KDV Tutari" in df else 0.0
    giden_tevkifat = float(giden["Tev.Tutarı"].sum()) if not df.empty and "Tev.Tutarı" in df else 0.0
    gelen_tevkifat = float(gelen["Tev.Tutarı"].sum()) if not df.empty and "Tev.Tutarı" in df else 0.0

    return {
        "giden_toplam": round(giden_toplam, 2), "gelen_toplam": round(gelen_toplam, 2),
        "giden_kdv": round(giden_kdv, 2), "gelen_kdv": round(gelen_kdv, 2),
        "giden_tevkifat": round(giden_tevkifat, 2), "gelen_tevkifat": round(gelen_tevkifat, 2),
        "net_fark": round(giden_toplam - gelen_toplam, 2), "kdv_fark": round(giden_kdv - gelen_kdv, 2),
    }

def build_pdf(df: pd.DataFrame, ozet: dict) -> bytes:
    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=landscape(A4), rightMargin=14*mm, leftMargin=14*mm, topMargin=14*mm, bottomMargin=14*mm
    )

    pdfmetrics.registerFont(TTFont("Quicksand", str(APP_FONT_DIR / "Quicksand-Regular.ttf")))
    pdfmetrics.registerFont(TTFont("Quicksand-Bold", str(APP_FONT_DIR / "Quicksand-Bold.ttf")))
    report_palette = get_theme_palette("light")

    def add_pdf_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Quicksand", 8)
        canvas.setFillColor(colors.HexColor(report_palette["report_muted"]))
        current_year = datetime.now().year
        year_str = f"2026 - {current_year}"
        footer_text = f"© {_t('owner_name', 'Ümit Arik')} {year_str}"
        
        canvas.drawString(14*mm, 8*mm, _t("pdf_financial_report", "Financial Analysis Report"))
        canvas.drawCentredString(297*mm / 2.0, 8*mm, footer_text)
        canvas.drawRightString(297*mm - 14*mm, 8*mm, f"{_t('pdf_page', 'Page')} {doc.page}")
        canvas.restoreState()

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="QuicksandTitle", fontName="Quicksand-Bold", fontSize=18, spaceAfter=10))
    styles.add(ParagraphStyle(name="QuicksandNormal", fontName="Quicksand", fontSize=10, spaceAfter=6))
    styles.add(ParagraphStyle(name="QuicksandHeading", fontName="Quicksand-Bold", fontSize=14, spaceAfter=8))
    
    story = []

    title = Paragraph("Gelen / Giden Fatura Analizi", styles["QuicksandTitle"])
    story.append(title)
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles["QuicksandNormal"]))
    story.append(Spacer(1, 10))

    summary_data = [
        ["Finansal Durum (Metrik)", "Tutar"],
        ["Toplam Kestiğiniz Faturalar (Gelir)", format_currency(ozet["giden_toplam"])],
        ["Toplam Aldığınız Faturalar (Gider)", format_currency(ozet["gelen_toplam"])],
        ["Net Kâr / Zarar Durumu", format_currency(ozet["net_fark"])],
        ["Net KDV (Devlete Ödenecek KDV / Devreden)", format_currency(ozet["kdv_fark"])],
    ]

    summary_table = Table(summary_data, colWidths=[110*mm, 55*mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(report_palette["report_header_teal"])),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(report_palette["report_grid"])),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("FONTNAME", (0, 0), (-1, -1), "Quicksand"),
        ("FONTNAME", (0, 0), (-1, 0), "Quicksand-Bold"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor(report_palette["report_row_bg"])),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 12))

    if not df.empty:
        report_df = df.copy()
        
        # Kağıttan taşmayı önlemek için PDF listesinde göstermeye gerek olmayan sütunları çıkaralım
        cols_to_drop = ["id", "Aciklama", "Tevkifat"]
        report_df = report_df.drop(columns=[col for col in cols_to_drop if col in report_df.columns])

        report_df["Matrah"] = report_df["Matrah"].map(lambda x: format_number(x))
        if "KDV Tutari" in report_df:
            report_df["KDV Tutari"] = report_df["KDV Tutari"].map(lambda x: format_number(x))
        if "Tev.Tutarı" in report_df:
            report_df["Tev.Tutarı"] = report_df["Tev.Tutarı"].map(lambda x: format_number(x) if pd.notnull(x) else "0,00")
        report_df["Toplam"] = report_df["Toplam"].map(lambda x: format_number(x))

        table_data = [report_df.columns.tolist()] + report_df.values.tolist()
        available_width = 269 * mm
        num_cols = len(report_df.columns)
        col_widths = [available_width / num_cols] * num_cols if num_cols > 0 else None
        
        invoice_table = Table(
            table_data, colWidths=col_widths, repeatRows=1
        )
        invoice_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(report_palette["info_hover"])),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Quicksand"),
            ("FONTNAME", (0, 0), (-1, 0), "Quicksand-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),  # Biraz daha büyük ve okunaklı
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(report_palette["group_border"])),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"), # Tüm metinleri ortala
            ("ALIGN", (5, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ]))
        story.append(Paragraph("Fatura Kayıt Detayları", styles["QuicksandHeading"]))
        story.append(Spacer(1, 6))
        story.append(invoice_table)
    else:
        story.append(Paragraph("Listede fatura kaydı bulunmamaktadır.", styles["QuicksandNormal"]))

    doc.build(story, onFirstPage=add_pdf_footer, onLaterPages=add_pdf_footer)
    return output.getvalue()


import time
from PyQt6.QtGui import QPixmap

class TickerWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedHeight(35)
        self.text = _t("ticker_bekleniyor", "Tüm sistem kayıtları kontrol ediliyor... Lütfen bekleyiniz.")
        self.x_pos = 1000.0
        self.speed = 40.0
        self.last_time = time.time()
        self.cached_pixmap = None

        self.timer = QTimer(self)
        self.timer.setTimerType(Qt.TimerType.PreciseTimer)
        self.timer.timeout.connect(self.update_tick)
        self.timer.start(16) 
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.apply_theme()

    def apply_theme(self):
        theme = app_config.get("theme", "dark")
        palette = get_theme_palette(theme)
        if theme == "dark":
            self.bg_color = palette["ticker_dark_bg"]
            self.border_color = palette["ticker_dark_border"]
            self.text_color = palette["ticker_dark_text"]
        else:
            self.bg_color = palette["muted_panel"]
            self.border_color = palette["ticker_light_border"]
            self.text_color = palette["ticker_light_text"]

        self.setStyleSheet(f"""
            QWidget {{
                background-color: {self.bg_color};
                border: 1px solid {self.border_color};
                border-radius: 6px;
            }}
        """)
        self.cached_pixmap = None
        self.update()

    def set_text(self, text):
        if self.text != text:
            self.text = text
            self.cached_pixmap = None
            self.x_pos = float(self.width())
            self.update()

    def update_tick(self):
        now = time.time()
        dt = now - self.last_time
        if dt <= 0: return
        self.last_time = now
        
        self.x_pos -= self.speed * dt
        
        tw = (self.cached_pixmap.width() / self.devicePixelRatioF()) if self.cached_pixmap else 800.0
        if self.x_pos < -(tw + 50.0):
            self.x_pos = float(self.width())
                
        self.update()

    def rebuild_cache(self):
        from PyQt6.QtGui import QFontMetrics, QPainter, QColor, QPixmap
        from PyQt6.QtCore import Qt, QRectF
        font = get_custom_font("Quicksand", 11, QFont.Weight.Bold)
        fm = QFontMetrics(font)
        text_w = float(fm.horizontalAdvance(self.text))
        
        ratio = self.devicePixelRatioF()
        w = max(10.0, text_w + 150.0)
        h = float(self.height())
        
        self.cached_pixmap = QPixmap(int(w * ratio), int(h * ratio))
        self.cached_pixmap.setDevicePixelRatio(ratio)
        self.cached_pixmap.fill(Qt.GlobalColor.transparent)
        
        p = QPainter(self.cached_pixmap)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        p.setRenderHint(QPainter.RenderHint.TextAntialiasing)
        p.setFont(font)
        p.setPen(QColor(self.text_color))
        rect = QRectF(0.0, 0.0, w, h)
        p.drawText(rect, Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft, self.text)
        p.end()

    def paintEvent(self, event):
        from PyQt6.QtGui import QPainter
        from PyQt6.QtCore import QPointF, QRectF
        if self.cached_pixmap is None:
            self.rebuild_cache()

        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        
        rect = self.cached_pixmap.rect()
        painter.drawPixmap(QPointF(self.x_pos, 0.0), self.cached_pixmap, QRectF(rect))


class AracEkleDialog(QDialog):
    def __init__(self, parent=None, arac_data=None):
        super().__init__(parent)
        palette = get_theme_palette(app_config.get("theme", "dark"))
        self.arac_data = arac_data or {}
        self.setWindowTitle("Araç Ekle / Düzenle" if arac_data else "Yeni Araç Ekle")
        self.resize(750, 480)  # Ekrandan taşmaması için genişletip kısalttık
        self.setStyleSheet(f"""
            QDialog {{ background-color: transparent; font-weight: bold; }}
            QLabel {{ background-color: transparent; }}
            QLineEdit, QComboBox, QDateEdit {{ background-color: {palette['input_bg']}; border: 1px solid {palette['input_border']}; border-radius: 4px; padding: 6px; color: {palette['text'] if app_config.get('theme', 'dark') == 'light' else palette['white']}; }}
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus {{ border: 1px solid {palette['focus_border']}; }}
            QPushButton {{ background-color: {palette['primary_accent']}; color: {palette['white']}; padding: 8px 16px; border-radius: 4px; font-weight: bold; }}
            QPushButton:hover {{ background-color: {palette['primary_hover']}; }}
        """)

        layout = QVBoxLayout(self)

        h_layout = QHBoxLayout()
        
        left_form = QFormLayout()
        left_form.setSpacing(12)
        
        right_form = QFormLayout()
        right_form.setSpacing(12)

        self.in_plaka = QLineEdit(self.arac_data.get("plaka", ""))
        self.in_firma = QLineEdit(self.arac_data.get("firma", ""))
        self.in_kurum = QLineEdit(self.arac_data.get("kurum", ""))
        self.in_vergi_no = QLineEdit(self.arac_data.get("vergi_no", ""))
        self.in_sofor = QLineEdit(self.arac_data.get("sofor", ""))
        self.in_sofor_tel = QLineEdit(self.arac_data.get("sofor_tel", "")) # Yeni Eklenen Şoför Tel
        self.in_marka = QLineEdit(self.arac_data.get("marka", ""))
        self.in_model = QLineEdit(self.arac_data.get("model", ""))
        self.in_yil = QLineEdit(self.arac_data.get("yil", ""))
        
        self.in_yakit = QComboBox()
        self.in_yakit.addItem(_t("yakit_dizel", "Dizel"), "Dizel")
        self.in_yakit.addItem(_t("yakit_benzin", "Benzin"), "Benzin")
        self.in_yakit.addItem(_t("yakit_lpg", "Benzin & LPG"), "Benzin & LPG")
        self.in_yakit.addItem(_t("yakit_elektrik", "Elektrik"), "Elektrik")
        self.in_yakit.addItem(_t("yakit_hibrit", "Hibrit"), "Hibrit")
        self.in_yakit.setCurrentText(self.arac_data.get("yakit", "Dizel"))

        self.in_vites = QComboBox()
        self.in_vites.addItems([_t("vites_manuel", "Manuel"), _t("vites_otomatik", "Otomatik"), _t("vites_yari_otomatik", "Yarı Otomatik")])
        self.in_vites.setCurrentText(self.arac_data.get("vites", "Manuel"))

        self.in_km = QLineEdit(str(self.arac_data.get("km", "")))
        self.in_sasi = QLineEdit(self.arac_data.get("sasi", ""))
        
        self.in_utts = QComboBox()
        self.in_utts.addItems([_t("yok", "Yok"), _t("var", "Var")])
        self.in_utts.setCurrentText(self.arac_data.get("utts", _t("yok", "Yok")))

        self.in_muayene_tarihi = QDateEdit()
        self.in_muayene_tarihi.setCalendarPopup(True)
        if "muayene_tarihi" in self.arac_data:
            self.in_muayene_tarihi.setDate(QDate.fromString(self.arac_data["muayene_tarihi"], "dd.MM.yyyy"))
        else:
            self.in_muayene_tarihi.setDate(QDate.currentDate())

        self.in_trafik_tarihi = QDateEdit()
        self.in_trafik_tarihi.setCalendarPopup(True)
        if "trafik_tarihi" in self.arac_data:
            self.in_trafik_tarihi.setDate(QDate.fromString(self.arac_data["trafik_tarihi"], "dd.MM.yyyy"))
        else:
            self.in_trafik_tarihi.setDate(QDate.currentDate())

        self.in_kasko_tarihi = QDateEdit()
        self.in_kasko_tarihi.setCalendarPopup(True)
        self.kasko_yok_cb = QCheckBox(_t("yok_chk", "Yok"))
        
        kasko_lyt = QHBoxLayout()
        kasko_lyt.addWidget(self.in_kasko_tarihi)
        kasko_lyt.addWidget(self.kasko_yok_cb)
        kasko_lyt.setContentsMargins(0,0,0,0)
        
        def disable_kasko_date():
            self.in_kasko_tarihi.setDisabled(self.kasko_yok_cb.isChecked())
            
        self.kasko_yok_cb.toggled.connect(disable_kasko_date)

        if "kasko_tarihi" in self.arac_data and self.arac_data["kasko_tarihi"] not in ["-", "Yok", ""]:
            self.in_kasko_tarihi.setDate(QDate.fromString(self.arac_data["kasko_tarihi"], "dd.MM.yyyy"))
            self.kasko_yok_cb.setChecked(False)
        else:
            self.in_kasko_tarihi.setDate(QDate.currentDate())
            if self.arac_data: # Mevcut araçta kasko yok demektir
                self.kasko_yok_cb.setChecked(True)
                self.in_kasko_tarihi.setDisabled(True)
            
        self.in_koltuk_sigortasi = QDateEdit()
        self.in_koltuk_sigortasi.setCalendarPopup(True)
        if "koltuk_sigortasi" in self.arac_data and self.arac_data["koltuk_sigortasi"] not in ["Var", "Yok", "-"]:
            self.in_koltuk_sigortasi.setDate(QDate.fromString(self.arac_data["koltuk_sigortasi"], "dd.MM.yyyy"))
        else:
            self.in_koltuk_sigortasi.setDate(QDate.currentDate())

        # Yağ Bakım Tarihi
        self.in_yag_bakim_tarihi = QDateEdit()
        self.in_yag_bakim_tarihi.setCalendarPopup(True)
        if "yag_bakim_tarihi" in self.arac_data and self.arac_data.get("yag_bakim_tarihi") not in ["-", "Yok", ""]:
            self.in_yag_bakim_tarihi.setDate(QDate.fromString(self.arac_data["yag_bakim_tarihi"], "dd.MM.yyyy"))
        else:
            self.in_yag_bakim_tarihi.setDate(QDate.currentDate())
            
        self.yag_bakim_yok_cb = QCheckBox(_t("yok_chk", "Yok / Bilinmiyor"))
        yag_lyt = QHBoxLayout()
        yag_lyt.addWidget(self.in_yag_bakim_tarihi)
        yag_lyt.addWidget(self.yag_bakim_yok_cb)
        yag_lyt.setContentsMargins(0,0,0,0)
        
        def disable_yag_date():
            self.in_yag_bakim_tarihi.setDisabled(self.yag_bakim_yok_cb.isChecked())
        self.yag_bakim_yok_cb.toggled.connect(disable_yag_date)
        if "yag_bakim_tarihi" not in self.arac_data or self.arac_data.get("yag_bakim_tarihi") in ["-", "Yok", ""]:
            self.yag_bakim_yok_cb.setChecked(True)
            self.in_yag_bakim_tarihi.setDisabled(True)

        # Genel Bakım Tarihi
        self.in_genel_bakim_tarihi = QDateEdit()
        self.in_genel_bakim_tarihi.setCalendarPopup(True)
        if "genel_bakim_tarihi" in self.arac_data and self.arac_data.get("genel_bakim_tarihi") not in ["-", "Yok", ""]:
            self.in_genel_bakim_tarihi.setDate(QDate.fromString(self.arac_data["genel_bakim_tarihi"], "dd.MM.yyyy"))
        else:
            self.in_genel_bakim_tarihi.setDate(QDate.currentDate())
            
        self.genel_bakim_yok_cb = QCheckBox(_t("yok_chk", "Yok / Bilinmiyor"))
        genel_lyt = QHBoxLayout()
        genel_lyt.addWidget(self.in_genel_bakim_tarihi)
        genel_lyt.addWidget(self.genel_bakim_yok_cb)
        genel_lyt.setContentsMargins(0,0,0,0)
        
        def disable_genel_date():
            self.in_genel_bakim_tarihi.setDisabled(self.genel_bakim_yok_cb.isChecked())
        self.genel_bakim_yok_cb.toggled.connect(disable_genel_date)
        if "genel_bakim_tarihi" not in self.arac_data or self.arac_data.get("genel_bakim_tarihi") in ["-", "Yok", ""]:
            self.genel_bakim_yok_cb.setChecked(True)
            self.in_genel_bakim_tarihi.setDisabled(True)

        self.in_ruhsat_no = QLineEdit(self.arac_data.get("ruhsat_no", ""))
        self.in_ruhsat_no.setPlaceholderText(_t("ruhsat_belge_no", "Ruhsat Belge No"))

        left_form.addRow(_t("plaka_lbl", "Plaka:"), self.in_plaka)
        left_form.addRow(_t("firma_sahibi_lbl", "Araç Sahibi (Firma):"), self.in_firma)
        left_form.addRow(_t("kurum_lbl", "Çalıştığı Kurum:"), self.in_kurum)
        left_form.addRow(_t("vergi_tc_lbl", "Kimlik/Vergi No:"), self.in_vergi_no)
        left_form.addRow(_t("sofor_lbl", "Araç Şoförü:"), self.in_sofor)
        left_form.addRow(_t("sofor_tel_lbl", "Şoför Tel:"), self.in_sofor_tel)
        left_form.addRow(_t("marka_lbl", "Marka:"), self.in_marka)
        left_form.addRow(_t("model_lbl", "Model:"), self.in_model)
        left_form.addRow(_t("model_yili_lbl", "Model Yılı:"), self.in_yil)
        left_form.addRow(_t("yakit_tipi_lbl", "Yakıt Tipi:"), self.in_yakit)
        left_form.addRow(_t("vites_tipi_lbl", "Vites Tipi:"), self.in_vites)

        right_form.addRow(_t("guncel_km_lbl", "Güncel KM:"), self.in_km)
        right_form.addRow(_t("sasi_no_lbl", "Şasi No:"), self.in_sasi)
        right_form.addRow(_t("ruhsat_belge_no_lbl", "Ruhsat Belge No:"), self.in_ruhsat_no)
        right_form.addRow(_t("utts_durumu_lbl", "UTTS Durumu:"), self.in_utts)
        right_form.addRow(_t("muayene_bitis_lbl", "Muayene Bitiş:"), self.in_muayene_tarihi)
        right_form.addRow(_t("trafik_bitis_lbl", "Trafik Sigortası Bitiş:"), self.in_trafik_tarihi)
        right_form.addRow(_t("kasko_bitis_lbl", "Kasko Bitiş:"), kasko_lyt)
        right_form.addRow(_t("koltuk_sigortasi_lbl", "Koltuk Sig. Bitiş:"), self.in_koltuk_sigortasi)
        right_form.addRow(_t("yag_bakim_lbl", "Son Yağ Bakımı:"), yag_lyt)
        right_form.addRow(_t("genel_bakim_lbl", "Son Genel Bakım:"), genel_lyt)
        
        h_layout.addLayout(left_form)
        h_layout.addLayout(right_form)
        layout.addLayout(h_layout)
        layout.addStretch()

        # Dialog Buttons
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.button(QDialogButtonBox.StandardButton.Ok).setText(_t("kaydet", "Kaydet"))
        btn_box.button(QDialogButtonBox.StandardButton.Cancel).setText(_t("iptal_btn", "İptal"))
        btn_box.button(QDialogButtonBox.StandardButton.Cancel).setStyleSheet("background-")
        btn_box.accepted.connect(self.validate_and_accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def validate_and_accept(self):
        if not self.in_plaka.text().strip():
            QMessageBox.warning(self, _t("msg_hata", "Hata"), _t("msg_plakaalanzor", "Plaka alanı zorunludur."))
            return
        self.accept()

    def get_data(self):
        return {
            "id": self.arac_data.get("id", str(uuid.uuid4())),
            "plaka": self.in_plaka.text().strip().upper(),
            "firma": self.in_firma.text().strip(),
            "kurum": self.in_kurum.text().strip(),
            "vergi_no": self.in_vergi_no.text().strip(),
            "sofor": self.in_sofor.text().strip(),
            "sofor_tel": self.in_sofor_tel.text().strip(),
            "marka": self.in_marka.text().strip(),
            "model": self.in_model.text().strip(),
            "yil": self.in_yil.text().strip(),
            "yakit": self.in_yakit.currentText(),
            "vites": self.in_vites.currentText(),
            "km": self.in_km.text().strip(),
            "sasi": self.in_sasi.text().strip(),
            "utts": self.in_utts.currentText(),
            "muayene_tarihi": self.in_muayene_tarihi.date().toString("dd.MM.yyyy"),
            "trafik_tarihi": self.in_trafik_tarihi.date().toString("dd.MM.yyyy"),
            "kasko_tarihi": "-" if self.kasko_yok_cb.isChecked() else self.in_kasko_tarihi.date().toString("dd.MM.yyyy"),
            "koltuk_sigortasi": self.in_koltuk_sigortasi.date().toString("dd.MM.yyyy"),
            "yag_bakim_tarihi": "-" if self.yag_bakim_yok_cb.isChecked() else self.in_yag_bakim_tarihi.date().toString("dd.MM.yyyy"),
            "genel_bakim_tarihi": "-" if self.genel_bakim_yok_cb.isChecked() else self.in_genel_bakim_tarihi.date().toString("dd.MM.yyyy"),
            "ruhsat_no": self.in_ruhsat_no.text().strip()
        }

class EvrakListWidget(QListWidget):
    """Sürükle-bırak destekli evrak listesi"""
    files_dropped = None  # callback

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setDragDropMode(QAbstractItemView.DragDropMode.DropOnly)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            paths = [u.toLocalFile() for u in event.mimeData().urls() if u.isLocalFile()]
            if paths and callable(self.files_dropped):
                self.files_dropped(paths)
            event.acceptProposedAction()

class EvrakDialog(QDialog):
    def __init__(self, parent=None, arac=None, mode="arac"):
        super().__init__(parent)
        self.arac = arac
        self.mode = mode
        
        base_evrak_path = str(APP_EVRAK_DIR)
        
        if self.mode == "sofor":
            sofor_adi = self.arac.get('sofor', '').strip() or 'Bilinmeyen_Sofor'
            self.setWindowTitle(f"{_t('sofor_evraklari', 'Şoför Evrakları')} - {sofor_adi}")
            self.lbl_aciklama_metni = _t("qlabel_sofore_ait_eklenen", "Şoföre ait eklenen dosyalar (Ehliyet, SRC, Psikoteknik vs.)")
            self.evrak_klasoru = os.path.join(base_evrak_path, "Şoförler", sofor_adi)
        else:
            plaka = self.arac.get('plaka', 'Bilinmeyen_Arac')
            self.setWindowTitle(f"{_t('evrak_yonetimi_baslik', 'Evrak Yönetimi')} - {plaka}")
            self.lbl_aciklama_metni = _t("qlabel_araca_ait_eklenen_pd", "Araca ait eklenen PDF ve diğer dosyalar (Ruhsat, Sigorta, Güzergah vs.)")
            self.evrak_klasoru = os.path.join(base_evrak_path, "Araçlar", plaka)
            
        self.resize(600, 450)
            
        self.init_ui()
        
        # Eğer eski isimli (plaka_id) bir klasör varsa yeni yerine (Araçlar/Plaka) taşı (Sadece araç için)
        if self.mode == "arac":
            eski_klasor = os.path.join(base_evrak_path, f"{self.arac.get('plaka', '')}_{self.arac.get('id', '')}")
            if os.path.exists(eski_klasor) and self.evrak_klasoru != eski_klasor:
                if not os.path.exists(self.evrak_klasoru):
                    os.makedirs(self.evrak_klasoru)
                import shutil
                try:
                    for f_name in os.listdir(eski_klasor):
                        shutil.move(os.path.join(eski_klasor, f_name), os.path.join(self.evrak_klasoru, f_name))
                    os.rmdir(eski_klasor)
                except Exception as e:
                    pass

        self.check_license()

    def check_license(self):
        lic_file = os.path.join(os.path.expanduser('~'), '.gelir_gider_license.json')
        try:
            from datetime import datetime
            with open(lic_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            expire_date = datetime.fromisoformat(data['expire_date'])
        except:
            # First run, force lock
            from datetime import datetime
            expire_date = datetime(2000, 1, 1)

        if datetime.now() > expire_date:
            from PyQt6.QtWidgets import QGraphicsBlurEffect
            blur = QGraphicsBlurEffect()
            blur.setBlurRadius(10)
            self.centralWidget().setGraphicsEffect(blur)
            dlg = LicenseDialog(self)
            dlg.exec()
            self.centralWidget().setGraphicsEffect(None)

    def init_ui(self):
        palette = get_theme_palette(app_config.get("theme", "dark"))
        self.setStyleSheet(f"""
            QDialog {{ background-color: {palette['main_bg']}; }}
            QLabel {{ font-weight: bold; }}
            QListWidget {{ background-color: {palette['input_bg']}; border: 1px solid {palette['input_border']}; border-radius: 4px; color: {palette['text'] if app_config.get('theme', 'dark') == 'light' else palette['white']}; padding: 5px; font-size: 14px; }}
            QPushButton {{ background-color: {palette['primary_accent']}; color: {palette['white']}; padding: 8px; border-radius: 4px; font-weight: bold; }}
            QPushButton:hover {{ background-color: {palette['primary_hover']}; }}
            QPushButton#btnSil {{ background-color: {palette['danger_bg']}; }}
            QPushButton#btnSil:hover {{ background-color: {palette['danger_hover']}; }}
        """)
        
        main_lyt = QVBoxLayout(self)
        
        self.lbl_bilgi = QLabel(self.lbl_aciklama_metni)
        main_lyt.addWidget(self.lbl_bilgi)
        self.lbl_drop_hint = QLabel(_t("evrak_surukle_birak", "↓ Dosyaları buraya sürükleyip bırakabilirsiniz"))
        self.lbl_drop_hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_drop_hint.setStyleSheet(f"color: {palette['muted_hint']}; font-size: 12px; padding: 2px;")
        main_lyt.addWidget(self.lbl_drop_hint)
        self.list_evrak = EvrakListWidget()
        self.list_evrak.files_dropped = self.evraklar_yukle
        self.list_evrak.itemDoubleClicked.connect(self.evrak_ac)
        main_lyt.addWidget(self.list_evrak)
        self.load_evraklar()
        btn_lyt = QHBoxLayout()
        self.btn_ekle = QPushButton(_t("qpushbutton_evrak_yukle", "Evrak Yükle"))
        self.btn_ekle.clicked.connect(self.evrak_ekle)
        self.btn_ac = QPushButton(_t("qpushbutton_secili_evragi_ac", "Seçili Evrağı Aç"))
        self.btn_ac.clicked.connect(self.evrak_ac)
        self.btn_sil = QPushButton(_t("qpushbutton_secili_evragi_sil", "Seçili Evrağı Sil"))
        self.btn_sil.setObjectName("btnSil")
        self.btn_sil.clicked.connect(self.evrak_sil)
        self.btn_klasor_ac = QPushButton(_t("qpushbutton_evrak_klasorunu_ac", "Evrak Klasörünü Aç"))
        self.btn_klasor_ac.setStyleSheet(build_button_style(palette["success_bg"], palette["white"], palette["success_hover"], padding="8px"))
        self.btn_klasor_ac.clicked.connect(self.klasor_ac)
        btn_lyt.addWidget(self.btn_ekle)
        btn_lyt.addWidget(self.btn_ac)
        btn_lyt.addWidget(self.btn_klasor_ac)
        btn_lyt.addWidget(self.btn_sil)
        main_lyt.addLayout(btn_lyt)
        
    def load_evraklar(self):
        self.list_evrak.clear()
        if os.path.exists(self.evrak_klasoru):
            for file in os.listdir(self.evrak_klasoru):
                self.list_evrak.addItem(file)
                
    def evrak_ekle(self):
        files, _ = QFileDialog.getOpenFileNames(self, _t("evrak_sec_title", "Evrak Seç (PDF, Görsel vs.)"), "", _t("tum_dosyalar", "Tüm Dosyalar (*.*)"))
        if files:
            self.evraklar_yukle(files)

    def evraklar_yukle(self, file_paths):
        import shutil
        if not os.path.exists(self.evrak_klasoru):
            os.makedirs(self.evrak_klasoru)
            
        for file_path in file_paths:
            filename = os.path.basename(file_path)
            hedef_yol = os.path.join(self.evrak_klasoru, filename)
            
            # Eğer kaynak ve hedef yol tamamen aynıysa işlemi atla
            if os.path.normcase(os.path.abspath(file_path)) == os.path.normcase(os.path.abspath(hedef_yol)):
                continue
                
            try:
                shutil.copy2(file_path, hedef_yol)
            except shutil.SameFileError:
                continue
            except Exception as e:
                if "are the same file" in str(e):
                    continue
                QMessageBox.warning(self, _t("hata_baslik", "Hata"), f"{filename} kopyalanırken hata oluştu: {str(e)}")
        self.load_evraklar()
        QMessageBox.information(self, _t("msg_baarl", "Başarılı"), _t("msg_seilenevrakla", "Seçilen evraklar başarıyla yüklendi."))
            
    def get_secili_evrak(self):
        item = self.list_evrak.currentItem()
        if not item:
            return None
        return os.path.join(self.evrak_klasoru, item.text())

    def evrak_ac(self, item=None):
        hedef_yol = self.get_secili_evrak()
        if hedef_yol and os.path.exists(hedef_yol):
            try:
                os.startfile(hedef_yol)
            except Exception as e:
                QMessageBox.warning(self, _t("hata_baslik", "Hata"), f"Dosya açılamadı:\n{str(e)}")
        else:
            QMessageBox.warning(self, _t("msg_uyar", "Uyarı"), _t("msg_ltfenamaki", "Lütfen açmak için bir evrak seçin."))
            
    def evrak_sil(self):
        hedef_yol = self.get_secili_evrak()
        if hedef_yol and os.path.exists(hedef_yol):
            msg = QMessageBox(self)
            msg.setWindowTitle(_t("onay_baslik", "Onay"))
            onay_metni = _t("msg_evrak_sil_emin", "Seçili evrağı silmek istediğinize emin misiniz?\n")
            msg.setText(f"{onay_metni} {os.path.basename(hedef_yol)}")
            btn_evet = msg.addButton(_t("evet_sil", "Evet, Sil"), QMessageBox.ButtonRole.YesRole)
            btn_hayir = msg.addButton(_t("hayir_iptal", "Hayır, İptal"), QMessageBox.ButtonRole.NoRole)
            msg.exec()
            
            if msg.clickedButton() == btn_evet:
                try:
                    os.remove(hedef_yol)
                    if not os.listdir(self.evrak_klasoru):
                        os.rmdir(self.evrak_klasoru)
                        
                    self.load_evraklar()
                except Exception as e:
                    QMessageBox.warning(self, _t("hata_baslik", "Hata"), f"Dosya silinirken hata oluştu:\n{str(e)}")
        else:
            QMessageBox.warning(self, _t("msg_uyar", "Uyarı"), _t("msg_ltfensilmeki", "Lütfen silmek için bir evrak seçin."))
            
    def klasor_ac(self):
        if not os.path.exists(self.evrak_klasoru):
            os.makedirs(self.evrak_klasoru)
        os.startfile(self.evrak_klasoru)

    def closeEvent(self, event):
        # Pencere kapanırken eğer klasör oluşturulmuş ama içi boş kalmışsa silsin (Gereksiz boş klasör oluşumunu engellemek için)
        if hasattr(self, 'evrak_klasoru') and os.path.exists(self.evrak_klasoru):
            if not os.listdir(self.evrak_klasoru):
                try:
                    os.rmdir(self.evrak_klasoru)
                except Exception:
                    pass
        super().closeEvent(event)

DATA_FILE = app_data_path("fatura_kayitlari.json")


class LicenseDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        dialog_theme = get_license_dialog_theme(app_config.get("theme", "dark"))
        self.setWindowTitle(_t("lisans_aktivasyonu", "Lisans Etkinleştirme"))
        self.setFixedSize(420, 230)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.CustomizeWindowHint |
            Qt.WindowType.WindowTitleHint
        )
        self.setStyleSheet(dialog_theme["dialog"])
        
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(16, 16, 16, 16)
        
        info_frame = QFrame()
        info_frame.setFrameShape(QFrame.Shape.StyledPanel)
        info_frame.setStyleSheet(dialog_theme["info_frame"])
        info_layout = QVBoxLayout(info_frame)
        info_layout.setContentsMargins(0, 0, 0, 0)
        
        self.info_lbl = QLabel(_t("uyari_sure_doldu", "Lisans süreniz sona erdi.\nLütfen aşağıya geçerli lisans anahtarınızı girerek uygulamayı etkinleştirin."))
        self.info_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.info_lbl.setWordWrap(True)
        self.info_lbl.setStyleSheet(dialog_theme["info_text"])
        info_layout.addWidget(self.info_lbl)
        layout.addWidget(info_frame)
        
        self.key_input = QLineEdit()
        self.key_input.setPlaceholderText("LCS-XXXX-XXXX")
        self.key_input.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.key_input.setStyleSheet(dialog_theme["input"])
        layout.addWidget(self.key_input)
        
        self.btn_verify = QPushButton(_t("dogrula_buton", "Lisansı Etkinleştir"))
        self.btn_verify.setStyleSheet(dialog_theme["button"])
        self.btn_verify.clicked.connect(self.verify_key)
        layout.addWidget(self.btn_verify)

        self.error_lbl = QLabel("")
        self.error_lbl.setStyleSheet(dialog_theme["error"])
        self.error_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.error_lbl)

    def verify_key(self):
        key = self.key_input.text().strip().upper()
        if not key:
            self.error_lbl.setText(_t("lutfen_lisans_gir", "Lütfen bir lisans anahtarı girin."))
            return

        import hashlib
        parts = key.split('-')
        if len(parts) == 3 and parts[0] == 'LCS':
            expected_hash = hashlib.sha256((parts[1] + "GELIR_GIDER_SECRET_2026").encode()).hexdigest()[:4].upper()
            if parts[2] == expected_hash:
                # Add 15 days
                self.save_new_license()
                self.accept()
                return

        self.error_lbl.setText(_t("gecersiz_lisans", "Geçersiz lisans anahtarı!"))

    def save_new_license(self):
        lic_file = os.path.join(os.path.expanduser('~'), '.gelir_gider_license.json')
        from datetime import datetime, timedelta
        try:
            with open(lic_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except:
            data = {}
        
        new_expire = datetime.now() + timedelta(days=15)
        data['expire_date'] = new_expire.isoformat()
        
        with open(lic_file, 'w', encoding='utf-8') as f:
            json.dump(data, f)



from PyQt6.QtWidgets import QPushButton
from PyQt6.QtCore import QVariantAnimation, QEasingCurve, Qt
from PyQt6.QtGui import QPainter, QPen, QColor

class MorphingHamburger(QPushButton):
    def __init__(self, title='', parent=None):
        super().__init__('', parent)
        self.setFixedSize(40, 40)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._progress = 1.0
        self.anim = QVariantAnimation(self)
        self.anim.setDuration(400)
        self.anim.setEasingCurve(QEasingCurve.Type.InOutQuart)
        self.anim.valueChanged.connect(self.set_progress)
        palette = get_theme_palette(app_config.get("theme", "dark"))
        self.setStyleSheet(
            f"QPushButton {{ background: transparent; border: none; }} QPushButton:hover {{ background-color: {palette['hamburger_hover_bg']}; border-radius: 4px; }}"
        )
    
    def set_progress(self, p):
        self._progress = p
        self.update()
        
    def toggle(self, is_open):
        self.anim.setStartValue(self._progress)
        self.anim.setEndValue(1.0 if is_open else 0.0)
        self.anim.start()
        
    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        cx, cy = self.width()/2, self.height()/2
        w = 18.0
        
        palette = get_theme_palette(app_config.get("theme", "dark"))
        pen = QPen(QColor(palette["hamburger_color"]))
        pen.setWidthF(2.5)
        pen.setCapStyle(Qt.PenCapStyle.RoundCap)
        painter.setPen(pen)
        
        painter.translate(cx, cy)
        p = self._progress
        
        painter.rotate(90 * p)
        
        painter.save()
        painter.translate(0, -6 * (1 - p))
        painter.rotate(45 * p)
        painter.drawLine(int(-w/2), 0, int(w/2), 0)
        painter.restore()
        
        painter.save()
        painter.setOpacity(1 - p)
        painter.drawLine(int(-w/2), 0, int(w/2), 0)
        painter.restore()
        
        painter.save()
        painter.translate(0, 6 * (1 - p))
        painter.rotate(-45 * p)
        painter.drawLine(int(-w/2), 0, int(w/2), 0)
        painter.restore()
        
        painter.end()


class FaturaApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.records = []
        self.araclar = []
        self.editing_record_id = None
        self.load_records()
        
        self.flash_state = False
        self.pulse_step = 0
        self.flash_timer = QTimer(self)
        self.flash_timer.timeout.connect(self.update_pulse_labels)
        self.flash_timer.start(50) # 50ms for smooth 20 fps animation
        
        self.init_ui()
        self.update_check_timer = QTimer(self)
        self.update_check_timer.timeout.connect(self._auto_update_check_tick)
        self.configure_update_timer()
        QTimer.singleShot(8000, self.maybe_check_for_updates_on_startup)

    def update_pulse_labels(self):
        import math
        self.pulse_step += 1
        # Sine wave ranges from -1 to 1. Normalize to 0 to 1.
        factor = (math.sin(self.pulse_step * 0.1) + 1) / 2.0  
        
        def mix_colors(c1, c2, f):
            c1_r, c1_g, c1_b = int(c1[1:3], 16), int(c1[3:5], 16), int(c1[5:7], 16)
            c2_r, c2_g, c2_b = int(c2[1:3], 16), int(c2[3:5], 16), int(c2[5:7], 16)
            r = int(c1_r + (c2_r - c1_r) * f)
            g = int(c1_g + (c2_g - c1_g) * f)
            b = int(c1_b + (c2_b - c1_b) * f)
            return f"#{r:02x}{g:02x}{b:02x}"

        palette = get_theme_palette(app_config.get("theme", "dark"))
        base_color = palette["pulse_base"]

        labels = [getattr(self, 'lbl_arac_muayene', None), getattr(self, 'lbl_arac_trafik', None), 
                  getattr(self, 'lbl_arac_kasko', None), getattr(self, 'lbl_arac_koltuk', None)]
                  
        for lbl in labels:
            if lbl and hasattr(lbl, 'target_color'):
                color = lbl.target_color
                if color == "stable":
                    lbl.setStyleSheet(f"color: {palette['positive_text']}; font-weight: bold;")
                elif color: # yanıp sönen renk
                    mixed_color = mix_colors(base_color, color, factor)
                    lbl.setStyleSheet(f"color: {mixed_color}; font-weight: bold;")
                else: # tarih yok veya geçersiz
                    lbl.setStyleSheet("")

    def check_license(self):
        lic_file = os.path.join(os.path.expanduser('~'), '.gelir_gider_license.json')
        try:
            from datetime import datetime
            with open(lic_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            expire_date = datetime.fromisoformat(data['expire_date'])
        except:
            # First run, force lock
            from datetime import datetime
            expire_date = datetime(2000, 1, 1)

        if datetime.now() > expire_date:
            from PyQt6.QtWidgets import QGraphicsBlurEffect
            blur = QGraphicsBlurEffect()
            blur.setBlurRadius(10)
            self.centralWidget().setGraphicsEffect(blur)
            
            dlg = LicenseDialog(self)
            dlg.exec()
            
            self.centralWidget().setGraphicsEffect(None)

    def load_records(self):
        if os.path.exists(DATA_FILE):
            try:
                with open(DATA_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if isinstance(data, list): # Eski liste formatından çevir
                        self.records = data
                        self.araclar = []
                    elif isinstance(data, dict):
                        self.records = data.get("faturalar", [])
                        self.araclar = data.get("araclar", [])
            except Exception as e:
                print(f"Kayıtlar yüklenirken hata oluştu: {e}")
                self.records = []
                self.araclar = []
        else:
            self.records = []
            self.araclar = []

    def save_records(self):
        try:
            data = {
                "faturalar": self.records,
                "araclar": self.araclar
            }
            with open(DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Kayıtlar kaydedilirken hata oluştu: {e}")

    def init_ui(self, is_reload=False):
        self.setWindowTitle("Mali & Araç Yönetim Paneli")
        if not is_reload: self.resize(1300, 850)
        current_theme = app_config.get("theme", "dark")
        palette = get_theme_palette(current_theme)
        dash_theme = get_dashboard_theme(current_theme)
        vehicle_theme = get_vehicle_summary_styles(current_theme)

        # APPLY THEME
        self.apply_theme(current_theme)

        self.main_widget = QWidget()
        self.setCentralWidget(self.main_widget)
        
        self.main_vbox = QVBoxLayout(self.main_widget)
        self.main_vbox.setContentsMargins(0, 0, 0, 0)
        self.main_vbox.setSpacing(0)

        self.root_layout = QHBoxLayout()
        self.root_layout.setContentsMargins(0, 0, 0, 0)
        self.root_layout.setSpacing(0)
        
        self.main_vbox.addLayout(self.root_layout, stretch=1)

        # LEFT SIDEBAR MENU
        self.sidebar_frame = QFrame()
        self.sidebar_frame.setFixedWidth(260)
        self.sidebar_frame.setObjectName("SidebarFrame")
        self.sidebar_layout = QVBoxLayout(self.sidebar_frame)
        self.sidebar_layout.setContentsMargins(10, 10, 10, 35)  # Alt bosluk taskbar uzerinde kalmasi icin artirildi
        self.sidebar_layout.setSpacing(10)

        # HAMBURGER BUTTON
        self.btn_hamburger = MorphingHamburger()
        self.btn_hamburger.setFixedSize(40, 40)
        self.btn_hamburger.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        self.btn_hamburger.setStyleSheet(
            f"QPushButton {{ background: transparent; color: {palette['hamburger_color']}; border: none; }} "
            f"QPushButton:hover {{ color: {palette['hamburger_hover_text']}; background-color: {palette['hamburger_hover_bg']}; border-radius: 4px; }}"
        )
        self.btn_hamburger.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_hamburger.clicked.connect(self.toggle_sidebar)
        ham_lyt = QHBoxLayout()
        ham_lyt.addWidget(self.btn_hamburger)
        ham_lyt.addStretch()
        self.sidebar_layout.addLayout(ham_lyt)
        self.sidebar_menu_scroll = QScrollArea()
        self.sidebar_menu_scroll.setWidgetResizable(True)
        self.sidebar_menu_scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.sidebar_menu_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.sidebar_menu_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.sidebar_menu_scroll.setStyleSheet("background: transparent;")

        self.sidebar_menu_widget = QWidget()
        self.sidebar_menu_widget.setStyleSheet("background: transparent;")
        self.sidebar_menu_layout = QVBoxLayout(self.sidebar_menu_widget)
        self.sidebar_menu_layout.setContentsMargins(3, 3, 3, 3)
        self.sidebar_menu_layout.setSpacing(10)
        self.sidebar_menu_scroll.setWidget(self.sidebar_menu_widget)
        self.sidebar_layout.addWidget(self.sidebar_menu_scroll, stretch=1)


        self.btn_menu_dashboard = QToolButton()
        self.btn_menu_dashboard.setText(_t("menu_dashboard", "Ana Sayfa\nPaneli"))
        self.btn_menu_dashboard.setIcon(QIcon(str(APP_ANIM_DIR / "Ana_Sayfa.png")))
        self.btn_menu_dashboard.setIconSize(QSize(110, 110))
        self.btn_menu_dashboard.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.btn_menu_dashboard.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        self.btn_menu_fatura = QToolButton()
        self.btn_menu_fatura.setText(_t("menu_fatura", "Fatura\nYönetimi"))
        self.btn_menu_fatura.setIcon(QIcon(str(APP_ANIM_DIR / "Menu_fatura.png")))
        self.btn_menu_fatura.setIconSize(QSize(110, 110))
        self.btn_menu_fatura.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.btn_menu_fatura.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.btn_menu_arac = QToolButton()
        self.btn_menu_arac.setText(_t("menu_arac", "Araç\nYönetimi"))
        self.btn_menu_arac.setIcon(QIcon(str(APP_ANIM_DIR / "menu_arac.png")))
        self.btn_menu_arac.setIconSize(QSize(110, 110))
        self.btn_menu_arac.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.btn_menu_arac.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.btn_menu_ayarlar = QToolButton()
        self.btn_menu_ayarlar.setText(_t("menu_ayarlar", "Ayarlar"))
        self.btn_menu_ayarlar.setIcon(QIcon(str(APP_ANIM_DIR / "Ayarlar.png")))
        self.btn_menu_ayarlar.setIconSize(QSize(110, 110))
        self.btn_menu_ayarlar.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
        self.btn_menu_ayarlar.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        
        for btn in (self.btn_menu_dashboard, self.btn_menu_fatura, self.btn_menu_arac, self.btn_menu_ayarlar):
            btn.setFixedHeight(176)
            btn.setObjectName("SidebarButton")
            btn.setCheckable(True)
            self.sidebar_menu_layout.addWidget(btn)

        self.sidebar_menu_layout.addStretch()

        # Compute remaining days
        remaining_days = 0
        lic_file = os.path.join(os.path.expanduser('~'), '.gelir_gider_license.json')
        try:
            from datetime import datetime
            with open(lic_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            expire_date = datetime.fromisoformat(data['expire_date'])
            delta = expire_date.date() - datetime.now().date()
            remaining_days = max(0, delta.days)
        except Exception:
            pass

        self.lbl_version = QLabel(f"v1.0.0 | {_t('kalan_lisans', 'Lisans')} : {remaining_days} {_t('gun', 'Gün')}")
        self.lbl_version.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_version.setStyleSheet(f"color: {palette['positive_text']}; font-weight: bold; font-size: 11px;")
        self.sidebar_layout.addWidget(self.lbl_version)

        # Connect buttons 
        self.btn_menu_dashboard.clicked.connect(lambda: self.switch_page(0))
        self.btn_menu_fatura.clicked.connect(lambda: self.switch_page(1))
        self.btn_menu_arac.clicked.connect(lambda: self.switch_page(2))
        self.btn_menu_ayarlar.clicked.connect(lambda: self.switch_page(3))
        
        self.root_layout.addWidget(self.sidebar_frame)

        # RIGHT STACKED WIDGET PAGE
        self.stack = QStackedWidget()
        
        self.main_scroll_area = QScrollArea()
        self.main_scroll_area.setWidgetResizable(True)
        self.main_scroll_area.setWidget(self.stack)
        self.main_scroll_area.setFrameShape(QFrame.Shape.NoFrame)
        self.main_scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.main_scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        self.root_layout.addWidget(self.main_scroll_area)

        # --- FATURA WIDGET PAGE ---
        self.page_fatura = QWidget()
        main_layout = QVBoxLayout(self.page_fatura)
        main_layout.setContentsMargins(20, 20, 20, 35)

        # Title
        title_label = QLabel(_t("qlabel_fatura_yonetim_panel", "Fatura Yönetim Paneli"))
        title_label.setFont(get_custom_font("Quicksand", 20, QFont.Weight.DemiBold))
        title_label.setStyleSheet(" padding: 10px; background-color: transparent;")
        main_layout.addWidget(title_label)

        # Top section: Form and Metrics
        top_layout = QHBoxLayout()
        main_layout.addLayout(top_layout)

        # Input Form
        form_group = QGroupBox(_t("qgroupbox_yeni_fatura_ekle", "Yeni Fatura Ekle"))
        form_group.setStyleSheet(get_invoice_form_styles(current_theme))
        form_layout = QGridLayout(form_group)
        form_layout.setSpacing(12)
        form_layout.setContentsMargins(15, 20, 15, 15)

        self.date_input = QDateEdit()
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setCalendarPopup(True)

        self.tip_input = QComboBox()
        self.tip_input.addItem(_t("tip_gelen", "Gelen"), "Gelen"); self.tip_input.addItem(_t("tip_giden", "Giden"), "Giden")

        self.fatura_no_input = QLineEdit()
        self.fatura_no_input.setPlaceholderText(_t("orn_fat", "ÖRN: FAT-2026-0001"))

        self.firma_input = QLineEdit()
        self.firma_input.setPlaceholderText(_t("tedarikci_musteri", "Tedarikçi / Müşteri"))

        self.aciklama_input = QLineEdit()
        self.aciklama_input.setPlaceholderText(_t("hizmet_urun_not", "Hizmet, ürün veya not"))

        self.kdv_input = QComboBox()
        self.kdv_input.setEditable(True)
        for opt in KDV_OPTIONS:
            self.kdv_input.addItem(f"% {opt}", opt)
        self.kdv_input.setCurrentText("% 20")
        self.kdv_input.currentIndexChanged.connect(self.update_live_preview)
        self.kdv_input.editTextChanged.connect(self.update_live_preview)

        self.tevkifat_input = QComboBox()
        for text, value in TEVKIFAT_OPTIONS:
            if text == "yok_tevkifat_key":
                self.tevkifat_input.addItem(_t("yok_tevkifat", "Yok (0/10)"), value)
            else:
                self.tevkifat_input.addItem(text, value)
        self.tevkifat_input.currentIndexChanged.connect(self.update_live_preview)

        self.matrah_input = QLineEdit()
        self.matrah_input.setPlaceholderText(_t("tutar_girin", "Tutar girin..."))
        self.matrah_input.textChanged.connect(self.update_live_preview)

        # Create a container to show "TL" as a faded label
        matrah_widget = QWidget()
        matrah_lyt = QHBoxLayout(matrah_widget)
        matrah_lyt.setContentsMargins(0, 0, 0, 0)
        matrah_lyt.setSpacing(5)
        matrah_lyt.addWidget(self.matrah_input)
        
        lbl_tl = QLabel(_t("qlabel_tl", "TL"))
        lbl_tl.setStyleSheet(" background-color: transparent; border: none; font-weight: bold;")
        matrah_lyt.addWidget(lbl_tl)

        self.live_preview_label = QLabel(_t("qlabel_kdv_000_tl__tevkifat", "KDV: 0.00 TL | Tevkifat: 0.00 TL | Toplam: 0.00 TL"))
        self.live_preview_label.setFont(get_custom_font("Quicksand", 11, QFont.Weight.DemiBold))
        self.live_preview_label.setStyleSheet(" background-color: transparent; padding: 8px; border-radius: 4px;")
        self.live_preview_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.btn_ekle = QPushButton(_t("qpushbutton_faturayi_kaydet", "Faturayı Kaydet"))
        self.btn_ekle.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_ekle.setStyleSheet(build_button_style(palette["primary_accent"], palette["white"], palette["primary_hover"], padding="12px", font_size="14px"))
        self.btn_ekle.clicked.connect(self.add_record)

        self.btn_iptal = QPushButton(_t("qpushbutton_iptal_et", "İptal Et"))
        self.btn_iptal.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_iptal.setStyleSheet("QPushButton { background- color: white; padding: 12px; font-size: 14px; font-weight: bold; border-radius: 6px; } QPushButton:hover { background- }")
        self.btn_iptal.clicked.connect(self.cancel_edit)
        self.btn_iptal.hide()

        btn_layout_widget = QWidget()
        btn_layout = QHBoxLayout(btn_layout_widget)
        btn_layout.setContentsMargins(0, 0, 0, 0)
        btn_layout.addWidget(self.btn_ekle)
        btn_layout.addWidget(self.btn_iptal)

        # Re-arrange grid for 2 matching columns of inputs exactly (4 columns total)
        # Row 0: Tarih, Fatura Tipi
        form_layout.addWidget(QLabel(_t("qlabel_tarih", "Tarih:")), 0, 0)
        form_layout.addWidget(self.date_input, 0, 1)
        form_layout.addWidget(QLabel(_t("qlabel_fatura_tipi", "Fatura Tipi:")), 0, 2)
        form_layout.addWidget(self.tip_input, 0, 3)

        # Row 1: Fatura No, Firma
        form_layout.addWidget(QLabel(_t("qlabel_fatura_no", "Fatura No:")), 1, 0)
        form_layout.addWidget(self.fatura_no_input, 1, 1)
        form_layout.addWidget(QLabel(_t("qlabel_firma", "Firma:")), 1, 2)
        form_layout.addWidget(self.firma_input, 1, 3)

        # Row 2: Açıklama, Matrah
        form_layout.addWidget(QLabel(_t("qlabel_aciklama", "Açıklama:")), 2, 0)
        form_layout.addWidget(self.aciklama_input, 2, 1)
        form_layout.addWidget(QLabel(_t("qlabel_matrah", "Matrah:")), 2, 2)
        form_layout.addWidget(matrah_widget, 2, 3)

        # Row 3: KDV Oranı, Tevkifat Oranı
        form_layout.addWidget(QLabel(_t("qlabel_kdv_orani", "KDV Oranı:")), 3, 0)
        form_layout.addWidget(self.kdv_input, 3, 1)
        form_layout.addWidget(QLabel(_t("qlabel_tevkifat_orani", "Tevkifat Oranı:")), 3, 2)
        form_layout.addWidget(self.tevkifat_input, 3, 3)

        # Preview and Button
        form_layout.addWidget(self.live_preview_label, 4, 0, 1, 4)
        form_layout.addWidget(btn_layout_widget, 5, 0, 1, 4)

        # Make input columns stretch evenly
        form_layout.setColumnStretch(1, 1)
        form_layout.setColumnStretch(3, 1)

        top_layout.addWidget(form_group, stretch=2)

        # Metrics Summary
        metrics_group = QGroupBox(_t("qgroupbox_ozet_bilgiler", "Özet Bilgiler"))
        self.metrics_layout = QVBoxLayout(metrics_group)
        self.lbl_gelen_toplam = QLabel()
        self.lbl_giden_toplam = QLabel()
        self.lbl_net_fark = QLabel()
        self.lbl_kdv_fark = QLabel()

        for lbl in (self.lbl_gelen_toplam, self.lbl_giden_toplam, self.lbl_net_fark, self.lbl_kdv_fark):
            lbl.setFont(get_custom_font("Quicksand", 12))
            self.metrics_layout.addWidget(lbl)

        # KDV Dipnotu
        self.lbl_kdv_dipnot = QLabel()
        self.lbl_kdv_dipnot.setStyleSheet("background-color: transparent; margin-top: 5px;")
        self.lbl_kdv_dipnot.setTextFormat(Qt.TextFormat.RichText)
        self.lbl_kdv_dipnot.setWordWrap(True)
        self.metrics_layout.addWidget(self.lbl_kdv_dipnot)

        top_layout.addWidget(metrics_group, stretch=1)

        # Tabs for Tables
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)

        self.table_all = QTableWidget()
        self.table_gelen = QTableWidget()
        self.table_giden = QTableWidget()

        for tbl in (self.table_all, self.table_gelen, self.table_giden):
            tbl.setColumnCount(11)
            tbl.setHorizontalHeaderLabels([_t("tarih_hdr", "Tarih"), _t("tip_hdr", "Tip"), _t("fatura_no_hdr", "Fatura No"), _t("firma_hdr", "Firma"), _t("aciklama_hdr", "Açıklama"), _t("matrah_hdr", "Matrah"), _t("kdv_yuzde_hdr", "KDV %"), _t("kdv_tutari_hdr", "KDV Tutarı"), _t("tevkifat_hdr", "Tevkifat"), _t("tev_tutari_hdr", "Tev.Tutarı"), _t("toplam_hdr", "Toplam")])
            tbl.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
            tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
            tbl.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
            tbl.customContextMenuRequested.connect(lambda pos, t=tbl: self.show_context_menu(pos, t))
            tbl.itemClicked.connect(lambda item, t=tbl: self.handle_item_click(item, t))

        self.tabs.addTab(self.table_all, _t("tab_tm_kaytlar", "Tüm Kayıtlar"))
        self.tabs.addTab(self.table_gelen, _t("tab_gelen_faturalar", "Gelen Faturalar"))
        self.tabs.addTab(self.table_giden, _t("tab_giden_faturalar", "Giden Faturalar"))

        # Bottom Buttons
        bottom_layout = QHBoxLayout()
        self.btn_excel = QPushButton(_t("qpushbutton_excele_aktar_xlsx", "Excel'e Aktar (XLSX)"))
        self.btn_pdf = QPushButton(_t("qpushbutton_pdf_raporu_indir", "PDF Raporu İndir"))
        self.btn_clear = QPushButton(_t("qpushbutton_tum_kayitlari_temizl", "Tüm Kayıtları Temizle"))

        self.btn_excel.clicked.connect(self.export_excel)
        self.btn_pdf.clicked.connect(self.export_pdf)
        self.btn_clear.clicked.connect(self.clear_records)

        bottom_layout.addWidget(self.btn_excel)
        bottom_layout.addWidget(self.btn_pdf)
        bottom_layout.addWidget(self.btn_clear)
        main_layout.addLayout(bottom_layout)
        

        # --- DASHBOARD PAGE ---
        self.page_dashboard = QWidget()
        self.page_dashboard.setObjectName("DashboardPage")
        dash_lyt = QVBoxLayout(self.page_dashboard)
        dash_lyt.setContentsMargins(20, 20, 20, 35)
        dash_lyt.setSpacing(14)

        self.page_dashboard.setStyleSheet("""
            QWidget#DashboardPage {
                background: %s;
            }
        """ % dash_theme["page_bg"])
        
        dash_title = QLabel(_t("dashboard_title", "Sistem Genel Özeti (Dashboard)"))
        dash_title.setFont(get_custom_font("Quicksand", 24, QFont.Weight.Bold))
        dash_title.setStyleSheet(f"padding: 8px 4px 0 4px; background-color: transparent; color: {dash_theme['title']};")
        dash_lyt.addWidget(dash_title)

        dash_subtitle = QLabel(_t("dashboard_subtitle", "Tüm filo, fatura ve yaklaşan bakım süreçleri tek bakışta"))
        dash_subtitle.setFont(get_custom_font("Quicksand", 11, QFont.Weight.Medium))
        dash_subtitle.setStyleSheet(f"padding: 0 4px 8px 4px; background-color: transparent; color: {dash_theme['subtitle']};")
        dash_lyt.addWidget(dash_subtitle)

        self.lbl_dash_info = QLabel("-")
        self.lbl_dash_info.setFont(get_custom_font("Quicksand", 10, QFont.Weight.DemiBold))
        self.lbl_dash_info.setStyleSheet(f"""
            QLabel {{
                background-color: {dash_theme['info_bg']};
                border: 1px solid {dash_theme['info_border']};
                border-radius: 8px;
                color: {dash_theme['info_text']};
                padding: 8px 12px;
            }}
        """)
        dash_lyt.addWidget(self.lbl_dash_info)

        # Top KPI Cards
        kpi_w = QWidget()
        kpi_lyt = QHBoxLayout(kpi_w)
        kpi_lyt.setContentsMargins(0, 0, 0, 0)
        kpi_lyt.setSpacing(12)
        kpi_w.setStyleSheet("background-color: transparent;")
        dash_lyt.addWidget(kpi_w)

        def create_kpi(title, value_lbl_obj, accent_color, soft_bg):
            f = QFrame()
            f.setMinimumHeight(120)
            if current_theme == "light":
                f.setStyleSheet("""
                    QFrame {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 %s, stop:1 %s);
                        border-radius: 12px;
                        border: 1px solid %s;
                    }
                    QLabel { background-color: transparent; }
                """ % (soft_bg, dash_theme["card_gradient_end"], accent_color))
            else:
                f.setStyleSheet("""
                    QFrame {
                        background-color: %s;
                        border-radius: 12px;
                        border: 1px solid %s;
                    }
                    QLabel { background-color: transparent; }
                """ % (soft_bg, accent_color))
            l = QVBoxLayout(f)
            l.setContentsMargins(14, 10, 14, 10)
            l.setSpacing(4)
            t = QLabel(title)
            t.setFont(get_custom_font("Quicksand", 10, QFont.Weight.DemiBold))
            title_color = accent_color if current_theme == "light" else dash_theme['kpi_title']
            t.setStyleSheet(f"color: {title_color}; letter-spacing: 0.5px;")
            t.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
            value_lbl_obj.setFont(get_custom_font("Quicksand", 22, QFont.Weight.Bold))
            value_lbl_obj.setStyleSheet(f"color: {dash_theme['kpi_value']};")
            value_lbl_obj.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignBottom)
            l.addWidget(t)
            l.addWidget(value_lbl_obj)
            return f

        self.kpi_arac_sayisi = QLabel("0")
        self.kpi_gider_aylik = QLabel("0.00 TL")
        self.kpi_gelir_aylik = QLabel("0.00 TL")
        self.kpi_uyarilar = QLabel("0")

        kpi_lyt.addWidget(create_kpi(_t("kpi_toplam_arac", "Toplam Aktif Araç"), self.kpi_arac_sayisi, palette["info_hover"], dash_theme["card_blue_bg"]))
        kpi_lyt.addWidget(create_kpi(_t("kpi_aylik_gider", "Bu Ayki Giderler"), self.kpi_gider_aylik, palette["danger_hover"], dash_theme["card_red_bg"]))
        kpi_lyt.addWidget(create_kpi(_t("kpi_aylik_gelir", "Bu Ayki Gelirler"), self.kpi_gelir_aylik, palette["positive_text"], dash_theme["card_green_bg"]))
        kpi_lyt.addWidget(create_kpi(_t("kpi_kritik_uyari", "Kritik Uyarılar"), self.kpi_uyarilar, palette["warning_bg"], dash_theme["card_amber_bg"]))

        dash_badges = QWidget()
        dash_badges_lyt = QHBoxLayout(dash_badges)
        dash_badges_lyt.setContentsMargins(0, 0, 0, 0)
        dash_badges_lyt.setSpacing(10)
        dash_badges.setStyleSheet("background-color: transparent;")

        badge_style = f"""
            QLabel {{
                background-color: {dash_theme['badge_bg']};
                color: {dash_theme['badge_text']};
                border: 1px solid {dash_theme['badge_border']};
                border-radius: 8px;
                padding: 6px 10px;
            }}
        """
        self.lbl_alert_count = QLabel("-")
        self.lbl_alert_count.setFont(get_custom_font("Quicksand", 10, QFont.Weight.DemiBold))
        self.lbl_alert_count.setStyleSheet(badge_style)
        self.lbl_alert_count.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.lbl_event_count = QLabel("-")
        self.lbl_event_count.setFont(get_custom_font("Quicksand", 10, QFont.Weight.DemiBold))
        self.lbl_event_count.setStyleSheet(badge_style)
        self.lbl_event_count.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Filo İstatistikleri (Genel İstatistikler)
        self.ozet_gb = QGroupBox(_t("filo_istatistikleri", "Araç İstatistikleri"))
        self.ozet_gb.setFont(get_custom_font("Quicksand", 10, QFont.Weight.Bold))
        self.ozet_gb.setStyleSheet(vehicle_theme["group"])
        self.ozet_gb.setMaximumHeight(120)
        ozet_gb_lyt = QVBoxLayout(self.ozet_gb)
        ozet_gb_lyt.setContentsMargins(10, 15, 10, 10)
        self.lbl_arac_ozet = QLabel(_t("yukleniyor", "Yükleniyor..."))
        self.lbl_arac_ozet.setFont(get_custom_font("Quicksand", 10, QFont.Weight.Medium))
        self.lbl_arac_ozet.setStyleSheet(vehicle_theme["label"])
        self.lbl_arac_ozet.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.lbl_arac_ozet.setWordWrap(True)
        ozet_gb_lyt.addWidget(self.lbl_arac_ozet)
        dash_lyt.addWidget(self.ozet_gb)

        dash_filter_w = QWidget()
        dash_filter_w.setStyleSheet("background-color: transparent;")
        dash_filter_lyt = QHBoxLayout(dash_filter_w)
        dash_filter_lyt.setContentsMargins(0, 0, 0, 0)
        dash_filter_lyt.setSpacing(8)
        self.lbl_dash_period = QLabel(_t("dash_period_label", "Dönem:"))
        self.lbl_dash_period.setFont(get_custom_font("Quicksand", 10, QFont.Weight.DemiBold))
        self.lbl_dash_period.setStyleSheet(f"color: {palette['text']}; background-color: transparent;")
        self.combo_dash_period = QComboBox()
        self.combo_dash_period.setMinimumWidth(190)
        self.combo_dash_period.setStyleSheet(
            f"QComboBox {{ background-color: {palette['input_bg']}; color: {palette['text']}; border: 1px solid {palette['input_border']}; border-radius: 6px; padding: 5px 8px; }} "
            f"QComboBox QAbstractItemView {{ background-color: {palette['surface_bg']}; color: {palette['text']}; border: 1px solid {palette['input_border']}; selection-background-color: {palette['surface_alt']}; selection-color: {palette['text']}; }}"
        )
        self.combo_dash_period.addItem(_t("dash_period_last_6", "Son 6 Ay"), "6m")
        self.combo_dash_period.addItem(_t("dash_period_last_12", "Son 12 Ay"), "12m")
        self.combo_dash_period.addItem(_t("dash_period_this_year", "Bu Yıl"), "ytd")
        self.combo_dash_period.addItem(_t("dash_period_all", "Tüm Dönem"), "all")
        self.combo_dash_period.setCurrentIndex(1)
        self.combo_dash_period.currentIndexChanged.connect(lambda _: self.update_dashboard())
        dash_filter_lyt.addWidget(self.lbl_dash_period)
        dash_filter_lyt.addWidget(self.combo_dash_period)
        dash_filter_lyt.addStretch(1)
        dash_lyt.addWidget(dash_filter_w)

        self.gp_dash_chart = QGroupBox(_t("dash_monthly_trend", "Aylık Gelir / Gider Trendi"))
        self.gp_dash_chart.setStyleSheet(
            f"QGroupBox {{ background-color: {palette['surface_bg']}; font-weight: bold; border: 1px solid {palette['group_border']}; border-radius: 8px; margin-top: 15px; }} "
            f"QGroupBox::title {{ subcontrol-origin: margin; subcontrol-position: top center; color: {dash_theme['title']}; padding: 0 6px; }}"
        )
        dash_chart_lyt = QVBoxLayout(self.gp_dash_chart)
        dash_chart_lyt.setContentsMargins(10, 16, 10, 10)
        if MATPLOTLIB_AVAILABLE:
            chart_split_w = QWidget()
            chart_split_w.setStyleSheet("background-color: transparent;")
            chart_split_lyt = QHBoxLayout(chart_split_w)
            chart_split_lyt.setContentsMargins(0, 0, 0, 0)
            chart_split_lyt.setSpacing(10)

            self.dash_trend_figure = Figure(figsize=(6.8, 2.8), dpi=100)
            self.dash_trend_canvas = NonBlockingFigureCanvas(self.dash_trend_figure)
            self.dash_trend_canvas.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            self.dash_trend_canvas.setStyleSheet("background-color: transparent;")

            self.dash_pie_figure = Figure(figsize=(3.4, 2.8), dpi=100)
            self.dash_pie_canvas = NonBlockingFigureCanvas(self.dash_pie_figure)
            self.dash_pie_canvas.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            self.dash_pie_canvas.setStyleSheet("background-color: transparent;")

            chart_split_lyt.addWidget(self.dash_trend_canvas, 2)
            chart_split_lyt.addWidget(self.dash_pie_canvas, 1)
            dash_chart_lyt.addWidget(chart_split_w)
        else:
            self.lbl_dash_chart_fallback = QLabel(_t("dash_chart_missing", "Grafik için matplotlib gerekli. Lütfen matplotlib kurun."))
            self.lbl_dash_chart_fallback.setStyleSheet(f"color: {palette['subtle_text']}; font-size: 12px;")
            self.lbl_dash_chart_fallback.setAlignment(Qt.AlignmentFlag.AlignCenter)
            dash_chart_lyt.addWidget(self.lbl_dash_chart_fallback)
        dash_lyt.addWidget(self.gp_dash_chart)

        # Bottom Tables Left/Right split
        dash_tables_w = QWidget()
        dash_tables_lyt = QHBoxLayout(dash_tables_w)
        dash_tables_lyt.setContentsMargins(0, 0, 0, 0)
        dash_tables_lyt.setSpacing(12)
        dash_lyt.addWidget(dash_tables_w, stretch=1)

        # Left Table (Alerts)
        alerts_col = QWidget()
        alerts_col.setStyleSheet("background-color: transparent;")
        alerts_col_lyt = QVBoxLayout(alerts_col)
        alerts_col_lyt.setContentsMargins(0, 0, 0, 0)
        alerts_col_lyt.setSpacing(4)
        alerts_col_lyt.addWidget(self.lbl_alert_count)
        gp_alerts = QGroupBox(_t("dash_alerts", "Acil Durumlar & Yaklaşan Bakımlar"))
        alerts_panel_bg = "#f8fbff" if current_theme == "light" else dash_theme["alerts_panel_bg"]
        gp_alerts.setStyleSheet(f"QGroupBox {{ background-color: {alerts_panel_bg}; font-weight: bold; border: 1px solid {dash_theme['alerts_border']}; border-radius: 6px; margin-top: 15px;}} QGroupBox::title {{ subcontrol-origin: margin; subcontrol-position: top center; color: {dash_theme['alerts_title']}; padding: 0 6px; }}")
        al_lyt = QVBoxLayout(gp_alerts)
        self.table_alerts = QTableWidget()
        self.table_alerts.setColumnCount(4)
        self.table_alerts.setHorizontalHeaderLabels([_t("plaka", "Plaka"), _t("uyari_tipi", "Uyarı Tipi/İşlem"), _t("uyari", "Kalan"), _t("buton", "")])
        self.table_alerts.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.table_alerts.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table_alerts.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.table_alerts.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.table_alerts.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table_alerts.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_alerts.setAlternatingRowColors(True)
        self.table_alerts.setSortingEnabled(True)
        self.table_alerts.verticalHeader().setVisible(False)
        alerts_table_bg = "#ffffff" if current_theme == "light" else "transparent"
        alerts_row_bg = "#ffffff" if current_theme == "light" else "transparent"
        alerts_row_alt_bg = "#f8fbff" if current_theme == "light" else "transparent"
        self.table_alerts.setStyleSheet(
            f"QTableWidget {{ background-color: {alerts_table_bg}; border: none; color: {palette['text']}; }} "
            f"QTableWidget::item {{ background-color: {alerts_row_bg}; color: {palette['text']}; }} "
            f"QTableWidget::item:alternate {{ background-color: {alerts_row_alt_bg}; }} "
            f"QTableWidget::item:selected {{ background-color: {dash_theme['table_selected_bg']}; color: {dash_theme['table_selected_text']}; }} "
            f"QHeaderView::section {{ background-color: {dash_theme['alerts_header_bg']}; color: {dash_theme['alerts_header_text']}; }}"
        )
        al_lyt.addWidget(self.table_alerts)
        alerts_col_lyt.addWidget(gp_alerts)
        dash_tables_lyt.addWidget(alerts_col, stretch=1)
        events_col = QWidget()
        events_col.setStyleSheet("background-color: transparent;")
        events_col_lyt = QVBoxLayout(events_col)
        events_col_lyt.setContentsMargins(0, 0, 0, 0)
        events_col_lyt.setSpacing(4)
        events_col_lyt.addWidget(self.lbl_event_count)
        gp_events = QGroupBox(_t("dash_events", "Son Hareketler (İşlemler & Adımlar)"))
        events_panel_bg = "#f8fbff" if current_theme == "light" else "transparent"
        gp_events.setStyleSheet(f"QGroupBox {{ background-color: {events_panel_bg}; font-weight: bold; border: 1px solid {dash_theme['events_border']}; border-radius: 6px; margin-top: 15px;}} QGroupBox::title {{ subcontrol-origin: margin; subcontrol-position: top center; color: {dash_theme['events_title']}; }}")
        ev_lyt = QVBoxLayout(gp_events)
        self.table_events = QTableWidget()
        self.table_events.setColumnCount(4)
        self.table_events.setHorizontalHeaderLabels([_t("tarih", "Tarih"), _t("plaka2", "İlgili Araç/Fatura"), _t("olay", "Açıklama"), _t("tutar", "Tutar")])
        self.table_events.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.table_events.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.table_events.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.table_events.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.table_events.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table_events.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_events.setAlternatingRowColors(True)
        self.table_events.setSortingEnabled(True)
        self.table_events.verticalHeader().setVisible(False)
        events_table_bg = "#ffffff" if current_theme == "light" else "transparent"
        events_row_bg = "#ffffff" if current_theme == "light" else "transparent"
        events_row_alt_bg = "#f8fbff" if current_theme == "light" else "transparent"
        self.table_events.setStyleSheet(
            f"QTableWidget {{ background-color: {events_table_bg}; border: none; color: {palette['text']}; }} "
            f"QTableWidget::item {{ background-color: {events_row_bg}; color: {palette['text']}; }} "
            f"QTableWidget::item:alternate {{ background-color: {events_row_alt_bg}; }} "
            f"QTableWidget::item:selected {{ background-color: {dash_theme['table_selected_bg']}; color: {dash_theme['table_selected_text']}; }} "
            f"QHeaderView::section {{ background-color: {dash_theme['events_header_bg']}; color: {dash_theme['events_header_text']}; }}"
        )
        ev_lyt.addWidget(self.table_events)
        events_col_lyt.addWidget(gp_events)
        dash_tables_lyt.addWidget(events_col, stretch=1)

        self.stack.addWidget(self.page_dashboard)

        self.stack.addWidget(self.page_fatura)

        # --- ARAÇ YÖNETİMİ PAGE ---
        self.page_arac = QWidget()
        arac_lyt = QVBoxLayout(self.page_arac)
        arac_lyt.setContentsMargins(20, 20, 20, 35)
        
        arac_title = QLabel(_t("qlabel_filo_ve_arac_yonetim", "Araç Yönetim Paneli"))
        arac_title.setFont(get_custom_font("Quicksand", 20, QFont.Weight.DemiBold))
        arac_title.setStyleSheet(" padding: 10px; background-color: transparent;")
        
        arac_title_lyt = QHBoxLayout()
        arac_title_lyt.addWidget(arac_title)
        
        self.arac_title = arac_title
        self.arac_title_lyt = arac_title_lyt

        arac_lyt.addLayout(arac_title_lyt)
        arac_top_bar = QGroupBox(_t("qgroupbox_arac_secimi_ve_islem", "Araç Seçimi ve İşlemler"))
        arac_top_bar.setStyleSheet("QGroupBox { background-color: transparent; }")
        arac_top_lyt = QHBoxLayout(arac_top_bar)
        
        arac_top_lyt.addWidget(QLabel(_t("qlabel_mevcut_arac", "Mevcut Araç:")), 0)
        self.combo_araclar = QComboBox()
        self.combo_araclar.setMinimumWidth(250)
        self.combo_araclar.setMaximumWidth(400) # Çok uzun olmasını engelle
        self.combo_araclar.currentIndexChanged.connect(self.arac_secildi)
        arac_top_lyt.addWidget(self.combo_araclar, 0)
        
        self.btn_yeni_arac = QPushButton(_t("qpushbutton__yeni_arac_ekle", "+ Yeni Araç Ekle"))
        self.btn_yeni_arac.setStyleSheet(build_button_style(palette["success_bg"], palette["white"], palette["success_hover"], padding="6px 12px"))
        self.btn_yeni_arac.clicked.connect(self.yeni_arac_ekle)
        
        self.btn_duzenle_arac = QPushButton(_t("qpushbutton_arac_bilgilerini_duz", "Araç Bilgilerini Düzenle"))
        self.btn_duzenle_arac.setStyleSheet(build_button_style(palette["info_bg"], palette["white"], palette["info_hover"], padding="6px 12px"))
        self.btn_duzenle_arac.clicked.connect(self.arac_duzenle)
        
        self.btn_sil_arac = QPushButton(_t("qpushbutton_araci_sil", "Aracı Sil"))
        self.btn_sil_arac.setStyleSheet(build_button_style(palette["danger_bg"], palette["white"], palette["danger_hover"], padding="6px 12px"))
        self.btn_sil_arac.clicked.connect(self.arac_sil)
        
        # EXPORT BUTTONS FOR VEHICLES
        self.btn_excel_arac = QPushButton(_t("qpushbutton_araclari_excele_akta", "Araçları Excel'e Aktar"))
        self.btn_excel_arac.setStyleSheet(build_button_style(palette["emerald_bg"], palette["white"], palette["success_hover"], padding="6px 12px", extra="margin-left: 15px;"))
        self.btn_excel_arac.clicked.connect(self.export_arac_excel)
        
        self.btn_pdf_arac = QPushButton(_t("qpushbutton_araclari_pdfe_aktar", "Araçları PDF'e Aktar"))
        self.btn_pdf_arac.setStyleSheet(build_button_style(palette["rose_bg"], palette["white"], palette["danger_hover"], padding="6px 12px"))
        self.btn_pdf_arac.clicked.connect(self.export_arac_pdf)
        
        arac_top_lyt.addWidget(self.btn_yeni_arac)
        arac_top_lyt.addWidget(self.btn_duzenle_arac)
        arac_top_lyt.addWidget(self.btn_sil_arac)
        arac_top_lyt.addWidget(self.btn_excel_arac)
        arac_top_lyt.addWidget(self.btn_pdf_arac)
        arac_top_lyt.addStretch(1) # Bütün butonlar yan yana geldikten sonra geri kalan boşluğu sağa iterek kapat
        
        arac_lyt.addWidget(arac_top_bar)
        
        # Main Layout split: Left (Car Details), Right (Tabs for Logs)
        arac_split_lyt = QHBoxLayout()
        
        # Left Details
        group_arac_detay = QGroupBox(_t("qgroupbox_genel_arac_bilgileri", "Genel Araç Bilgileri"))
        form_arac_detay = QFormLayout(group_arac_detay)
        form_arac_detay.setSpacing(10)
        
        self.lbl_arac_plaka = QLabel(_t("qlabel_secilmedi", "Seçilmedi"))
        self.lbl_arac_firma = QLabel("-")
        self.lbl_arac_kurum = QLabel("-")
        self.lbl_arac_vergi_no = QLabel("-")
        self.lbl_arac_sofor = QLabel("-")
        self.lbl_arac_sofor_tel = QLabel("-")
        self.lbl_arac_marka = QLabel("-")
        self.lbl_arac_model = QLabel("-")
        self.lbl_arac_yil = QLabel("-")
        self.lbl_arac_yakit = QLabel("-")
        self.lbl_arac_vites = QLabel("-")
        self.lbl_arac_km = QLabel(_t("qlabel_0_km", "0 km"))
        self.lbl_arac_sasi = QLabel("-")
        self.lbl_arac_utts = QLabel("-")
        self.lbl_arac_muayene = QLabel("-")
        self.lbl_arac_trafik = QLabel("-")
        self.lbl_arac_kasko = QLabel("-")
        self.lbl_arac_koltuk = QLabel("-")
        self.lbl_arac_ruhsat = QLabel("-")
        
        for lbl in (self.lbl_arac_plaka, self.lbl_arac_firma, self.lbl_arac_kurum, self.lbl_arac_vergi_no, self.lbl_arac_sofor, self.lbl_arac_sofor_tel, self.lbl_arac_marka, self.lbl_arac_model, self.lbl_arac_yil, self.lbl_arac_yakit, self.lbl_arac_vites, self.lbl_arac_km, self.lbl_arac_sasi,
                    self.lbl_arac_utts, self.lbl_arac_muayene, self.lbl_arac_trafik, self.lbl_arac_kasko, self.lbl_arac_koltuk, self.lbl_arac_ruhsat):
            lbl.setStyleSheet(" font-weight: bold; background-color: transparent;")
        
        form_arac_detay.addRow(_t("plaka_lbl", "Plaka:"), self.lbl_arac_plaka)
        form_arac_detay.addRow(_t("firma_lbl", "Araç Sahibi (Firma):"), self.lbl_arac_firma)
        form_arac_detay.addRow(_t("kurum_lbl", "Çalıştığı Kurum:"), self.lbl_arac_kurum)
        form_arac_detay.addRow(_t("vergi_tc_lbl", "Kimlik/Vergi No:"), self.lbl_arac_vergi_no)
        form_arac_detay.addRow(_t("sofor_lbl", "Araç Şoförü:"), self.lbl_arac_sofor)
        form_arac_detay.addRow(_t("sofor_tel_lbl", "Şoför Tel:"), self.lbl_arac_sofor_tel)
        form_arac_detay.addRow(_t("marka_lbl", "Marka:"), self.lbl_arac_marka)
        form_arac_detay.addRow(_t("model_lbl", "Model:"), self.lbl_arac_model)
        form_arac_detay.addRow(_t("model_yili_lbl", "Model Yılı:"), self.lbl_arac_yil)
        form_arac_detay.addRow(_t("yakit_tipi_lbl", "Yakıt Tipi:"), self.lbl_arac_yakit)
        form_arac_detay.addRow(_t("vites_tipi_lbl", "Vites Tipi:"), self.lbl_arac_vites)
        form_arac_detay.addRow(_t("guncel_km_lbl", "Güncel KM:"), self.lbl_arac_km)
        form_arac_detay.addRow(_t("sasi_no_lbl", "Şasi No:"), self.lbl_arac_sasi)
        form_arac_detay.addRow(_t("utts_durumu_lbl", "UTTS Durumu:"), self.lbl_arac_utts)
        form_arac_detay.addRow(_t("muayene_bitis_lbl", "Muayene Bitiş Tarihi:"), self.lbl_arac_muayene)
        form_arac_detay.addRow(_t("trafik_bitis_lbl", "Trafik Sigortası Bitiş Tarihi:"), self.lbl_arac_trafik)
        form_arac_detay.addRow(_t("kasko_bitis_lbl", "Kasko Bitiş Tarihi:"), self.lbl_arac_kasko)
        form_arac_detay.addRow(_t("koltuk_sigortasi_lbl", "Koltuk Sig. Bitiş Tarihi:"), self.lbl_arac_koltuk)
        form_arac_detay.addRow(_t("ruhsat_belge_no_lbl", "Ruhsat Belge No:"), self.lbl_arac_ruhsat)
        
        # --- EVRAKLAR (ARAÇ VE ŞOFÖR BÖLMESİ) ---
        evrak_lyt = QVBoxLayout()
        btn_evrak_lyt = QHBoxLayout()
        
        self.btn_arac_evraklar = QPushButton(_t("arac_evrak_yükle_btn", "Araç Evrakları(Yükle/İncele)"))
        self.btn_arac_evraklar.setStyleSheet(build_button_style(palette["indigo_bg"], palette["white"], palette["indigo_hover"], padding="4px", border_radius=4, font_size="10px"))
        self.btn_arac_evraklar.setFixedHeight(28)
        self.btn_arac_evraklar.clicked.connect(lambda: self.evrak_yonetimi_ac("arac"))
        self.btn_sofor_evraklar = QPushButton(_t("sofor_evrak_yükle_btn", "Şoför Evrakları(Yükle/İncele)"))
        self.btn_sofor_evraklar.setStyleSheet(build_button_style(palette["warning_bg"], palette["white"], palette["warning_hover"], padding="4px", border_radius=4, font_size="10px"))
        self.btn_sofor_evraklar.setFixedHeight(28)
        self.btn_sofor_evraklar.clicked.connect(lambda: self.evrak_yonetimi_ac("sofor"))

        btn_evrak_lyt.addWidget(self.btn_arac_evraklar)
        btn_evrak_lyt.addWidget(self.btn_sofor_evraklar)
        
        self.lbl_evrak_ozet = QLabel(_t("arac_secilmedi", "Henüz araç seçilmedi."))
        self.lbl_evrak_ozet.setStyleSheet(f"color: {palette['subtle_text']}; font-size: 11px;")
        self.lbl_evrak_ozet.setWordWrap(True)
        
        evrak_lyt.addLayout(btn_evrak_lyt)
        evrak_lyt.addWidget(self.lbl_evrak_ozet)
        
        form_arac_detay.addRow(f"<b>{_t('evrak_yonetimi_baslik', 'Evrak Yönetimi')}:</b>", evrak_lyt)
        
        arac_split_lyt.addWidget(group_arac_detay, stretch=1)
        
        # Right Tabs
        self.tabs_arac_islem = QTabWidget()
        
        self.arac_tab_yakit = QWidget()
        self.arac_tab_bakim = QWidget()
        self.layout_yakit = QVBoxLayout(self.arac_tab_yakit)
        
        self.yakit_table = QTableWidget(0, 9)
        self.yakit_table.setHorizontalHeaderLabels([
            _t("tarih_km", "Değişim Tarihi & KM"), _t("yag_cinsi", "Kullanılan Yağ Cinsi"), _t("yag_lt", "Yağ lt."), _t("yag_filtre", "Yağ Filtresi"), 
            _t("mazot_filtre", "Mazot Filtresi"), _t("hava_filtre", "Hava Filt."), _t("degisim_usta", "Değişimi Yapan Usta"), _t("isi_yaptiran", "İşi Yaptıran"), _t("not_hdr", "Not")
        ])
        
        self.yakit_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.yakit_table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeMode.Stretch)
        self.yakit_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.yakit_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.layout_yakit.addWidget(self.yakit_table)
        
        self.yakit_alt_lyt = QHBoxLayout()
        self.btn_yakit_ekle = QPushButton(_t("yeni_icerik_ekle", "+ Ekle"))
        self.btn_yakit_ekle.setStyleSheet(build_button_style(palette["success_bg"], palette["white"], palette["success_hover"], padding="5px"))
        self.btn_yakit_ekle.clicked.connect(lambda: self.arac_islem_ekle('yakit'))

        self.btn_yakit_duzenle = QPushButton(_t("qpushbutton_duzenle", "✎ Düzenle"))
        self.btn_yakit_duzenle.setStyleSheet(build_button_style(palette["info_bg"], palette["white"], palette["info_hover"], padding="5px"))
        self.btn_yakit_duzenle.clicked.connect(lambda: self.arac_secili_islem_tetikle('yakit', 'duzenle'))

        self.btn_yakit_sil = QPushButton(_t("qpushbutton_sil", "X Sil"))
        self.btn_yakit_sil.setStyleSheet(build_button_style(palette["danger_hover"], palette["white"], palette["danger_bg"], padding="5px"))
        self.btn_yakit_sil.clicked.connect(lambda: self.arac_secili_islem_tetikle('yakit', 'sil'))
        
        self.yakit_alt_lyt.addWidget(self.btn_yakit_ekle, 0, Qt.AlignmentFlag.AlignLeft)
        self.yakit_alt_lyt.addWidget(self.btn_yakit_duzenle, 0, Qt.AlignmentFlag.AlignLeft)
        self.yakit_alt_lyt.addWidget(self.btn_yakit_sil, 0, Qt.AlignmentFlag.AlignLeft)
        self.yakit_alt_lyt.addStretch(1)
        self.layout_yakit.addLayout(self.yakit_alt_lyt)
        
        # Bakım Sekmesi İçeriği
        self.layout_bakim = QVBoxLayout(self.arac_tab_bakim)
        
        self.bakim_table = QTableWidget(0, 4)
        self.bakim_table.setHorizontalHeaderLabels([_t("tarih_hdr", "Tarih"), _t("aciklama_hdr", "Açıklama"), _t("tutar_hdr", "Tutar"), _t("iscilik_hdr", "İşçilik")])
        self.bakim_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.bakim_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.bakim_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.layout_bakim.addWidget(self.bakim_table)
        
        self.bakim_alt_lyt = QHBoxLayout()
        self.btn_bakim_ekle = QPushButton(_t("yeni_icerik_ekle", "+ Ekle"))
        self.btn_bakim_ekle.setStyleSheet(build_button_style(palette["success_bg"], palette["white"], palette["success_hover"], padding="5px"))
        self.btn_bakim_ekle.clicked.connect(lambda: self.arac_islem_ekle('bakim'))

        self.btn_bakim_duzenle = QPushButton(_t("qpushbutton_duzenle", "✎ Düzenle"))
        self.btn_bakim_duzenle.setStyleSheet(build_button_style(palette["info_bg"], palette["white"], palette["info_hover"], padding="5px"))
        self.btn_bakim_duzenle.clicked.connect(lambda: self.arac_secili_islem_tetikle('bakim', 'duzenle'))

        self.btn_bakim_sil = QPushButton(_t("qpushbutton_sil", "X Sil"))
        self.btn_bakim_sil.setStyleSheet(build_button_style(palette["danger_hover"], palette["white"], palette["danger_bg"], padding="5px"))
        self.btn_bakim_sil.clicked.connect(lambda: self.arac_secili_islem_tetikle('bakim', 'sil'))

        self.lbl_bakim_toplam = QLabel(_t("toplam_lbl", "Toplam:") + " 0,00 TL")
        self.lbl_bakim_toplam.setFont(get_custom_font("Quicksand", 10, QFont.Weight.Bold))
        self.lbl_bakim_toplam.setAlignment(Qt.AlignmentFlag.AlignRight)
        
        self.bakim_alt_lyt.addWidget(self.btn_bakim_ekle, 0, Qt.AlignmentFlag.AlignLeft)
        self.bakim_alt_lyt.addWidget(self.btn_bakim_duzenle, 0, Qt.AlignmentFlag.AlignLeft)
        self.bakim_alt_lyt.addWidget(self.btn_bakim_sil, 0, Qt.AlignmentFlag.AlignLeft)
        self.bakim_alt_lyt.addStretch(1)
        self.bakim_alt_lyt.addWidget(self.lbl_bakim_toplam, 0, Qt.AlignmentFlag.AlignRight)
        self.layout_bakim.addLayout(self.bakim_alt_lyt)
            
        self.tabs_arac_islem.addTab(self.arac_tab_yakit, _t("tab_yakt_kaytlar", "Yakıt Kayıtları"))
        self.tabs_arac_islem.addTab(self.arac_tab_bakim, _t("tab_bakm__servis", "Bakım"))
        
        arac_split_lyt.addWidget(self.tabs_arac_islem, stretch=2)
        
        arac_lyt.addLayout(arac_split_lyt)
        
        self.stack.addWidget(self.page_arac)

        # --- AYARLAR PAGE ---
        self.page_ayarlar = QWidget()
        ayarlar_lyt = QVBoxLayout(self.page_ayarlar)
        ayarlar_lyt.setContentsMargins(40, 40, 40, 40)
        
        self.lbl_ayarlar_title = QLabel(_t("settings_title", "Ayarlar"))
        self.lbl_ayarlar_title.setFont(get_custom_font("Quicksand", 24, QFont.Weight.DemiBold))
        self.lbl_ayarlar_title.setStyleSheet(f"color: {palette['settings_title']}; background-color: transparent;")
        ayarlar_lyt.addWidget(self.lbl_ayarlar_title)
        
        self.ayarlar_group = QGroupBox(_t("settings_basic", "Temel Yapılandırma"))
        
        ayarlar_h_lyt = QHBoxLayout(self.ayarlar_group)
        ayarlar_form = QFormLayout()
        ayarlar_form.setSpacing(20)
        
        self.combo_tema = QComboBox()
        self.combo_tema.setFixedWidth(250)
        self.combo_tema.setFixedHeight(35)
        self.combo_tema.addItem(_t("theme_dark", "Koyu Tema"), "dark")
        self.combo_tema.addItem(_t("theme_light", "Açık Tema"), "light")
        idx_t = self.combo_tema.findData(app_config.get("theme", "dark"))
        if idx_t >= 0: self.combo_tema.setCurrentIndex(idx_t)
        
        self.lbl_settings_theme = QLabel(_t("settings_theme", "Uygulama Teması:"))
        self.lbl_settings_theme.setStyleSheet("background-color: transparent; font-size: 14px;")
        ayarlar_form.addRow(self.lbl_settings_theme, self.combo_tema)
        
        self.combo_dil = QComboBox()
        self.combo_dil.setFixedWidth(250)
        self.combo_dil.setFixedHeight(35)
        self.combo_dil.addItem(_t("lang_tr", "Türkçe"), "tr")
        self.combo_dil.addItem(_t("lang_en", "English"), "en")
        idx_l = self.combo_dil.findData(app_config.get("lang", "tr"))
        if idx_l >= 0: self.combo_dil.setCurrentIndex(idx_l)
        
        self.lbl_settings_lang = QLabel(_t("settings_lang", "Uygulama Dili:"))
        self.lbl_settings_lang.setStyleSheet("background-color: transparent; font-size: 14px;")
        ayarlar_form.addRow(self.lbl_settings_lang, self.combo_dil)

        self.in_update_repo = QLineEdit(app_config.get("update_repo", ""))
        self.in_update_repo.setFixedWidth(250)
        self.in_update_repo.setFixedHeight(35)
        self.in_update_repo.setPlaceholderText(_t("settings_update_repo_ph", "ornek: kullanici/proje"))
        self.lbl_settings_update_repo = QLabel(_t("settings_update_repo", "GitHub Repo (owner/repo):"))
        self.lbl_settings_update_repo.setStyleSheet("background-color: transparent; font-size: 14px;")
        ayarlar_form.addRow(self.lbl_settings_update_repo, self.in_update_repo)

        self.in_update_asset = QLineEdit(app_config.get("update_asset_keyword", "fatura"))
        self.in_update_asset.setFixedWidth(250)
        self.in_update_asset.setFixedHeight(35)
        self.in_update_asset.setPlaceholderText(_t("settings_update_asset_ph", "exe dosya adinda gececek anahtar"))
        self.lbl_settings_update_asset = QLabel(_t("settings_update_asset", "Guncelleme Dosya Anahtari:"))
        self.lbl_settings_update_asset.setStyleSheet("background-color: transparent; font-size: 14px;")
        ayarlar_form.addRow(self.lbl_settings_update_asset, self.in_update_asset)

        self.chk_update_auto = QCheckBox(_t("settings_update_auto", "12 saatte bir otomatik kontrol et"))
        self.chk_update_auto.setChecked(bool(app_config.get("update_enabled", True)))
        ayarlar_form.addRow("", self.chk_update_auto)

        self.btn_update_check = QPushButton(_t("settings_update_check", "Guncellemeleri Kontrol Et"))
        self.btn_update_check.setFixedWidth(250)
        self.btn_update_check.setFixedHeight(38)
        self.btn_update_check.setStyleSheet(build_button_style(palette["info_bg"], palette["white"], palette["info_hover"], padding="6px 12px"))
        self.btn_update_check.clicked.connect(lambda: self.check_for_updates(show_no_update=True, force=True))
        ayarlar_form.addRow("", self.btn_update_check)
        
        btn_lyt = QHBoxLayout()
        self.btn_ayarlar_kaydet = QPushButton(_t("settings_save", "Ayarları Kaydet"))
        self.btn_ayarlar_kaydet.setFixedWidth(250)
        self.btn_ayarlar_kaydet.setFixedHeight(40)
        self.btn_ayarlar_kaydet.setStyleSheet(build_button_style(palette["success_bg"], palette["white"], palette["success_hover"], padding="6px 12px"))
        self.btn_ayarlar_kaydet.clicked.connect(self.save_app_settings_gui)
        
        self.lbl_settings_feedback = QLabel("")
        self.lbl_settings_feedback.setStyleSheet(f"color: {palette['positive_text']}; font-weight: bold; background-color: transparent;")
        
        btn_lyt.addWidget(self.btn_ayarlar_kaydet)
        btn_lyt.addWidget(self.lbl_settings_feedback)
        btn_lyt.addStretch()
        
        ayarlar_form.addRow("", btn_lyt)
        
        ayarlar_h_lyt.addLayout(ayarlar_form)
        ayarlar_h_lyt.addStretch() # Prevents stretching full width
        
        ayarlar_lyt.addWidget(self.ayarlar_group)
        ayarlar_lyt.addStretch()
        
        self.stack.addWidget(self.page_ayarlar)

        # Footer applies to the entire app now (centered along the very bottom of the screen)
        current_year = datetime.now().year
        year_str = f"2026 - {current_year}"
        self.lbl_footer = QLabel(f"© {_t('owner_name', 'Ümit Arik')} {year_str}")
        self.lbl_footer.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_footer.setStyleSheet(" font-size: 14px; font-weight: bold; background- padding: 12px; ")
        self.main_vbox.addWidget(self.lbl_footer)

        self.update_ui()
        self.reload_arac_combo() # Araçları yükle
        self.switch_page(0) # Select first page by default

    def toggle_sidebar(self):
        if getattr(self, "_sidebar_animating", False):
            return

        width = self.sidebar_frame.width()
        new_width = 260 if width == 110 else 110
        is_expanding = new_width > width
        self._sidebar_animating = True
        self.btn_hamburger.toggle(new_width == 260)

        # Keep sidebar margins stable to avoid a second mini-layout jump.
        self.sidebar_layout.setContentsMargins(8, 8, 8, 8)

        if is_expanding:
            self.lbl_version.show()
            for btn in (self.btn_menu_dashboard, self.btn_menu_fatura, self.btn_menu_arac, self.btn_menu_ayarlar):
                btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon)
                btn.setIconSize(QSize(96, 96))
                btn.setFixedHeight(176)
            self.btn_menu_dashboard.setText(_t("menu_dashboard", "Ana Sayfa\nPaneli"))
            self.btn_menu_fatura.setText(_t("menu_fatura", "Fatura\nYönetimi"))
            self.btn_menu_arac.setText(_t("menu_arac", "Araç\nYönetimi"))
            self.btn_menu_ayarlar.setText(_t("menu_ayarlar", "Ayarlar"))
        else:
            # Apply collapsed visual mode before animation so there is no second jump at the end.
            self.lbl_version.hide()
            for btn in (self.btn_menu_dashboard, self.btn_menu_fatura, self.btn_menu_arac, self.btn_menu_ayarlar):
                btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonIconOnly)
                btn.setIconSize(QSize(62, 62))
                btn.setFixedHeight(94)
            self.btn_menu_dashboard.setText("")
            self.btn_menu_fatura.setText("")
            self.btn_menu_arac.setText("")
            self.btn_menu_ayarlar.setText("")

        # Single width animation is smoother than animating min/max in parallel.
        self.sidebar_width_anim = QVariantAnimation(self)
        self.sidebar_width_anim.setDuration(360)
        self.sidebar_width_anim.setStartValue(width)
        self.sidebar_width_anim.setEndValue(new_width)
        self.sidebar_width_anim.setEasingCurve(QEasingCurve.Type.InOutSine)

        def _on_sidebar_width(value):
            w = int(value)
            self.sidebar_frame.setMinimumWidth(w)
            self.sidebar_frame.setMaximumWidth(w)

        self.sidebar_width_anim.valueChanged.connect(_on_sidebar_width)

        def _after_sidebar_anim():
            # Re-assert active button style after layout/style changes.
            self.switch_page(self.stack.currentIndex())
            self._sidebar_animating = False

        self.sidebar_width_anim.finished.connect(_after_sidebar_anim)
        self.sidebar_width_anim.start()

    def switch_page(self, index: int):
        self.stack.setCurrentIndex(index)
        self.btn_menu_dashboard.setChecked(index == 0)
        self.btn_menu_fatura.setChecked(index == 1)
        self.btn_menu_arac.setChecked(index == 2)
        self.btn_menu_ayarlar.setChecked(index == 3)

    def apply_theme(self, theme_name):
        self.setStyleSheet(get_stylesheet(theme_name))
    def retranslate_ui(self):
        self.btn_menu_fatura.setText(_t("menu_fatura", "Fatura\nYönetimi"))
        self.btn_menu_arac.setText(_t("menu_arac", "Araç\nYönetimi"))
        self.btn_menu_ayarlar.setText(_t("menu_ayarlar", "Ayarlar"))
        
        self.ozet_gb.setTitle(_t("filo_istatistikleri", "G İstatistikleri"))
        
        self.lbl_ayarlar_title.setText(_t("settings_title", "Ayarlar"))
        self.ayarlar_group.setTitle(_t("settings_basic", "Temel Yapılandırma"))
        self.lbl_settings_theme.setText(_t("settings_theme", "Uygulama Teması:"))
        self.lbl_settings_lang.setText(_t("settings_lang", "Uygulama Dili:"))
        self.btn_ayarlar_kaydet.setText(_t("settings_save", "Ayarları Kaydet"))
        self.combo_tema.setItemText(0, _t("theme_dark", "Koyu Tema"))
        self.combo_tema.setItemText(1, _t("theme_light", "Açık Tema"))
        self.combo_dil.setItemText(0, _t("lang_tr", "Türkçe"))
        self.combo_dil.setItemText(1, _t("lang_en", "English"))
        if hasattr(self, "lbl_settings_update_repo"):
            self.lbl_settings_update_repo.setText(_t("settings_update_repo", "GitHub Repo (owner/repo):"))
        if hasattr(self, "lbl_settings_update_asset"):
            self.lbl_settings_update_asset.setText(_t("settings_update_asset", "Guncelleme Dosya Anahtari:"))
        if hasattr(self, "chk_update_auto"):
            self.chk_update_auto.setText(_t("settings_update_auto", "12 saatte bir otomatik kontrol et"))
        if hasattr(self, "btn_update_check"):
            self.btn_update_check.setText(_t("settings_update_check", "Guncellemeleri Kontrol Et"))
        self.ozet_guncelle()

    def save_app_settings_gui(self):
        new_lang = self.combo_dil.currentData()
        new_theme = self.combo_tema.currentData()
        was_sidebar_collapsed = hasattr(self, "sidebar_frame") and self.sidebar_frame.width() <= 110
        
        app_config["lang"] = new_lang
        app_config["theme"] = new_theme
        if hasattr(self, "in_update_repo"):
            app_config["update_repo"] = self.in_update_repo.text().strip()
        if hasattr(self, "in_update_asset"):
            app_config["update_asset_keyword"] = self.in_update_asset.text().strip() or "fatura"
        if hasattr(self, "chk_update_auto"):
            app_config["update_enabled"] = self.chk_update_auto.isChecked()
        save_app_settings()
        # Update language data in memory instantly
        global locale_data
        try:
            with open(str(APP_LANG_DIR / f"{app_config['lang']}.json"), "r", encoding="utf-8") as f:
                locale_data = json.load(f)
        except Exception:
            locale_data = {}
            
        current_idx = self.stack.currentIndex()
        if hasattr(self, "main_widget"):
            self.main_widget.deleteLater()
        
        self.init_ui(is_reload=True)

        # Preserve sidebar mode after UI rebuild (e.g. language/theme change).
        if was_sidebar_collapsed and hasattr(self, "sidebar_frame"):
            self.sidebar_frame.setMinimumWidth(110)
            self.sidebar_frame.setMaximumWidth(110)
            self.sidebar_layout.setContentsMargins(8, 8, 8, 8)
            self.lbl_version.hide()
            self.btn_hamburger.toggle(False)
            for btn in (self.btn_menu_dashboard, self.btn_menu_fatura, self.btn_menu_arac, self.btn_menu_ayarlar):
                btn.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonIconOnly)
                btn.setIconSize(QSize(62, 62))
                btn.setFixedHeight(94)
            self.btn_menu_dashboard.setText("")
            self.btn_menu_fatura.setText("")
            self.btn_menu_arac.setText("")
            self.btn_menu_ayarlar.setText("")

        self.switch_page(current_idx)
        self.configure_update_timer()
        
        # Saniyelik geri bildirim (No popup, label instead)
        if hasattr(self, "lbl_settings_feedback"):
            self.lbl_settings_feedback.setText(_t("settings_saved", "✔ Ayarlar kaydedildi!"))
            # Clear text after 3 seconds
            QTimer.singleShot(3000, lambda: self.lbl_settings_feedback.setText(""))

    def _auto_update_check_tick(self):
        self.check_for_updates(show_no_update=False, force=False)

    def configure_update_timer(self):
        if not hasattr(self, "update_check_timer"):
            return
        self.update_check_timer.stop()
        if not bool(app_config.get("update_enabled", True)):
            return
        hours = int(app_config.get("update_interval_hours", 12) or 12)
        hours = max(1, hours)
        self.update_check_timer.start(hours * 60 * 60 * 1000)

    def maybe_check_for_updates_on_startup(self):
        if not bool(app_config.get("update_enabled", True)):
            return
        last_check = app_config.get("update_last_check", "")
        if not last_check:
            self.check_for_updates(show_no_update=False, force=False)
            return
        try:
            last_dt = datetime.fromisoformat(last_check)
        except Exception:
            self.check_for_updates(show_no_update=False, force=False)
            return
        hours = int(app_config.get("update_interval_hours", 12) or 12)
        elapsed = (datetime.now() - last_dt).total_seconds() / 3600.0
        if elapsed >= hours:
            self.check_for_updates(show_no_update=False, force=False)

    def _version_tuple(self, raw_version: str):
        clean = re.sub(r"[^0-9.]", "", str(raw_version or ""))
        parts = [int(x) for x in clean.split(".") if x.isdigit()]
        while len(parts) < 3:
            parts.append(0)
        return tuple(parts[:3])

    def _is_newer_version(self, latest_version: str):
        return self._version_tuple(latest_version) > self._version_tuple(APP_VERSION)

    def _fetch_latest_release(self, repo_full: str):
        api_url = f"https://api.github.com/repos/{repo_full}/releases/latest"
        req = urllib.request.Request(api_url, headers={"User-Agent": "GelirGiderAppUpdater"})
        with urllib.request.urlopen(req, timeout=20) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
        return payload

    def check_for_updates(self, show_no_update=True, force=False):
        repo_full = str(app_config.get("update_repo", "")).strip()
        asset_keyword = str(app_config.get("update_asset_keyword", "fatura")).strip().lower()

        if not repo_full:
            if force:
                QMessageBox.warning(self, _t("msg_uyar", "Uyarı"), _t("update_repo_missing", "Guncelleme kontrolu icin GitHub repo bilgisi gerekli (owner/repo)."))
            return

        try:
            rel = self._fetch_latest_release(repo_full)
            app_config["update_last_check"] = datetime.now().isoformat(timespec="seconds")
            save_app_settings()

            latest_tag = rel.get("tag_name", "")
            assets = rel.get("assets", []) if isinstance(rel.get("assets", []), list) else []

            selected_asset = None
            if assets:
                if asset_keyword:
                    for a in assets:
                        name = str(a.get("name", "")).lower()
                        if asset_keyword in name and (name.endswith(".exe") or name.endswith(".zip")):
                            selected_asset = a
                            break
                if selected_asset is None:
                    for a in assets:
                        name = str(a.get("name", "")).lower()
                        if name.endswith(".exe"):
                            selected_asset = a
                            break
                if selected_asset is None:
                    selected_asset = assets[0]

            if not self._is_newer_version(latest_tag):
                if show_no_update:
                    QMessageBox.information(self, _t("msg_bilgi", "Bilgi"), _t("update_no_new", "Uygulama guncel. Yeni surum bulunamadi."))
                return

            asset_url = selected_asset.get("browser_download_url", "") if selected_asset else ""
            asset_name = selected_asset.get("name", "") if selected_asset else ""
            rel_name = rel.get("name", latest_tag)

            msg = QMessageBox(self)
            msg.setWindowTitle(_t("update_available_title", "Guncelleme Hazir"))
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setText(
                _t("update_available_text", "Yeni surum bulundu!\n\nMevcut: v{0}\nYeni: {1}\n\nSimdi indirip kurmak ister misiniz?").format(APP_VERSION, rel_name)
            )
            btn_download = msg.addButton(_t("update_download_install", "Indir ve Kur"), QMessageBox.ButtonRole.AcceptRole)
            msg.addButton(_t("update_later", "Daha Sonra"), QMessageBox.ButtonRole.RejectRole)
            msg.exec()

            if msg.clickedButton() == btn_download:
                if not asset_url:
                    QMessageBox.warning(self, _t("msg_hata", "Hata"), _t("update_asset_missing", "Guncelleme dosyasi bulunamadi."))
                    return
                self.download_and_apply_update(asset_url, asset_name)

        except urllib.error.URLError as exc:
            if force:
                QMessageBox.warning(self, _t("msg_hata", "Hata"), _t("update_check_failed", "Guncelleme kontrolu basarisiz.") + f"\n{exc}")
        except Exception as exc:
            if force:
                QMessageBox.warning(self, _t("msg_hata", "Hata"), _t("update_check_failed", "Guncelleme kontrolu basarisiz.") + f"\n{exc}")

    def download_and_apply_update(self, asset_url: str, asset_name: str):
        try:
            temp_dir = tempfile.mkdtemp(prefix="gelir_gider_update_")
            safe_name = asset_name or "update_package.exe"
            download_path = os.path.join(temp_dir, safe_name)

            req = urllib.request.Request(asset_url, headers={"User-Agent": "GelirGiderAppUpdater"})
            with urllib.request.urlopen(req, timeout=60) as resp, open(download_path, "wb") as out:
                out.write(resp.read())

            if not getattr(sys, "frozen", False):
                webbrowser.open(asset_url)
                QMessageBox.information(self, _t("msg_bilgi", "Bilgi"), _t("update_manual_install", "Gelisim modunda otomatik degisim yapilmadi. Indirilen surumu manuel kurun."))
                return

            if not download_path.lower().endswith(".exe"):
                webbrowser.open(asset_url)
                QMessageBox.information(self, _t("msg_bilgi", "Bilgi"), _t("update_manual_package", "Guncelleme paketi exe degil. Dosya indirildi, lutfen manuel kurulum yapin."))
                return

            current_exe = os.path.abspath(sys.executable)
            bat_path = os.path.join(temp_dir, "apply_update.bat")
            src_bat = download_path.replace("/", "\\")
            dst_bat = current_exe.replace("/", "\\")

            bat_content = f"""@echo off
setlocal
set "SRC={src_bat}"
set "DST={dst_bat}"

for /L %%i in (1,1,40) do (
  copy /Y "%SRC%" "%DST%" >nul 2>&1
  if not errorlevel 1 goto done
  timeout /t 1 /nobreak >nul
)

:done
start "" "%DST%"
del "%SRC%" >nul 2>&1
(del "%~f0" >nul 2>&1)
"""

            with open(bat_path, "w", encoding="utf-8") as f:
                f.write(bat_content)

            creationflags = 0
            if hasattr(subprocess, "CREATE_NO_WINDOW"):
                creationflags = subprocess.CREATE_NO_WINDOW
            subprocess.Popen(["cmd", "/c", bat_path], creationflags=creationflags)
            QApplication.quit()
        except Exception as exc:
            QMessageBox.warning(self, _t("msg_hata", "Hata"), _t("update_download_failed", "Guncelleme indirilemedi.") + f"\n{exc}")

    # --- ARAÇ YÖNETİMİ METOTLARI ---
    def reload_arac_combo(self):
        curr_id = self.combo_araclar.currentData()
        self.combo_araclar.blockSignals(True)
        self.combo_araclar.clear()
        self.combo_araclar.addItem(_t("lutfen_arac_seciniz", "-- Lütfen Araç Seçiniz --"), None)
        
        firma_counts = {}
        for a in self.araclar:
            self.combo_araclar.addItem(f"{a['plaka']} - {a['marka']} {a['model']}", a['id'])
            f_ad = a.get("firma", "").strip() or _t("belirtilmemis", "Belirtilmemiş")
            firma_counts[f_ad] = firma_counts.get(f_ad, 0) + 1
            
        if curr_id:
            idx = self.combo_araclar.findData(curr_id)
            if idx >= 0:
                self.combo_araclar.setCurrentIndex(idx)

        self.combo_araclar.blockSignals(False)
        self.arac_secildi() # Refresh UI
        toplam_arac = len(self.araclar)
        
        if hasattr(self, 'lbl_arac_ozet'):
            if toplam_arac > 0:
                ozet_parts = [f"{k}: {v} {_t('arac_adeti', 'araç')}" for k, v in firma_counts.items()]
                ozet_str = f"{_t('toplam_arac', 'Toplam Araç')}: {toplam_arac}\n\n" + "\n\n".join(ozet_parts)
            else:
                ozet_str = _t("sistemde_arac_yok", "Sistemde Kayıtlı Araç Bulunmamaktadır")
            self.lbl_arac_ozet.setText(ozet_str)

    def arac_secildi(self):
        arac_id = self.combo_araclar.currentData()
        if not arac_id:
            # Clear UI
            self.lbl_arac_plaka.setText(_t("secilmedi", "Seçilmedi"))
            for lbl in (self.lbl_arac_firma, self.lbl_arac_vergi_no, self.lbl_arac_sofor, self.lbl_arac_sofor_tel, self.lbl_arac_marka, self.lbl_arac_model, self.lbl_arac_yil, self.lbl_arac_yakit, self.lbl_arac_vites, 
                        self.lbl_arac_km, self.lbl_arac_sasi, self.lbl_arac_utts, self.lbl_arac_muayene, self.lbl_arac_trafik, self.lbl_arac_kasko, 
                        self.lbl_arac_koltuk, self.lbl_arac_ruhsat):
                lbl.setText("-")
                lbl.target_color = None
                lbl.setStyleSheet("")
            return

        # Find arac
        a = next((x for x in self.araclar if x['id'] == arac_id), None)
        if a:
            self.lbl_arac_plaka.setText(a.get("plaka", "-"))
            self.lbl_arac_firma.setText(a.get("firma", "-"))
            self.lbl_arac_kurum.setText(a.get("kurum", "-"))
            self.lbl_arac_vergi_no.setText(a.get("vergi_no", "-"))
            self.lbl_arac_sofor.setText(a.get("sofor", "-"))
            self.lbl_arac_sofor_tel.setText(a.get("sofor_tel", "-"))
            self.lbl_arac_marka.setText(a.get("marka", "-"))
            self.lbl_arac_model.setText(a.get("model", "-"))
            self.lbl_arac_yil.setText(a.get("yil", "-"))
            self.lbl_arac_yakit.setText(a.get("yakit", "-"))
            self.lbl_arac_vites.setText(a.get("vites", "-"))
            self.lbl_arac_km.setText(f"{a.get('km', '0')} km")
            self.lbl_arac_sasi.setText(a.get("sasi", "-"))
            self.lbl_arac_utts.setText(a.get("utts", "-"))
            palette = get_theme_palette(app_config.get("theme", "dark"))
            
            def set_date_label(lbl, date_str):
                lbl.setText(date_str)
                lbl.target_color = None # default reset
                if date_str and date_str != "-":
                    d = None
                    try:
                        d = datetime.strptime(date_str, "%d.%m.%Y")
                    except ValueError:
                        try:
                            d = datetime.strptime(date_str, "%Y-%m-%d")
                        except ValueError:
                            pass
                            
                    if d:
                        days_left = (d.date() - datetime.now().date()).days
                        
                        sohbet_balonu = ""
                        
                        if days_left < 0:
                            lbl.target_color = palette["danger_hover"]
                            sohbet_balonu = f"  ({_t('tarih_gecmis', 'Süresi Geçti!')})"
                        elif days_left < 15:
                            lbl.target_color = palette["danger_hover"]
                            sohbet_balonu = f"  ({_t('cok_az_kaldi', 'Çok Az Kaldı!')} - {days_left} " + _t('gun', 'gün') + ")"
                        elif days_left < 30:
                            lbl.target_color = palette["warning_bg"]
                            sohbet_balonu = f"  ({_t('yaklasti', 'Yaklaştı')} - {days_left} " + _t('gun', 'gün') + ")"
                        elif days_left < 40:
                            lbl.target_color = palette["warning_hover"]
                            sohbet_balonu = f"  ({_t('dikkat', 'Dikkat')} - {days_left} " + _t('gun', 'gün') + ")"
                        else:
                            lbl.target_color = "stable" # yeşil sabit dursun
                            sohbet_balonu = f"  ({_t('iyi_durumda', 'Süresi Var')} - {days_left} " + _t('gun', 'gün') + ")"
                            
                        lbl.setText(date_str + sohbet_balonu)
                        lbl.setToolTip(f"{days_left} {_t('gun_kaldi', 'gün kaldı')}")
                
                # Apply initial state immediately so we don't wait for timer
                if lbl.target_color == "stable":
                    lbl.setStyleSheet(f"color: {palette['positive_text']}; font-weight: bold;")
                elif not lbl.target_color:
                    lbl.setStyleSheet("")

            set_date_label(self.lbl_arac_muayene, a.get("muayene_tarihi", "-"))
            set_date_label(self.lbl_arac_trafik, a.get("trafik_tarihi", "-"))
            set_date_label(self.lbl_arac_kasko, a.get("kasko_tarihi", "-"))
            set_date_label(self.lbl_arac_koltuk, a.get("koltuk_sigortasi", "-"))
            
            self.lbl_arac_ruhsat.setText(a.get("ruhsat_no", "-"))
            
        self.arac_islemleri_tablo_guncelle()
        self.guncelle_evrak_kutu() # Mevcut aracın/şoförün evraklarını label'da göster

    def arac_islemleri_tablo_guncelle(self):
        curr_id = self.combo_araclar.currentData()
        if not curr_id:
            self.yakit_table.setRowCount(0)
            self.bakim_table.setRowCount(0)
            self.lbl_bakim_toplam.setText(_t("toplam_lbl", "Toplam:") + " 0,00 TL")
            return
            
        arac = next((a for a in self.araclar if a["id"] == curr_id), {})
                
        def populate_yakit(tbl: QTableWidget, keys_list: list):
            tbl.setRowCount(0)
            for row, item in enumerate(keys_list):
                tbl.insertRow(row)
                tk = item.get("tarih_km", "")
                
                tarih_item = QTableWidgetItem(tk)
                tarih_item.setData(Qt.ItemDataRole.UserRole, item.get("id")) # islem_id yi sakla
                tbl.setItem(row, 0, tarih_item)
                tbl.setItem(row, 1, QTableWidgetItem(item.get("yag_cinsi", "")))
                tbl.setItem(row, 2, QTableWidgetItem(item.get("yag_lt", "")))
                tbl.setItem(row, 3, QTableWidgetItem(item.get("yag_filtresi", "")))
                tbl.setItem(row, 4, QTableWidgetItem(item.get("mazot_filtresi", "")))
                tbl.setItem(row, 5, QTableWidgetItem(item.get("hava_filtresi", "")))
                tbl.setItem(row, 6, QTableWidgetItem(item.get("usta", "")))
                tbl.setItem(row, 7, QTableWidgetItem(item.get("yaptiran", "")))
                tbl.setItem(row, 8, QTableWidgetItem(item.get("notlar", "")))
                
                tbl.setRowHeight(row, 30)

        def populate_bakim(tbl: QTableWidget, lbl_toplam: QLabel, keys_list: list):
            tbl.setRowCount(0)
            toplam_tutar = 0.0
            
            # Tarihe göre sırala
            try:
                keys_list = sorted(keys_list, key=lambda x: datetime.strptime(x.get("tarih", ""), "%d.%m.%Y"))
            except:
                pass

            row = 0
            current_month = None

            for item in keys_list:
                item_date_str = item.get("tarih", "")
                try:
                    dt = datetime.strptime(item_date_str, "%d.%m.%Y")
                    item_month = dt.strftime("%Y-%m")
                except:
                    item_month = None

                separator_bg = QColor(get_theme_palette(app_config.get("theme", "dark"))["surface_alt"])
                separator_bg.setAlpha(50)

                if current_month is not None and item_month != current_month:
                    # Yeni aya geçerken 3 boşluk
                    for _ in range(3):
                        tbl.insertRow(row)
                        for col_idx in range(4):
                            it = QTableWidgetItem("")
                            it.setFlags(Qt.ItemFlag.NoItemFlags) # tıklanamaz
                            it.setBackground(separator_bg)
                            tbl.setItem(row, col_idx, it)
                        row += 1

                current_month = item_month

                tbl.insertRow(row)
                tarih_item = QTableWidgetItem(item_date_str)
                tarih_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                tarih_item.setData(Qt.ItemDataRole.UserRole, item.get("id")) # islem_id yi sakla
                tbl.setItem(row, 0, tarih_item)
                
                ac_item = QTableWidgetItem(item.get("aciklama", ""))
                tbl.setItem(row, 1, ac_item)

                tutar = item.get("tutar", 0.0)
                iscilik = item.get("iscilik", 0.0)
                toplam_tutar += float(tutar) + float(iscilik) 
                tutar_item = QTableWidgetItem(format_number(float(tutar)))
                tutar_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                tbl.setItem(row, 2, tutar_item)

                iscilik_item = QTableWidgetItem(format_number(float(iscilik)))
                iscilik_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                tbl.setItem(row, 3, iscilik_item)
                
                tbl.setRowHeight(row, 30)
                row += 1
                
            tbl.setColumnWidth(0, 100)
            tbl.setColumnWidth(2, 90)
            tbl.setColumnWidth(3, 100)
            lbl_toplam.setText(_t("toplam_lbl", "Toplam:") + f" {format_currency(toplam_tutar)}")

        populate_yakit(self.yakit_table, arac.get("yakit_kayitlari", []))
        populate_bakim(self.bakim_table, self.lbl_bakim_toplam, arac.get("bakim_kayitlari", []))

    def arac_secili_islem_tetikle(self, islem_tipi: str, islem_ne: str):
        curr_id = self.combo_araclar.currentData()
        if not curr_id:
            QMessageBox.warning(self, "Hata", "Lütfen önce bir araç seçin.")
            return

        tbl = self.yakit_table if islem_tipi == "yakit" else self.bakim_table
        items = tbl.selectedItems()
        
        if not items:
            QMessageBox.warning(self, "Hata", "Lütfen listeden bir kayıt seçerek üzerine tıklayın.")
            return
            
        row = items[0].row()
        item_0 = tbl.item(row, 0)
        
        if not item_0: return # Boş satırsa (ayırıcı boşluk vb)

        islem_id = item_0.data(Qt.ItemDataRole.UserRole)
        
        if not islem_id: 
            return # data yoksa muhtemelen yeni ay geçişi için atılmış boş ayırıcı satırdır.

        if islem_ne == "duzenle":
            self.arac_islem_duzenle(curr_id, islem_tipi, islem_id)
        elif islem_ne == "sil":
            self.arac_islem_sil(curr_id, islem_tipi, islem_id)

    def arac_islem_ekle(self, islem_tipi: str):
        curr_id = self.combo_araclar.currentData()
        if not curr_id:
            QMessageBox.warning(self, "Hata", "Lütfen önce bir araç seçin.")
            return
            
        dialog = QDialog(self)
        dialog.setWindowTitle("Yeni Kayıt Ekle")
        dialog.setMinimumWidth(350)
        layout = QFormLayout(dialog)
        
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(dialog.accept)
        btn_box.rejected.connect(dialog.reject)
        
        inputs = {}
        
        if islem_tipi == "yakit":
            dialog.setWindowTitle(_t("yeni_kayit_ekle", "Yeni Kayıt Ekle"))
            inputs['tarih_km'] = QLineEdit()
            inputs['tarih_km'].setPlaceholderText(_t("tarih_km_ph", "Örn: 21.04.2026 / 150.000"))
            inputs['yag_cinsi'] = QLineEdit()
            inputs['yag_lt'] = QLineEdit()
            inputs['yag_filtresi'] = QLineEdit()
            inputs['mazot_filtresi'] = QLineEdit()
            inputs['hava_filtresi'] = QLineEdit()
            inputs['usta'] = QLineEdit()
            inputs['yaptiran'] = QLineEdit()
            inputs['notlar'] = QLineEdit()
            
            layout.addRow(_t("tarih_kmsi_lbl", "Değişim Tarihi / KMsi:"), inputs['tarih_km'])
            layout.addRow(_t("yag_cinsi_lbl", "Kullanılan Yağ Cinsi:"), inputs['yag_cinsi'])
            layout.addRow(_t("yag_lt_lbl", "Yağ lt.:"), inputs['yag_lt'])
            layout.addRow(_t("yag_filtresi_lbl", "Kullanılan Yağ Filtresi:"), inputs['yag_filtresi'])
            layout.addRow(_t("mazot_filtresi_lbl", "Mazot Filtresi:"), inputs['mazot_filtresi'])
            layout.addRow(_t("hava_filtresi_lbl", "Hava Filtresi:"), inputs['hava_filtresi'])
            layout.addRow(_t("usta_lbl", "Değişimi Yapan Usta:"), inputs['usta'])
            layout.addRow(_t("yaptiran_lbl", "İşi Yaptıran:"), inputs['yaptiran'])
            layout.addRow(_t("not_lbl", "Not:"), inputs['notlar'])
        else:
            inputs['tarih'] = QDateEdit(QDate.currentDate())
            inputs['tarih'].setCalendarPopup(True)
            inputs['aciklama'] = QLineEdit()
            
            inputs['tutar'] = QDoubleSpinBox()
            inputs['tutar'].setMaximum(99999999.99)
            inputs['tutar'].setDecimals(2)

            inputs['iscilik'] = QDoubleSpinBox()
            inputs['iscilik'].setMaximum(99999999.99)
            inputs['iscilik'].setDecimals(2)
            
            layout.addRow(_t("tarih_lbl", "Tarih:"), inputs['tarih'])
            layout.addRow(_t("aciklama_nedegisti_lbl", "Açıklama (Ne Değişti?):"), inputs['aciklama'])
            layout.addRow(_t("malzeme_tutari_tl_lbl", "Malzeme Tutarı (TL):"), inputs['tutar'])
            layout.addRow(_t("iscilik_tl_lbl", "İşçilik (TL):"), inputs['iscilik'])
            
        layout.addRow(btn_box)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            for a in self.araclar:
                if a["id"] == curr_id:
                    key = "yakit_kayitlari" if islem_tipi == "yakit" else "bakim_kayitlari"
                    if key not in a:
                        a[key] = []
                    
                    if islem_tipi == "yakit":
                        yeni_islem = {
                            "id": str(uuid.uuid4()),
                            "tarih_km": inputs['tarih_km'].text(),
                            "yag_cinsi": inputs['yag_cinsi'].text(),
                            "yag_lt": inputs['yag_lt'].text(),
                            "yag_filtresi": inputs['yag_filtresi'].text(),
                            "mazot_filtresi": inputs['mazot_filtresi'].text(),
                            "hava_filtresi": inputs['hava_filtresi'].text(),
                            "usta": inputs['usta'].text(),
                            "yaptiran": inputs['yaptiran'].text(),
                            "notlar": inputs['notlar'].text()
                        }
                    else:
                        yeni_islem = {
                            "id": str(uuid.uuid4()),
                            "tarih": inputs['tarih'].date().toString("dd.MM.yyyy"),
                            "aciklama": inputs['aciklama'].text(),
                            "iscilik": inputs['iscilik'].value(),
                            "tutar": inputs['tutar'].value()
                        }
                    a[key].append(yeni_islem)
                    
                    if "olaylar" not in a: a["olaylar"] = []
                    today_str = QDate.currentDate().toString("dd.MM.yyyy")
                    if islem_tipi == "yakit":
                        msg = _t("msg_yeni_yag_bakim", "Araç {} ").format(a.get("plaka", ""))
                        a["olaylar"].append({"tarih": today_str, "mesaj": msg, "tip": "yeni_yag_bakim", "p1": a.get("plaka", "")})
                    else:
                        msg = _t("msg_yeni_servis_bakim", "Araç {} ").format(a.get("plaka", ""))
                        a["olaylar"].append({"tarih": today_str, "mesaj": msg, "tip": "yeni_servis_bakim", "p1": a.get("plaka", "")})
                    a["olaylar"] = a["olaylar"][-10:] # Son 10 olayı tut
                    
                    self.save_records()
                    self.arac_islemleri_tablo_guncelle()
                    self.reload_arac_combo()
                    break

    def arac_islem_duzenle(self, arac_id: str, islem_tipi: str, islem_id: str):
        arac = next((a for a in self.araclar if a["id"] == arac_id), None)
        if not arac: return

        if islem_tipi == "yakit":
            liste = arac.get("yakit_kayitlari", [])
        else:
            liste = arac.get("bakim_kayitlari", [])

        kayit = next((x for x in liste if x.get("id") == islem_id), None)
        if not kayit: return

        dialog = QDialog(self)
        dialog.setWindowTitle(_t("kayit_duzenle", "Kayıt Düzenle"))
        dialog.setMinimumWidth(350)
        layout = QFormLayout(dialog)
        
        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(dialog.accept)
        btn_box.rejected.connect(dialog.reject)

        inputs = {}
        if islem_tipi == "yakit":
            inputs['tarih_km'] = QLineEdit(kayit.get("tarih_km", ""))
            inputs['yag_cinsi'] = QLineEdit(kayit.get("yag_cinsi", ""))
            inputs['yag_lt'] = QLineEdit(kayit.get("yag_lt", ""))
            inputs['yag_filtresi'] = QLineEdit(kayit.get("yag_filtresi", ""))
            inputs['mazot_filtresi'] = QLineEdit(kayit.get("mazot_filtresi", ""))
            inputs['hava_filtresi'] = QLineEdit(kayit.get("hava_filtresi", ""))
            inputs['usta'] = QLineEdit(kayit.get("usta", ""))
            inputs['yaptiran'] = QLineEdit(kayit.get("yaptiran", ""))
            inputs['notlar'] = QLineEdit(kayit.get("notlar", ""))
            
            layout.addRow(_t("tarih_kmsi_lbl", "Değişim Tarihi / KMsi:"), inputs['tarih_km'])
            layout.addRow(_t("yag_cinsi_lbl", "Kullanılan Yağ Cinsi:"), inputs['yag_cinsi'])
            layout.addRow(_t("yag_lt_lbl", "Yağ lt.:"), inputs['yag_lt'])
            layout.addRow(_t("yag_filtresi_lbl", "Kullanılan Yağ Filtresi:"), inputs['yag_filtresi'])
            layout.addRow(_t("mazot_filtresi_lbl", "Mazot Filtresi:"), inputs['mazot_filtresi'])
            layout.addRow(_t("hava_filtresi_lbl", "Hava Filtresi:"), inputs['hava_filtresi'])
            layout.addRow(_t("usta_lbl", "Değişimi Yapan Usta:"), inputs['usta'])
            layout.addRow(_t("yaptiran_lbl", "İşi Yaptıran:"), inputs['yaptiran'])
            layout.addRow(_t("not_lbl", "Not:"), inputs['notlar'])
        else:
            inputs['tarih'] = QDateEdit()
            dt = QDate.fromString(kayit.get("tarih", ""), "dd.MM.yyyy")
            if not dt.isValid(): dt = QDate.fromString(kayit.get("tarih", ""), "yyyy-MM-dd") # Geriye dönük uyumluluk
            if dt.isValid():
                inputs['tarih'].setDate(dt)
            else:
                inputs['tarih'].setDate(QDate.currentDate())
            inputs['tarih'].setCalendarPopup(True)
            
            inputs['aciklama'] = QLineEdit(kayit.get("aciklama", ""))
            
            inputs['tutar'] = QDoubleSpinBox()
            inputs['tutar'].setMaximum(99999999.99)
            inputs['tutar'].setDecimals(2)
            inputs['tutar'].setValue(float(kayit.get("tutar", 0.0)))

            inputs['iscilik'] = QDoubleSpinBox()
            inputs['iscilik'].setMaximum(99999999.99)
            inputs['iscilik'].setDecimals(2)
            inputs['iscilik'].setValue(float(kayit.get("iscilik", 0.0)))
            
            layout.addRow(_t("tarih_lbl", "Tarih:"), inputs['tarih'])
            layout.addRow(_t("aciklama_nedegisti_lbl", "Açıklama (Ne Değişti?):"), inputs['aciklama'])
            layout.addRow(_t("malzeme_tutari_tl_lbl", "Malzeme Tutarı (TL):"), inputs['tutar'])
            layout.addRow(_t("iscilik_tl_lbl", "İşçilik (TL):"), inputs['iscilik'])

        layout.addWidget(btn_box)

        if dialog.exec():
            if islem_tipi == "yakit":
                kayit["tarih_km"] = inputs['tarih_km'].text()
                kayit["yag_cinsi"] = inputs['yag_cinsi'].text()
                kayit["yag_lt"] = inputs['yag_lt'].text()
                kayit["yag_filtresi"] = inputs['yag_filtresi'].text()
                kayit["mazot_filtresi"] = inputs['mazot_filtresi'].text()
                kayit["hava_filtresi"] = inputs['hava_filtresi'].text()
                kayit["usta"] = inputs['usta'].text()
                kayit["yaptiran"] = inputs['yaptiran'].text()
                kayit["notlar"] = inputs['notlar'].text()
            else:
                kayit["tarih"] = inputs['tarih'].date().toString("dd.MM.yyyy")
                kayit["aciklama"] = inputs['aciklama'].text()
                kayit["tutar"] = float(inputs['tutar'].value())
                kayit["iscilik"] = float(inputs['iscilik'].value())

            if "olaylar" not in arac: arac["olaylar"] = []
            today_str = QDate.currentDate().toString("dd.MM.yyyy")
            if islem_tipi == "yakit":
                arac["olaylar"].append({"tarih": today_str, "tip": "guncel_yag_bakim", "p1": arac.get("plaka", "")})
            else:
                arac["olaylar"].append({"tarih": today_str, "tip": "guncel_servis_bakim", "p1": arac.get("plaka", "")})
            
            arac["olaylar"] = arac["olaylar"][-10:] # Son 10 olayı tut

            self.save_records()
            self.arac_islemleri_tablo_guncelle()

    def arac_islem_sil(self, arac_id, tbl_name, islem_id):
        reply = QMessageBox.question(self, "Onay", "Bu kaydı silmek istediğinize emin misiniz?", 
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply == QMessageBox.StandardButton.Yes:
            for a in self.araclar:
                if a["id"] == arac_id:
                    key = "yakit_kayitlari" if tbl_name == "yakit" else "bakim_kayitlari"
                    if key in a:
                        a[key] = [item for item in a[key] if item.get("id") != islem_id]
                        
                        if "olaylar" in a:
                            hedef_tipler = ["yeni_yag_bakim", "guncel_yag_bakim", "yag_yapildi"] if tbl_name == "yakit" else ["yeni_servis_bakim", "guncel_servis_bakim", "genel_yapildi"]
                            for i in range(len(a["olaylar"]) - 1, -1, -1):
                                if a["olaylar"][i].get("tip") in hedef_tipler:
                                    del a["olaylar"][i]
                                    break  
                    self.save_records()
                    self.arac_islemleri_tablo_guncelle()
                    self.reload_arac_combo()
                    break

    def yeni_arac_ekle(self):
        dlg = AracEkleDialog(self)
        if dlg.exec():
            yeni_arac = dlg.get_data()
            self.araclar.append(yeni_arac)
            self.save_records()
            self.reload_arac_combo()
            index = self.combo_araclar.findData(yeni_arac['id'])
            if index >= 0:
                self.combo_araclar.setCurrentIndex(index)

    def evrak_yonetimi_ac(self, mode="arac"):
        arac_id = self.combo_araclar.currentData()
        if not arac_id:
            QMessageBox.warning(self, _t("msg_uyar", "Uyarı"), _t("msg_ltfenevraklar", "Lütfen evraklarını görüntülemek için bir araç seçin."))
            return
            
        arac = next((x for x in self.araclar if x['id'] == arac_id), None)
        if arac:
            if mode == "sofor" and not arac.get("sofor"):
                QMessageBox.warning(self, _t("uyari_baslik", "Uyarı"), _t("sofor_yok_uyari", "Bu araca kayıtlı bir şoför bulunmuyor. Önce araca bir şoför atayın."))
                return
            dlg = EvrakDialog(self, arac, mode=mode)
            dlg.exec()
            self.guncelle_evrak_kutu() # Dosya eklendikten sonra listeyi yenilemek için

    def guncelle_evrak_kutu(self):
        arac_id = self.combo_araclar.currentData()
        if not arac_id:
            self.lbl_evrak_ozet.setText(_t("arac_secilmedi", "Henüz araç seçilmedi."))
            return
            
        arac = next((x for x in self.araclar if x['id'] == arac_id), None)
        if not arac: return

        base_evrak_path = str(APP_EVRAK_DIR)
        
        arac_klasor = os.path.join(base_evrak_path, "Araçlar", arac.get('plaka', ''))
        arac_dosyalar = os.listdir(arac_klasor) if os.path.exists(arac_klasor) else []
        
        sofor_dosyalar = []
        if arac.get('sofor'):
            sofor_klasor = os.path.join(base_evrak_path, "Şoförler", arac.get('sofor', ''))
            sofor_dosyalar = os.listdir(sofor_klasor) if os.path.exists(sofor_klasor) else []

        ozet_metin = f"<b>{_t('kayitli_evraklar_baslik', 'Kayıtlı Evraklar:')}</b><br>"
        if not arac_dosyalar and not sofor_dosyalar:
            palette = get_theme_palette(app_config.get("theme", "dark"))
            ozet_metin += f"<span style='color: {palette['danger_hover']}; font-style: italic;'>{_t('hic_evrak_yok_uyari', 'Kayıtlı hiçbir evrak bulunmuyor.')}</span>"
        else:
            if arac_dosyalar:
                ozet_metin += f"<u>{_t('arac_evrak_text', 'Araç')} ({len(arac_dosyalar)}):</u> " + ", ".join(arac_dosyalar[:3])
                if len(arac_dosyalar) > 3: ozet_metin += "..."
                ozet_metin += "<br>"
            if sofor_dosyalar:
                ozet_metin += f"<u>{_t('sofor_evrak_text', 'Şoför')} ({len(sofor_dosyalar)}):</u> " + ", ".join(sofor_dosyalar[:3])
                if len(sofor_dosyalar) > 3: ozet_metin += "..."

        self.lbl_evrak_ozet.setText(ozet_metin)

    def arac_duzenle(self):
        arac_id = self.combo_araclar.currentData()
        if not arac_id:
            QMessageBox.warning(self, _t("msg_uyar", "Uyarı"), _t("msg_ltfendzenlem", "Lütfen düzenlemek için bir araç seçin."))
            return
            
        mevcut = next((x for x in self.araclar if x['id'] == arac_id), None)
        if mevcut:
            dlg = AracEkleDialog(self, mevcut)
            if dlg.exec():
                guncel_arac = dlg.get_data()
                
                # Değişiklik olaylarını izleyip listeye (max 10 işlem) ekle.
                olaylar = mevcut.get("olaylar", [])
                
                # Check for changes and log them
                q_bugun = QDate.currentDate().toString("dd.MM.yyyy")
                
                eski_sofor = mevcut.get("sofor", "")
                yeni_sofor = guncel_arac.get("sofor", "")
                if eski_sofor != yeni_sofor:
                    olaylar.append({"tarih": q_bugun, "tip": "sofor_degisti", "p1": guncel_arac.get("plaka", ""), "p2": yeni_sofor})
                
                eski_plaka = mevcut.get("plaka", "")
                yeni_plaka = guncel_arac.get("plaka", "")
                if eski_plaka != yeni_plaka:
                    olaylar.append({"tarih": q_bugun, "tip": "plaka_degisti", "p1": eski_plaka, "p2": yeni_plaka})

                eski_yag = mevcut.get("yag_bakim_tarihi", "")
                yeni_yag = guncel_arac.get("yag_bakim_tarihi", "")
                if eski_yag != yeni_yag and yeni_yag not in ["-", "Yok", ""]:
                    yeni_yag_dt = QDate.fromString(yeni_yag, "dd.MM.yyyy")
                    if yeni_yag_dt.isValid() and 0 <= yeni_yag_dt.daysTo(QDate.currentDate()) <= 7:
                        olaylar.append({"tarih": q_bugun, "tip": "yag_yapildi", "p1": guncel_arac.get("plaka", ""), "p2": yeni_yag})

                eski_genel = mevcut.get("genel_bakim_tarihi", "")
                yeni_genel = guncel_arac.get("genel_bakim_tarihi", "")
                if eski_genel != yeni_genel and yeni_genel not in ["-", "Yok", ""]:
                    yeni_genel_dt = QDate.fromString(yeni_genel, "dd.MM.yyyy")
                    if yeni_genel_dt.isValid() and 0 <= yeni_genel_dt.daysTo(QDate.currentDate()) <= 7:
                        olaylar.append({"tarih": q_bugun, "tip": "genel_yapildi", "p1": guncel_arac.get("plaka", ""), "p2": yeni_genel})

                today = QDate.currentDate()
                olaylar = [o for o in olaylar if QDate.fromString(o.get("tarih", ""), "dd.MM.yyyy").isValid() and 0 <= QDate.fromString(o["tarih"], "dd.MM.yyyy").daysTo(today) <= 7]
                guncel_arac["olaylar"] = olaylar[-10:] # Keep max 10 events

                # Update in list
                for i, a in enumerate(self.araclar):
                    if a['id'] == arac_id:
                        for k, v in a.items():
                            if k not in guncel_arac:
                                guncel_arac[k] = v
                        self.araclar[i] = guncel_arac
                        break
                self.save_records()
                self.reload_arac_combo()
                
                # Re-select
                idx = self.combo_araclar.findData(arac_id)
                if idx >= 0:
                    self.combo_araclar.setCurrentIndex(idx)

    def arac_sil(self):
        arac_id = self.combo_araclar.currentData()
        if not arac_id:
            QMessageBox.warning(self, _t("msg_uyar", "Uyarı"), _t("msg_ltfensilmeki", "Lütfen silmek için bir araç seçin."))
            return
            
        arac_isim = self.combo_araclar.currentText()
        msg = QMessageBox(self)
        msg.setWindowTitle(_t("arac_sil_baslik", "Araç Sil"))
        msg.setText(f"'{arac_isim}' {_t('msg_arac_sil_emin', 'sistemden tamamen silinecek. Emin misiniz?')}")
        btn_evet = msg.addButton(_t("evet_sil", "Evet, Sil"), QMessageBox.ButtonRole.YesRole)
        msg.addButton(_t("hayir_iptal", "Hayır, İptal"), QMessageBox.ButtonRole.NoRole)
        msg.exec()
        
        if msg.clickedButton() == btn_evet:
            self.araclar = [a for a in self.araclar if a['id'] != arac_id]
            self.save_records()
            self.reload_arac_combo()

    def _get_matrah(self) -> float:
        text = self.matrah_input.text().replace(',', '.')
        try:
            return float(text) if text.strip() else 0.0
        except ValueError:
            return 0.0

    def _get_kdv(self):
        txt = self.kdv_input.currentText().replace('%', '').strip()
        try:
            return float(txt)
        except ValueError:
            return 0.0

    def update_live_preview(self):
        matrah = self._get_matrah()
        kdv_orani = self._get_kdv()
        tevkifat_payi = self.tevkifat_input.currentData()

        if tevkifat_payi is None: tevkifat_payi = 0
            
        kdv_tutari, tevkifat_tutari, odenecek_kdv, toplam = hesapla_tutarlar(matrah, kdv_orani, tevkifat_payi)
        
        kdv_lbl = _t('kdv_lbl', "KDV:")
        tev_lbl = _t('tev_lbl', "Tevkifat:")
        toplam_lbl = _t('toplam_lbl', "Toplam:")
        preview_text = f"{kdv_lbl} {format_currency(kdv_tutari)}"
        if tevkifat_payi > 0:
            preview_text += f" | {tev_lbl} {format_currency(tevkifat_tutari)}"
        preview_text += f" | {toplam_lbl} {format_currency(toplam)}"
        
        self.live_preview_label.setText(preview_text)

    def add_record(self):
        fatura_no = self.fatura_no_input.text().strip()
        firma = self.firma_input.text().strip()

        if not fatura_no or not firma:
            QMessageBox.warning(self, _t("msg_hata", "Hata"), _t("msg_faturanovefi", "Fatura No ve Firma alanı zorunludur."))
            return

        matrah = self._get_matrah()
        kdv_orani = self._get_kdv()
        tevkifat_payi = self.tevkifat_input.currentData()
        tevkifat_etiket = self.tevkifat_input.currentText()

        kdv_tutari, tevkifat_tutari, odenecek_kdv, toplam = hesapla_tutarlar(matrah, kdv_orani, tevkifat_payi)

        record = {
            "Tarih": self.date_input.date().toString("dd.MM.yyyy"),
            "Tip": self.tip_input.currentText(),
            "Fatura No": fatura_no,
            "Firma": firma,
            "Aciklama": self.aciklama_input.text().strip(),
            "Matrah": round(matrah, 2),
            "KDV %": kdv_orani,
            "KDV Tutari": kdv_tutari,
            "Tevkifat": tevkifat_etiket,
            "Tev.Tutarı": tevkifat_tutari,
            "Toplam": toplam
        }

        if self.editing_record_id:
            for r in self.records:
                if r.get("id") == self.editing_record_id:
                    record["id"] = r["id"]
                    r.update(record)
                    break
            self.cancel_edit()
        else:
            record["id"] = str(uuid.uuid4())
            self.records.append(record)
            
            # Clear inputs
            self.fatura_no_input.clear()
            self.firma_input.clear()
            self.aciklama_input.clear()
            self.matrah_input.clear()
            self.date_input.setDate(QDate.currentDate())

        self.save_records()
        self.update_ui()

    def cancel_edit(self):
        self.editing_record_id = None
        self.btn_ekle.setText(_t("faturayi_kaydet", "Faturayı Kaydet"))
        palette = get_theme_palette(app_config.get("theme", "dark"))
        self.btn_ekle.setStyleSheet(build_button_style(palette["primary_accent"], palette["white"], palette["primary_hover"], padding="12px", font_size="14px"))
        self.btn_iptal.hide()
        
        self.fatura_no_input.clear()
        self.firma_input.clear()
        self.aciklama_input.clear()
        self.matrah_input.clear()
        self.date_input.setDate(QDate.currentDate())

    def populate_table(self, table: QTableWidget, filter_type: str = None):
        data = self.records
        if filter_type:
            data = [r for r in data if r["Tip"] == filter_type]

        table.setRowCount(len(data))
        for row_idx, r in enumerate(data):
            items = [
                QTableWidgetItem(r["Tarih"]),
                QTableWidgetItem(r["Tip"]),
                QTableWidgetItem(r["Fatura No"]),
                QTableWidgetItem(r["Firma"]),
                QTableWidgetItem(r.get("Aciklama", "")),
                QTableWidgetItem(format_currency(r["Matrah"])),
                QTableWidgetItem(f"% {r['KDV %']}"),
                QTableWidgetItem(format_currency(r.get("KDV Tutari", 0.0))),
                QTableWidgetItem(r.get("Tevkifat", _t("yok_tevkifat", "Yok (0/10)"))),
                QTableWidgetItem(format_currency(r.get("Tev.Tutarı", 0.0))),
                QTableWidgetItem(format_currency(r["Toplam"]))
            ]
            items[0].setData(Qt.ItemDataRole.UserRole, r.get("id"))
            items[10].setFont(get_custom_font("Quicksand", 10, QFont.Weight.DemiBold))
            
            for col_idx, item in enumerate(items):
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(row_idx, col_idx, item)

    def _shift_month(self, year: int, month: int, delta: int):
        total = year * 12 + (month - 1) + delta
        return total // 12, (total % 12) + 1

    def _iter_months(self, start_key, end_key):
        y, m = start_key
        while (y, m) <= end_key:
            yield y, m
            y, m = self._shift_month(y, m, 1)

    def _format_month_label(self, year: int, month: int):
        month_map = ["Oca", "Sub", "Mar", "Nis", "May", "Haz", "Tem", "Agu", "Eyl", "Eki", "Kas", "Ara"]
        return f"{month_map[month - 1]} {str(year)[-2:]}"

    def _render_dashboard_chart(self, monthly_series):
        if not MATPLOTLIB_AVAILABLE or not hasattr(self, "dash_trend_figure") or not hasattr(self, "dash_trend_canvas") or not hasattr(self, "dash_pie_figure") or not hasattr(self, "dash_pie_canvas"):
            return

        palette = get_theme_palette(app_config.get("theme", "dark"))
        chart_bg = palette["surface_bg"]
        text_color = palette["text"]
        grid_color = palette["gridline"]

        labels = [x["label"] for x in monthly_series]
        gelir_data = [x["gelir"] for x in monthly_series]
        gider_data = [x["gider"] for x in monthly_series]
        net_data = [x["net"] for x in monthly_series]
        toplam_gelir = sum(gelir_data)
        toplam_gider = sum(gider_data)

        self.dash_trend_figure.clear()
        ax = self.dash_trend_figure.add_subplot(111)
        self.dash_trend_figure.patch.set_facecolor(chart_bg)
        ax.set_facecolor(chart_bg)

        self.dash_pie_figure.clear()
        ax_pie = self.dash_pie_figure.add_subplot(111)
        self.dash_pie_figure.patch.set_facecolor(chart_bg)
        ax_pie.set_facecolor(chart_bg)

        if not labels:
            ax.text(0.5, 0.5, _t("dash_chart_no_data", "Grafik için veri bulunamadı."), color=text_color, ha="center", va="center")
            ax.set_xticks([])
            ax.set_yticks([])
            self.dash_trend_figure.tight_layout()
            self.dash_trend_canvas.draw_idle()

            ax_pie.text(0.5, 0.5, _t("dash_chart_no_data", "Grafik için veri bulunamadı."), color=text_color, ha="center", va="center")
            ax_pie.set_xticks([])
            ax_pie.set_yticks([])
            self.dash_pie_figure.tight_layout()
            self.dash_pie_canvas.draw_idle()
            return

        x = list(range(len(labels)))
        ax.plot(x, gelir_data, color="#16a34a", linewidth=2.8, marker="o", markersize=5, label=_t("dash_income_outgoing", "Giden (Gelir)"))
        ax.plot(x, gider_data, color="#dc2626", linewidth=2.8, marker="o", markersize=5, label=_t("dash_expense_incoming", "Gelen (Gider)"))
        ax.plot(x, net_data, color="#2563eb", linewidth=2.0, linestyle="--", marker=".", markersize=3, label=_t("dash_net", "Net"))
        ax.fill_between(x, gelir_data, gider_data, color="#94a3b8", alpha=0.08)
        ax.set_title(_t("dash_trend_title", "Aylık Trend"), color=text_color, fontsize=13, fontweight="bold", pad=10)

        ax.set_xticks(x)
        ax.set_xticklabels(labels, rotation=0, fontsize=9, color=text_color)
        ax.tick_params(axis="y", labelsize=9, colors=text_color)
        ax.tick_params(axis="x", colors=text_color)
        ax.grid(axis="y", linestyle="--", alpha=0.35, color=grid_color)
        ax.set_ylabel(_t("dash_amount_tl", "Tutar (TL)"), color=text_color, fontsize=9)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_color(grid_color)
        ax.spines["bottom"].set_color(grid_color)
        ax.legend(frameon=False, fontsize=9, loc="upper left", ncol=2)
        self.dash_trend_figure.tight_layout()
        self.dash_trend_canvas.draw_idle()

        if (toplam_gelir + toplam_gider) > 0:
            pie_labels = [_t("dash_income_outgoing", "Giden (Gelir)"), _t("dash_expense_incoming", "Gelen (Gider)")]
            pie_values = [toplam_gelir, toplam_gider]
            pie_colors = ["#16a34a", "#dc2626"]
            _, _, autotexts = ax_pie.pie(
                pie_values,
                labels=pie_labels,
                colors=pie_colors,
                autopct="%1.1f%%",
                startangle=90,
                textprops={"color": text_color, "fontsize": 8},
                wedgeprops={"linewidth": 1, "edgecolor": chart_bg, "width": 0.42},
            )
            for t in autotexts:
                t.set_color(palette["white"])
                t.set_fontsize(8)
            ax_pie.set_title(_t("dash_pie_title", "Dönem Dağılımı"), color=text_color, fontsize=13, fontweight="bold", pad=10)
            ax_pie.text(0, 0, f"{toplam_gelir - toplam_gider:,.0f}\nTL", ha="center", va="center", color=text_color, fontsize=9, fontweight="bold")
        else:
            ax_pie.text(0.5, 0.5, _t("dash_chart_no_data", "Grafik için veri bulunamadı."), color=text_color, ha="center", va="center")
            ax_pie.set_xticks([])
            ax_pie.set_yticks([])

        self.dash_pie_figure.tight_layout()
        self.dash_pie_canvas.draw_idle()

    def update_dashboard(self):
        try:
            from PyQt6.QtWidgets import QTableWidgetItem
            from PyQt6.QtCore import Qt
            from datetime import datetime

            def pick_value(d, *keys, default=""):
                for k in keys:
                    if k in d and d.get(k) not in (None, ""):
                        return d.get(k)
                return default

            def parse_date_any(value):
                if not value or str(value).strip() in ("-", "Yok", "Var"):
                    return None
                s = str(value).strip()
                for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y"):
                    try:
                        return datetime.strptime(s, fmt)
                    except Exception:
                        pass
                return None

            def to_float_any(value):
                if isinstance(value, (int, float)):
                    return float(value)
                s = str(value or "").strip()
                if not s:
                    return 0.0
                s = s.replace("TL", "").replace("tl", "").replace(" ", "")
                if "," in s and "." in s:
                    if s.rfind(",") > s.rfind("."):
                        s = s.replace(".", "").replace(",", ".")
                    else:
                        s = s.replace(",", "")
                elif "," in s:
                    s = s.replace(".", "").replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return 0.0

            def normalize_tip(tip_value):
                raw = str(tip_value or "").strip().lower()
                return unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode("ascii")

            def is_expense_tip(tip_value):
                tip_norm = normalize_tip(tip_value)
                return any(k in tip_norm for k in ("gelen", "gider", "incoming", "expense", "alis"))

            def is_income_tip(tip_value):
                tip_norm = normalize_tip(tip_value)
                return any(k in tip_norm for k in ("giden", "gelir", "outgoing", "income", "revenue", "satis"))

            dash_theme = get_dashboard_theme(app_config.get("theme", "dark"))
            warning_color = QColor(dash_theme["warning_text"])
            critical_color = QColor(dash_theme["critical_text"])
            positive_color = QColor(dash_theme["positive_text"])
            
            toplam_arac = len(self.araclar)
            if hasattr(self, 'kpi_arac_sayisi'):
                self.kpi_arac_sayisi.setText(f"{toplam_arac} Adet")

            now = datetime.now()
            bu_ay_gelir = 0.0
            bu_ay_gider = 0.0
            tum_faturalar_olaylar = []
            aylik_toplamlar = {}
            
            for r in self.records:
                try:
                    fd = parse_date_any(pick_value(r, "tarih", "Tarih"))
                    if not fd:
                        continue

                    tip_val = str(pick_value(r, "tip", "Tip", default="")).strip()
                    toplam_val = to_float_any(pick_value(r, "toplam", "Toplam", default=0) or 0)
                    ilgili_val = pick_value(r, "fatura_no", "Fatura No", default="")
                    firma_val = pick_value(r, "firma", "Firma", default="")

                    tum_faturalar_olaylar.append({
                        "tarih": fd, "ilgili": ilgili_val,
                        "aciklama": f"Fatura: {firma_val}",
                        "tutar": toplam_val, "tip": tip_val, "isFatura": True
                    })

                    ay_key = (fd.year, fd.month)
                    if ay_key not in aylik_toplamlar:
                        aylik_toplamlar[ay_key] = {"gelir": 0.0, "gider": 0.0}
                    if is_expense_tip(tip_val):
                        aylik_toplamlar[ay_key]["gider"] += toplam_val
                    elif is_income_tip(tip_val):
                        aylik_toplamlar[ay_key]["gelir"] += toplam_val
                    else:
                        # Backward-compatible fallback: unknown type is treated as expense.
                        aylik_toplamlar[ay_key]["gider"] += toplam_val

                    if fd.month == now.month and fd.year == now.year:
                        if is_expense_tip(tip_val):
                            bu_ay_gider += toplam_val
                        elif is_income_tip(tip_val):
                            bu_ay_gelir += toplam_val
                        else:
                            bu_ay_gider += toplam_val
                except: pass

            if hasattr(self, 'kpi_gelir_aylik'):
                self.kpi_gelir_aylik.setText(f"{bu_ay_gelir:,.2f} TL")
                self.kpi_gider_aylik.setText(f"{bu_ay_gider:,.2f} TL")

            if hasattr(self, "combo_dash_period"):
                secim = self.combo_dash_period.currentData()
            else:
                secim = "12m"

            now_key = (now.year, now.month)
            if secim == "6m":
                start_key = self._shift_month(now.year, now.month, -5)
                end_key = now_key
            elif secim == "ytd":
                start_key = (now.year, 1)
                end_key = now_key
            elif secim == "all":
                if aylik_toplamlar:
                    start_key = min(aylik_toplamlar.keys())
                    end_key = max(aylik_toplamlar.keys())
                else:
                    start_key = now_key
                    end_key = now_key
            else:
                start_key = self._shift_month(now.year, now.month, -11)
                end_key = now_key

            aylik_seri = []
            for y, m in self._iter_months(start_key, end_key):
                vals = aylik_toplamlar.get((y, m), {"gelir": 0.0, "gider": 0.0})
                gelir = vals["gelir"]
                gider = vals["gider"]
                aylik_seri.append({
                    "label": self._format_month_label(y, m),
                    "gelir": gelir,
                    "gider": gider,
                    "net": gelir - gider,
                })
            self._render_dashboard_chart(aylik_seri)

            yaklasan_uyarilar = []
            kritik_count = 0
            gecen_count = 0
            
            for a in self.araclar:
                plaka = a.get("plaka", "")
                for belge_key, belge_adi in [
                    (("muayene_tarihi",), _t("doc_inspection", "Muayene")),
                    (("trafik_sigortasi_tarihi", "trafik_tarihi"), _t("doc_traffic_insurance", "Sigorta")),
                    (("kasko_tarihi",), _t("doc_casco", "Kasko")),
                    (("koltuk_sigorta_tarihi", "koltuk_sigortasi"), _t("doc_seat_insurance", "Koltuk Sigortası")),
                ]:
                    tarih = ""
                    belgeler = a.get("belgeler", {}) if isinstance(a.get("belgeler", {}), dict) else {}
                    for k in belge_key:
                        tarih = belgeler.get(k, "") or a.get(k, "")
                        if tarih:
                            break
                    if tarih:
                        try:
                            td = parse_date_any(tarih)
                            if not td:
                                continue
                            kalan = (td.date() - now.date()).days
                            if kalan <= 30:
                                yaklasan_uyarilar.append((plaka, belge_adi, kalan))
                                if kalan < 0:
                                    gecen_count += 1
                                    kritik_count += 1
                                elif kalan <= 15:
                                    kritik_count += 1
                        except: pass
                
                for olay in a.get("olaylar", []):
                    try:
                        od = parse_date_any(olay.get('tarih', ''))
                        if not od:
                            continue
                        aciklama = (olay.get("mesaj", "") or "").strip()
                        if not aciklama:
                            tip_txt = (olay.get("tip", "") or "").replace("_", " ").strip()
                            p1 = (olay.get("p1", "") or "").strip()
                            p2 = (olay.get("p2", "") or "").strip()
                            aciklama = " ".join([x for x in [tip_txt, p1, p2] if x]).strip() or "Arac olayi"
                        tum_faturalar_olaylar.append({
                            "tarih": od, "ilgili": plaka, "aciklama": aciklama,
                            "tutar": to_float_any(olay.get("tutar", 0)), "tip": "Gider", "isFatura": False
                        })
                    except: pass
            
            tum_faturalar_olaylar.sort(key=lambda x: x["tarih"], reverse=True)
            
            if hasattr(self, 'kpi_uyarilar'):
                self.kpi_uyarilar.setText(f"{kritik_count} / {len(yaklasan_uyarilar)}")

            if hasattr(self, 'lbl_dash_info'):
                self.lbl_dash_info.setText(
                    f"{_t('dash_last_update', 'Son güncelleme')}: {now.strftime('%d.%m.%Y %H:%M')}   |   "
                    f"{_t('dash_total_invoice', 'Toplam fatura')}: {len(self.records)}   |   "
                    f"{_t('dash_last_event', 'Son olay')}: {len(tum_faturalar_olaylar)}"
                )

            if hasattr(self, 'lbl_alert_count'):
                self.lbl_alert_count.setText(
                    f"{_t('dash_tracking_summary', 'Yaklaşan/Geçen Takip')}: {len(yaklasan_uyarilar)}  "
                    f"({_t('dash_critical', 'Kritik')}: {kritik_count}, {_t('dash_overdue', 'Geçen')}: {gecen_count})"
                )
            if hasattr(self, 'lbl_event_count'):
                self.lbl_event_count.setText(
                    f"{_t('dash_event_records', 'Hareket Kayıtları')}: {len(tum_faturalar_olaylar)}  "
                    f"({_t('dash_listed', 'Listelenen')}: {min(50, len(tum_faturalar_olaylar))})"
                )

            if hasattr(self, 'table_alerts'):
                self.table_alerts.setSortingEnabled(False)
                self.table_alerts.setRowCount(0)
                yaklasan_uyarilar.sort(key=lambda x: x[2])
                for r_idx, u in enumerate(yaklasan_uyarilar):
                    self.table_alerts.insertRow(r_idx)
                    
                    p_itm = QTableWidgetItem(u[0])
                    b_itm = QTableWidgetItem(u[1])
                    
                    k_itm = QTableWidgetItem(f"{u[2]} {_t('day_short', 'gün')}")
                    if u[2] < 0:
                        k_itm.setForeground(critical_color)
                        k_itm.setText(f"{_t('status_overdue', 'GEÇTİ')} ({-u[2]} {_t('day_short', 'gün')})")
                    elif u[2] <= 15:
                        k_itm.setForeground(critical_color)
                    else:
                        k_itm.setForeground(warning_color)
                    k_font = k_itm.font()
                    k_font.setBold(True)
                    k_itm.setFont(k_font)
                        
                    if u[2] < 0:
                        oncelik_itm = QTableWidgetItem(_t("status_very_urgent", "Çok Acil"))
                        oncelik_itm.setForeground(critical_color)
                    elif u[2] <= 15:
                        oncelik_itm = QTableWidgetItem(_t("status_urgent", "Acil"))
                        oncelik_itm.setForeground(critical_color)
                    else:
                        oncelik_itm = QTableWidgetItem(_t("status_approaching", "Yaklaşıyor"))
                        oncelik_itm.setForeground(warning_color)
                    p_font = oncelik_itm.font()
                    p_font.setBold(True)
                    oncelik_itm.setFont(p_font)
                    
                    self.table_alerts.setItem(r_idx, 0, p_itm)
                    self.table_alerts.setItem(r_idx, 1, b_itm)
                    self.table_alerts.setItem(r_idx, 2, k_itm)
                    self.table_alerts.setItem(r_idx, 3, oncelik_itm)
                self.table_alerts.setSortingEnabled(True)

            if hasattr(self, 'table_events'):
                self.table_events.setSortingEnabled(False)
                self.table_events.setRowCount(0)
                for r_idx, ev in enumerate(tum_faturalar_olaylar[:50]):
                    self.table_events.insertRow(r_idx)
                    try:
                        t_str = ev["tarih"].strftime("%d.%m.%Y %H:%M") if not ev["isFatura"] else ev["tarih"].strftime("%d.%m.%Y")
                    except:
                        t_str = ""
                    
                    t_itm = QTableWidgetItem(t_str)
                    i_itm = QTableWidgetItem(ev["ilgili"])
                    o_itm = QTableWidgetItem(ev["aciklama"])
                    
                    tutar_val = float(ev["tutar"])
                    tu_itm = QTableWidgetItem(f"{tutar_val:,.2f} TL" if tutar_val else "-")
                    
                    if tutar_val > 0:
                        if is_expense_tip(ev.get("tip", "")):
                            tu_itm.setForeground(critical_color)
                            tu_itm.setText(f"- {tutar_val:,.2f} TL")
                        elif is_income_tip(ev.get("tip", "")):
                            tu_itm.setForeground(positive_color)
                            tu_itm.setText(f"+ {tutar_val:,.2f} TL")
                            
                    self.table_events.setItem(r_idx, 0, t_itm)
                    self.table_events.setItem(r_idx, 1, i_itm)
                    self.table_events.setItem(r_idx, 2, o_itm)
                    self.table_events.setItem(r_idx, 3, tu_itm)
                self.table_events.setSortingEnabled(True)
                    
        except Exception as e:
            print(f"Dashboard update error: {e}")

    def update_ui(self):
        df = records_to_df(self.records)
        ozet = hesapla_ozet(df)

        self.lbl_gelen_toplam.setText(f"{_t('ozet_gelen_fatura', 'Toplam Gelen Fatura (Alış / Gider):')} {format_currency(ozet['gelen_toplam'])}")
        self.lbl_giden_toplam.setText(f"{_t('ozet_giden_fatura', 'Toplam Giden Fatura (Satış / Gelir):')} {format_currency(ozet['giden_toplam'])}")
        self.update_dashboard()
        self.lbl_net_fark.setText(f"{_t('ozet_net_fark', 'Net Durum (Kâr / Zarar):')} {format_currency(ozet['net_fark'])}")
        self.lbl_kdv_fark.setText(f"{_t('ozet_kdv_farki', 'KDV (Ödenecek / Devreden):')} {format_currency(ozet['kdv_fark'])}")

        palette = get_theme_palette(app_config.get("theme", "dark"))
        net_color = palette["negative_soft_text"] if ozet['net_fark'] < 0 else palette["positive_soft_text"]
        kdv_color = palette["negative_soft_text"] if ozet['kdv_fark'] < 0 else palette["positive_soft_text"]
        self.lbl_net_fark.setStyleSheet(f"color: {net_color}; font-weight: bold; background-color: transparent;")
        self.lbl_kdv_fark.setStyleSheet(f"color: {kdv_color}; font-weight: bold; background-color: transparent;")
        
        # Dipnot Guncelleme
        dipnot_title = _t('kdv_dipnot_title', "DİPNOT:")
        dipnot_text = _t('kdv_dipnot_text', "KDV farkı Yeşil (+) ise Maliyeye 'Ödenecek KDV'dir.<br>Kırmızı (-) ise bir sonraki aya 'Devreden KDV'dir.")
        self.lbl_kdv_dipnot.setText(
            f"<b style=' font-size: 12px;'>{dipnot_title}</b> "
            f"<span style=' font-size: 11px; font-style: italic;'>{dipnot_text}</span>"
        )

        self.populate_table(self.table_all)
        self.populate_table(self.table_gelen, "Gelen")
        self.populate_table(self.table_giden, "Giden")

    def export_arac_excel(self):
        if not self.araclar:
            QMessageBox.information(self, _t("msg_bilgi", "Bilgi"), _t("msg_daaktarlaca", "Dışa aktarılacak araç kaydı yok."))
            return

        file_name, _ = QFileDialog.getSaveFileName(self, "Excel Olarak Kaydet", f"arac_kayitlari_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx", "Excel Files (*.xlsx)")
        if file_name:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = "Araç Kayıtları"

                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="475569")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                headers = [_t("hr_plaka", "Plaka"), _t("hr_marka", "Marka"), _t("hr_model", "Model"), _t("hr_yil", "Yıl"), _t("hr_yakit", "Yakıt"), _t("hr_vites", "Vites"), _t("hr_km", "Güncel KM"), _t("hr_sasi", "Şasi No"), _t("hr_utts", "UTTS Durumu"), _t("hr_muayene", "Muayene Bitiş"), _t("hr_trafik", "Trafik Bitiş"), _t("hr_kasko", "Kasko Bitiş"), _t("hr_koltuk", "Koltuk Sigortası"), _t("hr_ruhsat", "Ruhsat No")]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border

                for row_num, arac in enumerate(self.araclar, 2):
                    row_data = [
                        arac.get("plaka", ""), arac.get("marka", ""), arac.get("model", ""), arac.get("yil", ""),
                        arac.get("yakit", ""), arac.get("vites", ""), arac.get("km", ""), arac.get("sasi", ""),
                        arac.get("utts", ""), arac.get("muayene_tarihi", ""), arac.get("trafik_tarihi", ""),
                        arac.get("kasko_tarihi", ""), arac.get("koltuk_sigortasi", ""), arac.get("ruhsat_no", "")
                    ]
                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.border = thin_border

                for col in ws.columns:
                    if not col: continue
                    max_length = 0
                    column = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[column].width = min(max_length + 2, 40)

                wb.save(file_name)
                QMessageBox.information(self, _t("msg_baarl", "Başarılı"), _t("msg_arabilgileri", "Araç bilgileri Excel olarak başarıyla kaydedildi."))
            except Exception as e:
                QMessageBox.warning(self, _t("hata_baslik", "Hata"), f"Excel kaydedilirken hata oluştu:\n{str(e)}")

    def export_arac_pdf(self):
        if not self.araclar:
            QMessageBox.information(self, _t("msg_bilgi", "Bilgi"), _t("msg_daaktarlaca", "Dışa aktarılacak araç kaydı yok."))
            return

        file_name, _ = QFileDialog.getSaveFileName(self, "PDF Olarak Kaydet", f"arac_kayitlari_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf", "PDF Files (*.pdf)")
        if file_name:
            try:
                from reportlab.lib.pagesizes import landscape, A4
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import mm
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
                from reportlab.lib.enums import TA_CENTER

                try:
                    pdfmetrics.registerFont(TTFont("Quicksand", str(APP_FONT_DIR / "Quicksand-Regular.ttf")))
                    pdfmetrics.registerFont(TTFont("Quicksand-Bold", str(APP_FONT_DIR / "Quicksand-Bold.ttf")))
                    font_normal = "Quicksand"
                    font_bold = "Quicksand-Bold"
                except:
                    font_normal = "Helvetica"
                    font_bold = "Helvetica-Bold"

                doc = SimpleDocTemplate(file_name, pagesize=landscape(A4), rightMargin=14*mm, leftMargin=14*mm, topMargin=14*mm, bottomMargin=14*mm)
                elements = []
                styles = getSampleStyleSheet()
                
                if "QuicksandTitle" not in styles:
                    styles.add(ParagraphStyle(name="QuicksandTitle", fontName=font_bold, fontSize=18, spaceAfter=10, alignment=TA_CENTER))
                if "QuicksandNormal" not in styles:
                    styles.add(ParagraphStyle(name="QuicksandNormal", fontName=font_normal, fontSize=10, spaceAfter=6))
                if "QuicksandHeading" not in styles:
                    styles.add(ParagraphStyle(name="QuicksandHeading", fontName=font_bold, fontSize=14, spaceAfter=8))
                if "TableCellCenter" not in styles:
                    styles.add(ParagraphStyle(name="TableCellCenter", fontName=font_normal, fontSize=7.5, alignment=TA_CENTER, leading=9))
                if "TableHeaderCenter" not in styles:
                    styles.add(ParagraphStyle(name="TableHeaderCenter", fontName=font_bold, fontSize=8, alignment=TA_CENTER, leading=10, textColor=colors.white))

                title = Paragraph("Araç Kayıtları Analiz Raporu", styles['QuicksandTitle'])
                elements.append(title)
                elements.append(Spacer(1, 8))
                elements.append(Paragraph(f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles["QuicksandNormal"]))
                elements.append(Spacer(1, 10))

                # --- SUMMARY TABLE ---
                firma_counts = {}
                for a in self.araclar:
                    f_ad = a.get("firma", "").strip() or _t("belirtilmemis", "Belirtilmemiş")
                    firma_counts[f_ad] = firma_counts.get(f_ad, 0) + 1
                
                summary_data = [
                    ["Genel Araç İstatistikleri", "Adet"],
                    ["Sistemde Kayıtlı Toplam Araç", str(len(self.araclar))]
                ]
                for k, v in firma_counts.items():
                    summary_data.append([f"{k} Bünyesindeki Araçlar", str(v)])
                report_palette = get_theme_palette("light")

                summary_table = Table(summary_data, colWidths=[110*mm, 55*mm])
                summary_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(report_palette["report_header_teal"])),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(report_palette["report_grid"])),
                    ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, -1), font_normal),
                    ("FONTNAME", (0, 0), (-1, 0), font_bold),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor(report_palette["report_row_bg"])),
                ]))
                elements.append(summary_table)
                elements.append(Spacer(1, 14))

                elements.append(Paragraph("Detaylı Araç Listesi", styles["QuicksandHeading"]))
                elements.append(Spacer(1, 4))

                headers = [
                    Paragraph(_t("pdf_calistigi_kurum", "Çalıştığı Kurum"), styles["TableHeaderCenter"]),
                    Paragraph(_t("pdf_firma", "Araç Sahibi (Firma)"), styles["TableHeaderCenter"]),
                    Paragraph(_t("pdf_plaka", "Plaka"), styles["TableHeaderCenter"]), 
                    Paragraph(_t("pdf_sofor_bilgisi", "Şoför Bilgisi"), styles["TableHeaderCenter"]),
                    Paragraph(_t("pdf_sofor_tel", "Şoför Tel"), styles["TableHeaderCenter"]),
                    Paragraph(_t("pdf_marka_model", "Marka Model"), styles["TableHeaderCenter"]), 
                    Paragraph(_t("pdf_model_yili", "Model Yılı"), styles["TableHeaderCenter"]), 
                    Paragraph(_t("pdf_muayene_t", "Muayene T."), styles["TableHeaderCenter"]), 
                    Paragraph(_t("pdf_trafik_s", "Trafik S."), styles["TableHeaderCenter"]),
                    Paragraph(_t("pdf_koltuk_s", "Koltuk S."), styles["TableHeaderCenter"]), 
                    Paragraph(_t("pdf_kasko_b", "Kasko B."), styles["TableHeaderCenter"]),  
                    Paragraph(_t("pdf_utts", "UTTS"), styles["TableHeaderCenter"])
                ]
                data = [headers]

                for a in self.araclar:
                    f_ad = f"{a.get('firma', '')}" if a.get('firma') else _t("belirtilmemis", "Belirtilmemiş")
                    if a.get('vergi_no'):
                        f_ad += f"<br/>{a.get('vergi_no')}"
                        
                    kurum_ad = a.get("kurum", "-")

                    data.append([
                        Paragraph(kurum_ad, styles["TableCellCenter"]),
                        Paragraph(f_ad, styles["TableCellCenter"]),
                        Paragraph(a.get("plaka", ""), styles["TableCellCenter"]),
                        Paragraph(a.get("sofor", "-"), styles["TableCellCenter"]),
                        Paragraph(a.get("sofor_tel", "-"), styles["TableCellCenter"]),
                        Paragraph(f"{a.get('marka', '')} {a.get('model', '')}", styles["TableCellCenter"]),
                        Paragraph(str(a.get("yil", "")), styles["TableCellCenter"]),
                        Paragraph(a.get("muayene_tarihi", ""), styles["TableCellCenter"]),
                        Paragraph(a.get("trafik_tarihi", ""), styles["TableCellCenter"]),
                        Paragraph(a.get("koltuk_sigortasi", "-"), styles["TableCellCenter"]),
                        Paragraph(a.get("kasko_tarihi", ""), styles["TableCellCenter"]),
                        Paragraph(a.get("utts", ""), styles["TableCellCenter"])
                    ])

                available_width = 269 * mm
                # 12 Sütun için orantılı alanlar (Tam 269mm'ye ayarlı)
                col_widths = [31*mm, 31*mm, 18*mm, 24*mm, 23*mm, 23*mm, 15*mm, 20*mm, 20*mm, 20*mm, 20*mm, 11*mm]
                total_w = sum(col_widths)
                col_widths = [w * (available_width / total_w) for w in col_widths]

                t = Table(data, colWidths=col_widths, repeatRows=1)
                
                table_style = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(report_palette['report_header_teal'])),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor(report_palette['group_border'])),
                ]
                
                # Zebra striping (satırları dönüşümlü renklendir)
                for i in range(1, len(data)):
                    bg_color = colors.white if i % 2 != 0 else colors.HexColor(report_palette['report_row_bg'])
                    table_style.append(('BACKGROUND', (0, i), (-1, i), bg_color))

                t.setStyle(TableStyle(table_style))

                elements.append(t)
                
                def add_pdf_footer(canvas, doc):
                    canvas.saveState()
                    canvas.setFont(font_normal, 8)
                    canvas.setFillColor(colors.HexColor(report_palette["report_muted"]))
                    current_year = datetime.now().year
                    year_str = f"2026 - {current_year}"
                    footer_text = f"© {_t('owner_name', 'Ümit Arik')} {year_str}"
                    canvas.drawString(14*mm, 8*mm, _t("pdf_vehicle_report", "Vehicle Management Report"))
                    canvas.drawCentredString(297*mm / 2.0, 8*mm, footer_text)
                    canvas.drawRightString(297*mm - 14*mm, 8*mm, f"{_t('pdf_page', 'Page')} {doc.page}")
                    canvas.restoreState()

                doc.build(elements, onFirstPage=add_pdf_footer, onLaterPages=add_pdf_footer)
                QMessageBox.information(self, _t("msg_baarl", "Başarılı"), _t("msg_arabilgileri", "Araç bilgileri PDF olarak başarıyla kaydedildi."))
            except Exception as e:
                QMessageBox.warning(self, _t("hata_baslik", "Hata"), f"PDF oluşturulurken hata oluştu:\n{str(e)}")

    def export_excel(self):
        if not self.records:
            QMessageBox.information(self, _t("msg_bilgi", "Bilgi"), _t("msg_daaktarlaca", "Dışa aktarılacak kayıt yok."))
            return

        file_name, _ = QFileDialog.getSaveFileName(self, "Excel Olarak Kaydet", f"fatura_kayitlari_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx", "Excel Files (*.xlsx)")
        if file_name:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                df = records_to_df(self.records)
                if 'id' in df.columns:
                    df = df.drop(columns=['id'])
                
                ozet = hesapla_ozet(df)
                
                # Split data
                df_gelen = df[df["Tip"] == "Gelen"].drop(columns=["Tip"])
                df_giden = df[df["Tip"] == "Giden"].drop(columns=["Tip"])

                wb = Workbook()
                ws = wb.active
                ws.title = "Fatura Raporu"
                
                # Styles
                title_font = Font(bold=True, color="FFFFFF", size=12)
                title_fill_gelen = PatternFill("solid", fgColor="dc2626")  # Kırmızı
                title_fill_giden = PatternFill("solid", fgColor="059669")  # Yeşil
                header_font = Font(bold=True)
                header_fill = PatternFill("solid", fgColor="e2e8f0")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                # Layout Config
                gelen_start_col = 1
                giden_start_col = len(df_gelen.columns) + 3 # Add a spacer column
                
                # TITLES
                ws.cell(row=1, column=gelen_start_col, value="GELEN FATURALAR").font = title_font
                ws.cell(row=1, column=gelen_start_col).fill = title_fill_gelen
                if len(df_gelen.columns) > 1:
                    ws.merge_cells(start_row=1, start_column=gelen_start_col, end_row=1, end_column=gelen_start_col + len(df_gelen.columns) - 1)
                ws.cell(row=1, column=gelen_start_col).alignment = Alignment(horizontal="center")
                
                ws.cell(row=1, column=giden_start_col, value="GİDEN FATURALAR").font = title_font
                ws.cell(row=1, column=giden_start_col).fill = title_fill_giden
                if len(df_giden.columns) > 1:
                    ws.merge_cells(start_row=1, start_column=giden_start_col, end_row=1, end_column=giden_start_col + len(df_giden.columns) - 1)
                ws.cell(row=1, column=giden_start_col).alignment = Alignment(horizontal="center")
                
                # HEADERS
                row_idx = 2
                for c_idx, col_name in enumerate(df_gelen.columns, start=gelen_start_col):
                    cell = ws.cell(row=row_idx, column=c_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    
                for c_idx, col_name in enumerate(df_giden.columns, start=giden_start_col):
                    cell = ws.cell(row=row_idx, column=c_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    
                # DATA - GELEN
                gelen_row = 3
                for _, row_data in df_gelen.iterrows():
                    for c_idx, val in enumerate(row_data, start=gelen_start_col):
                        cell = ws.cell(row=gelen_row, column=c_idx, value=val)
                        cell.border = thin_border
                    gelen_row += 1
                    
                # TOTALS - GELEN
                ws.cell(row=gelen_row, column=gelen_start_col, value="TOPLAM:").font = Font(bold=True)
                ws.cell(row=gelen_row, column=gelen_start_col + len(df_gelen.columns) - 1, value=format_currency(ozet['gelen_toplam'])).font = Font(bold=True)
                
                # DATA - GIDEN
                giden_row = 3
                for _, row_data in df_giden.iterrows():
                    for c_idx, val in enumerate(row_data, start=giden_start_col):
                        cell = ws.cell(row=giden_row, column=c_idx, value=val)
                        cell.border = thin_border
                    giden_row += 1
                
                # TOTALS - GIDEN
                ws.cell(row=giden_row, column=giden_start_col, value="TOPLAM:").font = Font(bold=True)
                ws.cell(row=giden_row, column=giden_start_col + len(df_giden.columns) - 1, value=format_currency(ozet['giden_toplam'])).font = Font(bold=True)
                
                # SUMMARY AREA (ÖZET BİLGİLER)
                summary_start_row = max(gelen_row, giden_row) + 3
                ws.cell(row=summary_start_row, column=1, value="ÖZET BİLGİLER VE KDV DURUMU").font = Font(bold=True, size=13, color="4c3a73")
                
                summary_data = [
                    ("Toplam Kestiğiniz Faturalar (Gelir):", format_currency(ozet['giden_toplam']), None),
                    ("Toplam Aldığınız Faturalar (Gider):", format_currency(ozet['gelen_toplam']), None),
                    ("Net Kâr / Zarar Durumu:", format_currency(ozet['net_fark']), ozet['net_fark']),
                    ("Müşteriden Alınan KDV (Satışlardan):", format_currency(ozet['giden_kdv']), None),
                    ("Firmalara Ödenen KDV (Alışlardan):", format_currency(ozet['gelen_kdv']), None),
                    ("KDV Farkı (Artı: Devlete Ödenecek, Eksi: Bir Sonraki Aya Devreder):", format_currency(ozet['kdv_fark']), ozet['kdv_fark'])
                ]
                
                s_row = summary_start_row + 1
                for lbl_text, str_val, raw_val in summary_data:
                    ws.cell(row=s_row, column=1, value=lbl_text).font = Font(bold=True)
                    
                    val_cell = ws.cell(row=s_row, column=2, value=str_val)
                    val_cell.font = Font(bold=True)
                    if raw_val is not None:
                        val_cell.font = Font(bold=True, color="059669" if raw_val >= 0 else "dc2626")
                        
                    s_row += 1

                from openpyxl.utils import get_column_letter
                for col in ws.columns:
                    if not col: continue
                    max_length = 0
                    column = get_column_letter(col[0].column)
                    for cell in col:
                        try: 
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[column].width = min(max_length + 2, 40) # cap width at 40

                wb.save(file_name)
                QMessageBox.information(self, _t("msg_baarl", "Başarılı"), _t("msg_excelraporuba", "Excel raporu başarıyla kaydedildi."))
            except Exception as e:
                QMessageBox.warning(self, _t("hata_baslik", "Hata"), f"Excel kaydedilirken hata oluştu:\n{str(e)}")

    def export_pdf(self):
        if not self.records:
            QMessageBox.information(self, _t("msg_bilgi", "Bilgi"), _t("msg_daaktarlaca", "Dışa aktarılacak kayıt yok."))
            return

        file_name, _ = QFileDialog.getSaveFileName(self, "PDF Olarak Kaydet", f"fatura_raporu_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf", "PDF Files (*.pdf)")
        if file_name:
            df = records_to_df(self.records)
            ozet = hesapla_ozet(df)
            pdf_bytes = build_pdf(df, ozet)
            with open(file_name, 'wb') as f:
                f.write(pdf_bytes)
            QMessageBox.information(self, _t("msg_baarl", "Başarılı"), _t("msg_pdfraporubaa", "PDF raporu başarıyla kaydedildi."))

    def clear_records(self):
        msg = QMessageBox(self)
        msg.setWindowTitle(_t("onay_baslik", "Onay"))
        msg.setText(_t("tum_kayitlari_sil_emin", "Tüm kayıtları silmek istediğinize emin misiniz?"))
        btn_evet = msg.addButton(_t("evet_temizle", "Evet, Temizle"), QMessageBox.ButtonRole.YesRole)
        btn_hayir = msg.addButton(_t("hayir_iptal", "Hayır, İptal"), QMessageBox.ButtonRole.NoRole)
        msg.exec()
        
        if msg.clickedButton() == btn_evet:
            self.records = []
            self.save_records()
            self.update_ui()

    def show_context_menu(self, pos, table: QTableWidget):
        item = table.itemAt(pos)
        if not item: return
        
        menu = QMenu()
        edit_action = QAction("Düzenle", self)
        delete_action = QAction("Sil", self)
        
        menu.addAction(edit_action)
        menu.addAction(delete_action)
        
        action = menu.exec(table.viewport().mapToGlobal(pos))
        
        row = item.row()
        record_id = table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        
        if action == edit_action:
            self.load_record_for_edit(record_id)
        elif action == delete_action:
            self.delete_record(record_id)

    def handle_item_click(self, item: QTableWidgetItem, table: QTableWidget):
        row = item.row()
        record_id = table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        self.load_record_for_edit(record_id)

    def load_record_for_edit(self, record_id):
        for r in self.records:
            if r.get("id") == record_id:
                self.editing_record_id = record_id
                
                self.date_input.setDate(QDate.fromString(r["Tarih"], "dd.MM.yyyy"))
                self.tip_input.setCurrentText(r["Tip"])
                self.fatura_no_input.setText(r["Fatura No"])
                self.firma_input.setText(r["Firma"])
                self.aciklama_input.setText(r.get("Aciklama", ""))
                self.matrah_input.setText(str(r["Matrah"]))
                self.kdv_input.setCurrentText(f"% {r['KDV %']}")
                self.tevkifat_input.setCurrentText(r.get("Tevkifat", _t("yok_tevkifat", "Yok (0/10)")))
                
                self.btn_ekle.setText(_t("faturayi_guncelle", "Faturayı Güncelle"))
                palette = get_theme_palette(app_config.get("theme", "dark"))
                self.btn_ekle.setStyleSheet(build_button_style(palette["success_bg"], palette["white"], palette["success_hover"], padding="12px", font_size="14px"))
                self.btn_iptal.show()
                break

    def delete_record(self, record_id):
        msg = QMessageBox(self)
        msg.setWindowTitle(_t("onay_baslik", "Onay"))
        msg.setText(_t("secili_faturayi_sil_emin", "Seçili faturayı silmek istediğinize emin misiniz?"))
        btn_evet = msg.addButton(_t("evet_sil", "Evet, Sil"), QMessageBox.ButtonRole.YesRole)
        btn_hayir = msg.addButton(_t("hayir_iptal", "Hayır, İptal"), QMessageBox.ButtonRole.NoRole)
        msg.exec()
        
        if msg.clickedButton() == btn_evet:
            self.records = [r for r in self.records if r.get("id") != record_id]
            if self.editing_record_id == record_id:
                self.cancel_edit()
            self.save_records()
            self.update_ui()

class SplashScreen(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(
            Qt.WindowType.SplashScreen |
            Qt.WindowType.FramelessWindowHint |
            Qt.WindowType.WindowStaysOnTopHint
        )
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)

        splash_size = QSize(800, 450)
        screen = QApplication.primaryScreen().geometry()
        self.setFixedSize(splash_size)
        self.move(
            screen.center().x() - splash_size.width() // 2,
            screen.center().y() - splash_size.height() // 2
        )

        self._lottie = None
        self._movie = None
        self._player = None
        self._audio = None
        self._main_window = None
        self._finished = False
        self._splash_available = False

        anim_dir = str(APP_ANIM_DIR)
        json_path = os.path.join(anim_dir, "GirisAnim.json")
        gif_path = os.path.join(anim_dir, "GirisAnim.gif")
        webm_path = os.path.join(anim_dir, "5Ehyk79dV61p1f5S0F.webm")
        html_path = os.path.join(anim_dir, "splash.html")

        if os.path.exists(json_path):
            with open(json_path, "r", encoding="utf-8") as f:
                json_data = f.read()
            html_content = f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><style>body {{ margin: 0; padding: 0; background-color: transparent; overflow: hidden; }} #lottie {{ width: 100vw; height: 100vh; }}</style><script src="lottie.min.js"></script></head>
<body><div id="lottie"></div><script>var animData = {json_data}; lottie.loadAnimation({{container: document.getElementById('lottie'), renderer: 'svg', loop: false, autoplay: true, animationData: animData}});</script></body>
</html>"""
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(html_content)

        if QLottieWidget is not None and os.path.exists(json_path):
            try:
                self._lottie = QLottieWidget(self)
                self._lottie.setStyleSheet("background: transparent;")
                self._lottie.setGeometry(0, 0, splash_size.width(), splash_size.height())
                source_url = QUrl.fromLocalFile(json_path)
                if hasattr(self._lottie, "setSource"):
                    self._lottie.setSource(source_url)
                elif hasattr(self._lottie, "load"):
                    self._lottie.load(source_url)
                if hasattr(self._lottie, "setLoops"):
                    self._lottie.setLoops(1)
                if hasattr(self._lottie, "play"):
                    self._lottie.play()
                elif hasattr(self._lottie, "start"):
                    self._lottie.start()
                self._splash_available = True
            except Exception:
                self._lottie = None

        if self._lottie is None and QWebEngineView is not None and os.path.exists(html_path):
            try:
                self._web = QWebEngineView(self)
                self._web.setStyleSheet("background: transparent;")
                self._web.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
                # Ensure the web view has a transparent background
                self._web.page().setBackgroundColor(Qt.GlobalColor.transparent)
                self._web.setGeometry(0, 0, splash_size.width(), splash_size.height())
                self._web.load(QUrl.fromLocalFile(html_path))
                self._web.show()
                self._splash_available = True
            except Exception:
                self._web = None
        else:
            self._web = None

        if self._lottie is None and os.path.exists(gif_path):
            try:
                self._movie_label = QLabel(self)
                self._movie_label.setStyleSheet("background: transparent;")
                self._movie_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                self._movie_label.setGeometry(0, 0, splash_size.width(), splash_size.height())

                self._movie = QMovie(gif_path)
                self._movie.setCacheMode(QMovie.CacheMode.CacheAll)
                self._movie.setScaledSize(splash_size)
                self._movie_label.setMovie(self._movie)
                self._movie.start()
                self._splash_available = True
            except Exception:
                self._movie = None

        # Eğer animasyon için Lottie veya WebEngine desteği yoksa,
        # sadece direkt ana pencereye geçilecek.

        if self._player is not None:
            self._player.mediaStatusChanged.connect(self._on_status)
            self._player.playbackStateChanged.connect(self._on_state)

        self._safety = QTimer()
        self._safety.setSingleShot(True)
        self._safety.timeout.connect(self._do_finish)
        random_delay = random.randint(4000, 9500)
        self._safety.start(random_delay)

    def _on_status(self, status):
        if status == QMediaPlayer.MediaStatus.EndOfMedia:
            self._do_finish()

    def _on_state(self, state):
        if state == QMediaPlayer.PlaybackState.StoppedState:
            self._do_finish()

    def _do_finish(self):
        if self._finished:
            return   # zaten tetiklendi
        self._finished = True
        self._safety.stop()
        if self._player is not None:
            self._player.stop()
        if self._movie is not None:
            self._movie.stop()
        if self._lottie is not None and hasattr(self._lottie, 'stop'):
            try:
                self._lottie.stop()
            except Exception:
                pass
        if hasattr(self, '_web') and self._web is not None:
            self._web.deleteLater()
            self._web = None
        if self._main_window:
            anim_out = QPropertyAnimation(self, b"windowOpacity")
            anim_out.setDuration(400)
            anim_out.setStartValue(1.0)
            anim_out.setEndValue(0.0)
            anim_out.setEasingCurve(QEasingCurve.Type.InCubic)
            anim_out.finished.connect(self.close)
            anim_out.finished.connect(self._main_window.showMaximized)
            if hasattr(self._main_window, 'check_license'):
                anim_out.finished.connect(self._main_window.check_license)
            anim_out.start()
            self._fadeout_anim = anim_out
        else:
            self._pending_show = True

    def finish(self, main_window):
        self._main_window = main_window
        if self._finished:
            main_window.showMaximized()
            if hasattr(main_window, 'check_license'):
                main_window.check_license()
            self.close()

if __name__ == "__main__":
    import signal
    signal.signal(signal.SIGINT, signal.SIG_DFL)

    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    
    from PyQt6.QtCore import QTranslator, QLibraryInfo
    qt_translator = QTranslator()
    qtbase_translator = QTranslator()
    
    if app_config.get("lang", "tr") == "tr":
        trans_path = QLibraryInfo.path(QLibraryInfo.LibraryPath.TranslationsPath)
        qt_translator.load("qt_tr", trans_path)
        qtbase_translator.load("qtbase_tr", trans_path)
        app.installTranslator(qt_translator)
        app.installTranslator(qtbase_translator)
    register_application_fonts()
    app_font = get_custom_font("Quicksand", 10, QFont.Weight.Medium)
    app_font.setStyleStrategy(QFont.StyleStrategy.PreferAntialias)
    app.setFont(app_font)
    window = FaturaApp()
    splash = None
    try:
        splash = SplashScreen()
        if splash._splash_available:
            splash.show()
            app.processEvents()
            splash.finish(window)
        else:
            print("Splash support not available on this system; opening main window directly.")
            splash.close()
            window.showMaximized()
    except Exception as exc:
        print("Splash initialization failed:", exc)
        if splash is not None:
            try:
                splash.close()
            except Exception:
                pass
        window.showMaximized()
    sys.exit(app.exec())